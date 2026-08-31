#!/usr/bin/env python3
"""Build a repo-ready, deduplicated snapshot of all assistant-generated artifacts.

The snapshot expands historical delivery bundles, deduplicates by SHA-256, keeps
all unique artifacts, records every original occurrence/alias, and produces a
classification/move plan for the user's current local directory listing.
"""
from __future__ import annotations

import csv
import hashlib
import json
import os
import posixpath
import re
import shutil
import sys
import textwrap
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import BinaryIO, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path('/mnt/data')
STAMP = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
SNAPSHOT_ROOT = ROOT / f'gpt56sol_complete_snapshot_{STAMP}'
GPT_ROOT = SNAPSHOT_ROOT / 'gpt56sol'
ZIP_PATH = ROOT / f'gpt56sol_complete_artifact_snapshot_{STAMP}.zip'

DIRECT_CURRENT = ROOT / 'ppi_repro_corrected'
DIRECT_INITIAL = ROOT / 'ppi_repro'
STANDALONE_FILES = [ROOT / 'historical_go_mapping_inventory.md']
BUNDLE_PATHS = sorted(ROOT.glob('B*.zip')) + [
    ROOT / 'ppi_reproduction_analysis_bundle.zip',
    ROOT / 'ppi_reproduction_corrected_bundle.zip',
    ROOT / 'ppi_batch_upload_protocol.zip',
]
BUNDLE_PATHS = [p for p in BUNDLE_PATHS if p.is_file()]

USER_LOCAL_LISTING = r'''
./2016-06-01-annotations-README
./2016-06-01-gene_ontology.obo
./2016-06-01-go.obo
./2016-06-01-go.owl
./2016-06-01-gp2protein.geneid.gz
./2016-06-01-gp2protein.human.gz
./2016-12-23-gene2go.gz
./2016-12-23-gene2go_human.tsv.gz
./2016_04
./2016_04/O95073_Q9Y620_2016_04.dat
./2016_04/O95073_Q9Y620_2016_04.tsv
./2016_04/O95073_Q9Y620_2016_04_provenance.json
./2016_04/RELEASE_2016_04.metalink
./2016_05
./2016_05/O95073_Q9Y620_2016_05.dat
./2016_05/O95073_Q9Y620_2016_05.tsv
./2016_05/O95073_Q9Y620_2016_05_provenance.json
./2016_05/RELEASE_2016_05.metalink
./2016_06
./2016_06/O95073_Q9Y620_2016_06.dat
./2016_06/O95073_Q9Y620_2016_06.tsv
./2016_06/O95073_Q9Y620_2016_06_provenance.json
./2016_06/RELEASE_2016_06.metalink
./2026-08-14-gene2go.gz
./2026-08-14-gene2go_human.tsv.gz
./B104D_GOA_date_screen_reference_pack_20260829T010311Z.zip
./bio-tissue-hierarchy.tar.gz
./bio-tissue-labels.tar.gz
./bio-tissue-networks.tar.gz
./bio-tissue-readme.txt
./bioconductor-annotation-org.Hs.eg.db_3.0.0.tar.gz
./bioconductor-annotation-org.Hs.eg.db_3.1.2.tar.gz
./bioconductor-annotation-org.Hs.eg.db_3.3.0.tar.gz
./bioconductor-annotation-org.Hs.eg.db_3.4.0.tar.gz
./biosnap_ppi_audit_results
./biosnap_ppi_audit_results/biosnap_ppi_provenance_audit_20260830T113352Z.json
./biosnap_ppi_audit_results/biosnap_ppi_provenance_audit_20260830T113352Z.zip
./biosnap_ppi_audit_results/README_20260830T113352Z.md
./blood_sample.tsv.gz
./chat_gpt_historical_go_mapping_inventory.md
./dgl_ppi.zip
./dhimmel-gene-ontology-962a5e1-GO_annotations-9606-direct-allev.tsv
./dhimmel-gene-ontology-962a5e1-GO_annotations-9606-direct-expev.tsv
./dhimmel-gene-ontology-962a5e1-GO_annotations-9606-inferred-allev.tsv
./dhimmel-gene-ontology-962a5e1-GO_annotations-9606-inferred-expev.tsv
./dhimmel_GO_annotations_962a5e1_manifest.csv
./download_and_verify_biosnap_ppi_sources.py
./download_extract_uniprot_2016_mapping_audit.py
./download_raw_dhimmel_annotations.py
./file_inventory_2026_08_18.pdf
./filter_gene2go.sh
./find_gene2go_2016.py
./gene2go_github_survey.tsv
./gene_association.goa_human.155.gz
./gene_association.goa_human.156.gz
./gene_association.goa_human.157.gz
./gene_association.goa_ref_human.157.gz
./goa_date_screen_cache
./goa_date_screen_results
./goa_date_screen_results/goa_date_screen_events.csv
./goa_date_screen_results/goa_date_screen_run_metadata.json
./goa_date_screen_results/goa_release_158_screen.json
./goa_date_screen_results/goa_release_159_screen.json
./goa_date_screen_results/goa_release_160_screen.json
./goa_date_screen_results/goa_release_161_screen.json
./goa_date_screen_results/goa_release_162_screen.json
./goa_date_screen_results/goa_release_163_screen.json
./goa_date_screen_results/goa_release_164_screen.json
./goa_date_screen_results/goa_release_165_screen.json
./goa_date_screen_results/goa_release_166_screen.json
./goa_date_screen_results/goa_release_167_screen.json
./goa_date_screen_results/goa_release_168_screen.json
./goa_date_screen_results/goa_release_169_screen.json
./goa_date_screen_results/goa_release_date_screen_summary_20260829T120614Z.csv
./goa_date_screen_results.zip
./goa_human.gaf.158.gz
./goa_human.gaf.159.gz
./goa_human.gaf.160.gz
./goa_human.gpa.158.gz
./goa_human.gpa.159.gz
./goa_human.gpa.160.gz
./goa_human.gpi.158.gz
./goa_human.gpi.159.gz
./goa_human.gpi.160.gz
./GO_annotations-9606-direct-allev.tsv
./GO_annotations-9606-direct-expev.tsv
./GO_annotations-9606-inferred-allev.tsv
./GO_annotations-9606-inferred-expev.tsv
./gp_association.goa_human.157.gz
./graphsage_ppi.zip
./GraphSAGE_PPI_GO_Label_Reconstruction_ChatGPT_2026_08_28.md
./Greene2015.pdf
./Greene2015_sup.pdf
./Greene2015_Table6.xlsx
./Greene2015_Table9.xlsx
./HuamnBase-kidney.dat
./HumanBase-blood.dat
./HumanBase-blood_top.gz
./idmapping_2026_08_27.tsv.gz
./investigation_summary_2026_08_23.md
./local_upload_inventory_20260827T145903Z.csv
./local_upload_inventory_full_20260827T160408Z.csv
./make_local_inventory.py
./make_local_inventory_v2.py
./msigdb_v5.0_files_to_download_locally.zip
./msigdb_v5.1_files_to_download_locally.zip
./msigdb_v5.2_chip_files_to_download_locally.zip
./msigdb_v5.2_files_to_download_locally.zip
./msigdb_v6.0_files_to_download_locally.zip
./ohmnet-master.zip
./OhmNet.pdf
./requested_file_patterns.txt
./sample_giant_network.sh
./screen_goa_release_date_range.py
./train_graph_id.npy
./uniprot_2016_mapping_audit_ledger.csv
./uniprot_2016_mapping_audit_ledger.zip
'''.strip().splitlines()

@dataclass
class Occurrence:
    kind: str  # direct or zip
    source: str
    member: str
    logical_group: str
    size: int = 0
    sha256: str = ''


def sha256_stream(f: BinaryIO) -> tuple[str, int]:
    h = hashlib.sha256()
    size = 0
    while True:
        block = f.read(1024 * 1024)
        if not block:
            break
        size += len(block)
        h.update(block)
    return h.hexdigest(), size


def file_sha256(path: Path) -> str:
    with path.open('rb') as f:
        return sha256_stream(f)[0]


def safe_member(name: str) -> str:
    name = name.replace('\\', '/')
    norm = posixpath.normpath(name).lstrip('/')
    if norm == '.' or norm.startswith('../') or '/..' in norm:
        raise ValueError(f'Unsafe zip member: {name}')
    return norm


def batch_from_text(text: str) -> str:
    m = re.search(r'\b(B\d{3}[A-Z]?)\b', text)
    if m:
        return m.group(1)
    if 'B000A' in text:
        return 'B000A'
    return ''


def category_for(path: str) -> str:
    p = path.lower()
    b = os.path.basename(p)
    if '/scripts/' in p or b.endswith('.py') or b.endswith('.sh'):
        return 'script'
    if 'source_ledger' in b or 'provenance_events' in b or 'actual_input_file_manifest' in b or 'input_manifest' in b:
        return 'provenance_manifest'
    if 'checksum' in b or 'sha256' in b or 'validation' in b or 'integrity' in b:
        return 'validation_or_checksum'
    if 'deletion_clearance' in b or 'receipt' in b or 'run_status' in b:
        return 'batch_administration'
    if b.endswith('.md') or b.endswith('.txt') or b.endswith('.pdf'):
        if 'report' in b or 'handoff' in b or 'readme' in b or 'instructions' in b or 'protocol' in b or 'backlog' in b or 'claim' in b or 'assessment' in b:
            return 'report_or_documentation'
        return 'documentation'
    if 'retained_inputs' in p or 'normalized' in b or 'compact_inputs' in p:
        return 'retained_normalized_input'
    if b.endswith(('.csv', '.csv.gz', '.tsv', '.tsv.gz', '.json', '.npy', '.npz')):
        return 'analysis_result_or_derived_data'
    if b.endswith('.xlsx'):
        return 'provenance_workbook'
    if b.endswith('.zip'):
        return 'reference_pack_or_nested_bundle'
    return 'other_artifact'


def description_for(path: str) -> str:
    p = path.lower()
    b = os.path.basename(path)
    batch = batch_from_text(path)
    prefix = f'{batch} ' if batch else ''
    patterns = [
        ('graphsage_ppi_node_to_entrez_evidence', 'Evidence-rich GraphSAGE row/node-to-Entrez mapping with topology, feature, label, and legacy-dictionary checks.'),
        ('graphsage_ppi_node_to_entrez', 'Compact GraphSAGE row/node-to-Entrez mapping for all 56,944 node rows.'),
        ('full_graphsage_row_to_entrez_mapping', 'Complete GraphSAGE row-to-Entrez mapping produced by the recovered legacy Python/NetworkX node order.'),
        ('full_4301_gene_universe', 'Deduplicated 4,301-gene universe derived from the complete row mapping.'),
        ('entrez_split_membership', 'One-row-per-Entrez summary of train/validation/test tissue-graph membership.'),
        ('evidence_vs_literature', 'Claim-by-claim register comparing file-derived evidence with OhmNet, GraphSAGE, and related literature.'),
        ('github_issues_tracker', 'Tracker of GraphSAGE GitHub issues relevant to dataset provenance, mappings, features, and evaluation.'),
        ('ppi_provenance', 'PPI topology provenance audit or report linking GraphSAGE, OhmNet, and BioSNAP edge sets.'),
        ('leakage_experiment_backlog', 'Saved design for the gene-identity lookup, randomized-label, and gene-disjoint leakage experiments.'),
        ('label_to_go_mapping', 'Per-column GraphSAGE label-to-Gene Ontology term mapping.'),
        ('final_exact_121_column_mapping', 'Exact 121-column GO candidate mapping after qualifier, evidence, propagation, and identifier-component corrections.'),
        ('inferred_unique_121_go_column_order', 'Strongly supported provisional disambiguation/order for the 121 GO label columns.'),
        ('direct_label_match_grid', 'Exhaustive grid of direct annotation-source, evidence, mapping, and label-column comparisons.'),
        ('direct_label_match_summary', 'Summary of direct GO-label matching configurations.'),
        ('nonexact_gene_differences', 'Gene-level mismatch witnesses from an intermediate, later superseded GO reconstruction.'),
        ('residual_false_positive', 'Witness rows for false positives from an intermediate GO reconstruction before qualifier/mapping corrections.'),
        ('mapping_components', 'Bipartite identifier-mapping components and ambiguity-resolution evidence.'),
        ('accession_geneid_mapping_edges', 'Complete retained UniProt accession-to-GeneID mapping edge table.'),
        ('gp2protein_mapping_summary', 'Summary statistics and edge cases for the historical gp2protein GeneID mapping.'),
        ('gpi159_uniprot_to_geneid_mapping', 'GPI159 UniProt objects projected to historical GeneIDs with ambiguity retained.'),
        ('goa_human_gaf159_normalized', 'Row-preserving normalized derivative of GOA human GAF release 159.'),
        ('goa_human_gpad159_normalized', 'Row-preserving normalized derivative of GOA human GPAD release 159.'),
        ('goa_human_gpi159_normalized', 'Row-preserving normalized derivative of GOA human GPI release 159.'),
        ('goa_human_gaf158_normalized', 'Row-preserving normalized derivative of GOA human GAF release 158.'),
        ('goa_human_gpad158_normalized', 'Row-preserving normalized derivative of GOA human GPAD release 158.'),
        ('goa_human_gpi158_normalized', 'Row-preserving normalized derivative of GOA human GPI release 158.'),
        ('global_human_goa_term_counts', 'Full-human propagated GO term prevalence counts used to reconstruct the 121-term selection rule.'),
        ('term_selection_policy_robustness', 'Policy grid testing evidence, qualifier, and ontology-propagation rules for term selection.'),
        ('msigdb_direct_membership_summary', 'Direct MSigDB-to-GraphSAGE label membership comparison summary.'),
        ('msigdb_feature_version_screen', 'Cross-version screen showing which MSigDB releases reproduce the 50 feature vectors.'),
        ('exact_msigdb52_feature_generation_rule', 'Column-level reconstruction of the 50 MSigDB feature sets and their ordering.'),
        ('feature_generation_validation', 'Machine-readable exact feature-matrix reconstruction validation.'),
        ('full_feature_validation', 'Independent full 56,944-row feature reconstruction validation.'),
        ('full_label_validation', 'Independent full 56,944-row GO-label reconstruction validation.'),
        ('row_order_validation', 'Independent validation of the recovered CPython 2/NetworkX row order.'),
        ('column_order_model', 'Comparison or summary of candidate legacy-Python GO column-order models.'),
        ('goa_release158_169', 'GOA release-date screen results for releases 158 through 169.'),
        ('uniprot_o95073_q9y620', 'Date-matched UniProt flat-file audit for the O95073/FSBP and Q9Y620/RAD54B mapping component.'),
        ('temporal_identifier_audit', 'Temporal audit of identifier mappings that changed or required contextual interpretation.'),
        ('identifier_mapping_watchlist', 'Watchlist of uncertain or time-varying Entrez-UniProt mappings.'),
        ('source_ledger', 'Append-only source and artifact provenance ledger.'),
        ('provenance_events', 'Append-only provenance-event history, including receipt, validation, and deletion states.'),
        ('actual_input_file_manifest', 'Manifest of actual materialized inputs used by the analysis at that point in time.'),
        ('inventory', 'File inventory or enriched local-holdings inventory.'),
        ('execution_diagnostics', 'Execution diagnostics, including accepted runs, failures, retries, and limitations.'),
        ('deletion_clearance', 'Batch-specific record authorizing deletion of raw conversation attachments after retention checks.'),
        ('validation', 'Machine-readable validation checks for the associated analysis or delivery.'),
        ('report', 'Narrative scientific and provenance report for the associated analysis batch.'),
        ('handoff', 'Investigator handoff summarizing methods, findings, caveats, and next tests.'),
        ('reproduction_instructions', 'Instructions for reproducing the associated analysis from retained inputs.'),
        ('download_extract_uniprot', 'Sequential low-storage downloader and extractor for date-matched UniProt Swiss-Prot records.'),
        ('screen_goa_release_date_range', 'Low-storage script screening GOA releases 158–169 while recording provenance.'),
        ('download_raw_dhimmel', 'Validated downloader for commit-pinned raw dhimmel Gene Ontology annotation TSVs.'),
        ('download_and_verify_biosnap', 'Downloader and exact edge-set audit for BioSNAP OhmNet and global PPI sources.'),
        ('make_local_inventory', 'Local file inventory generator with SHA-256 and datestamped output support.'),
        ('source_acquisition', 'Source acquisition notes and download instructions for historical inputs.'),
        ('requested_additional_inputs', 'Prioritized list of additional historical inputs requested during the investigation.'),
        ('batch_upload_protocol', 'Batch-upload, retention, provenance, and deletion-gate protocol.'),
    ]
    for needle, desc in patterns:
        if needle in p:
            return desc
    cat = category_for(path)
    if cat == 'script':
        return f'{prefix}analysis, download, validation, or packaging script: {b}.'
    if cat == 'report_or_documentation':
        return f'{prefix}report or documentation artifact: {b}.'
    if cat == 'provenance_manifest':
        return f'{prefix}provenance or input-manifest snapshot: {b}.'
    if cat == 'analysis_result_or_derived_data':
        return f'{prefix}analysis result or compact derived dataset: {b}.'
    if cat == 'retained_normalized_input':
        return f'{prefix}compact normalized retained input used for later reproducibility: {b}.'
    return f'{prefix}assistant-generated investigation artifact: {b}.'


def status_for(path: str) -> str:
    p = path.lower()
    if 'archive/initial_incomplete' in p:
        return 'superseded_initial_delivery'
    if '/b000/' in p or '/b101/' in p or 'b102_' in p or '/b104_2026' in p:
        return 'historical_foundation'
    if any(k in p for k in ['b104a_', 'b104b_', 'b104c_', 'b104d_', 'b104e_', 'b104f_']):
        return 'supporting_analysis'
    if any(k in p for k in ['b104g_', 'b104h_', 'b104i_']):
        return 'current_core'
    if 'through_b104i' in p or 'through_b104h' in p:
        return 'current_provenance'
    if '/results/' in p and any(k in p for k in ['master_reproduction_report', 'run_status', 'source_ledger.csv', 'actual_input_file_manifest.csv']):
        return 'superseded_or_early_summary'
    if 'deletion_clearance' in p:
        return 'administrative_history'
    return 'historical_or_supporting'


def git_tracking_for(size: int, path: str) -> str:
    p = path.lower()
    if size >= 5_000_000 or p.endswith(('.zip', '.gz', '.npy', '.npz')) and size >= 2_000_000:
        return 'git_lfs_recommended'
    return 'normal_git'


def bundle_rank(name: str) -> int:
    order = ['B104I', 'B104H', 'B104G', 'B104F', 'B104E', 'B104D', 'B104C', 'B104B', 'B104A', 'B104_', 'B102', 'B101', 'ppi_reproduction_corrected', 'ppi_reproduction_analysis', 'ppi_batch_upload_protocol']
    for idx, key in enumerate(order):
        if key in name:
            return len(order) - idx
    return 0


def normalize_user_path(line: str) -> str:
    s = line.strip()
    if s.startswith('./'):
        s = s[2:]
    return s.rstrip('/')


def classify_local(path: str, known_source: dict[str, dict[str, str]]) -> dict[str, str]:
    p = path.replace('\\', '/')
    b = os.path.basename(p)
    lower = p.lower()
    is_dir = '.' not in b or b in {'2016_04', '2016_05', '2016_06', 'biosnap_ppi_audit_results', 'goa_date_screen_cache', 'goa_date_screen_results'}
    origin = 'external_or_other'
    kind = 'directory' if is_dir else 'file'
    destination = ''
    action = 'move'
    commit = 'commit_metadata_only'
    description = ''
    notes = ''

    assistant_scripts = {
        'download_and_verify_biosnap_ppi_sources.py',
        'download_extract_uniprot_2016_mapping_audit.py',
        'download_raw_dhimmel_annotations.py',
        'make_local_inventory.py',
        'make_local_inventory_v2.py',
        'screen_goa_release_date_range.py',
    }
    assistant_docs = {
        'B104D_GOA_date_screen_reference_pack_20260829T010311Z.zip',
        'chat_gpt_historical_go_mapping_inventory.md',
        'GraphSAGE_PPI_GO_Label_Reconstruction_ChatGPT_2026_08_28.md',
        'requested_file_patterns.txt',
        'local_upload_inventory_20260827T145903Z.csv',
        'local_upload_inventory_full_20260827T160408Z.csv',
        'dhimmel_GO_annotations_962a5e1_manifest.csv',
        'uniprot_2016_mapping_audit_ledger.csv',
        'uniprot_2016_mapping_audit_ledger.zip',
        'goa_date_screen_results.zip',
    }
    other_agent = {
        'investigation_summary_2026_08_23.md',
        'file_inventory_2026_08_18.pdf',
        'filter_gene2go.sh',
        'find_gene2go_2016.py',
        'gene2go_github_survey.tsv',
        'sample_giant_network.sh',
        'train_graph_id.npy',
    }

    if lower.startswith(('2016_04/', '2016_05/', '2016_06/')) or p in {'2016_04', '2016_05', '2016_06'}:
        origin = 'gpt56sol_script_output'
        destination = f'gpt56sol/user_runs/uniprot_2016_audit/{p}'
        commit = 'normal_git'
        description = 'Small extracted UniProt audit records, summaries, metalinks, and provenance generated by the assistant-provided low-storage script.'
    elif lower.startswith('biosnap_ppi_audit_results'):
        origin = 'gpt56sol_script_output'
        suffix = p[len('biosnap_ppi_audit_results'):].lstrip('/')
        destination = 'gpt56sol/user_runs/biosnap_ppi_audit' + (f'/{suffix}' if suffix else '')
        commit = 'normal_git'
        description = 'BioSNAP PPI provenance-audit output generated by the assistant-provided script.'
    elif lower.startswith('goa_date_screen_results'):
        origin = 'gpt56sol_script_output'
        suffix = p[len('goa_date_screen_results'):].lstrip('/')
        destination = 'gpt56sol/user_runs/goa_date_screen' + (f'/{suffix}' if suffix else '')
        commit = 'normal_git'
        description = 'GOA release 158–169 date-screen result generated by the assistant-provided script.'
    elif lower == 'goa_date_screen_cache':
        origin = 'working_cache'
        destination = 'data/cache/goa_date_screen'
        action = 'move_or_delete_after_validation'
        commit = 'gitignore'
        description = 'Working download/cache directory; not a canonical result.'
    elif b in assistant_scripts:
        origin = 'gpt56sol_script'
        destination = f'gpt56sol/scripts/{b}'
        commit = 'normal_git'
        description = 'Assistant-provided analysis, download, audit, or inventory script.'
    elif b in assistant_docs:
        origin = 'gpt56sol_artifact_or_script_output'
        sub = 'reference_packs' if b.endswith('.zip') and b.startswith('B104D_') else 'user_runs'
        if b.startswith('local_upload_inventory'):
            sub = 'provenance/local_inventories'
        elif b == 'dhimmel_GO_annotations_962a5e1_manifest.csv':
            sub = 'provenance/dhimmel_download'
        elif b.startswith('uniprot_2016_mapping_audit'):
            sub = 'user_runs/uniprot_2016_audit'
        elif b.startswith('GraphSAGE_PPI_GO_Label_Reconstruction') or b.startswith('chat_gpt_'):
            sub = 'docs'
        destination = f'gpt56sol/{sub}/{b}'
        commit = 'git_lfs_recommended' if b.endswith('.zip') and known_source.get(b, {}).get('size_bytes', '0').isdigit() and int(known_source[b]['size_bytes']) > 5_000_000 else 'normal_git'
        description = 'Assistant-provided artifact or output from an assistant-provided script.'
    elif b in other_agent:
        origin = 'other_agent_or_prior_investigation'
        destination = f'other_agent/{b}'
        commit = 'normal_git' if not b.endswith('.npy') else 'git_lfs_recommended'
        description = 'Artifact from the other investigation agent or pre-existing local workflow, not generated by gpt56sol.'
    elif b.startswith('dhimmel-gene-ontology-962a5e1-'):
        origin = 'invalid_download_html_disguised_as_tsv'
        destination = f'quarantine/invalid_html/{b}'
        action = 'delete_or_quarantine'
        commit = 'do_not_commit'
        description = 'Rejected saved GitHub HTML page, not a scientific TSV; superseded by the genuine GO_annotations files.'
    elif b.startswith('GO_annotations-9606-'):
        origin = 'external_raw_data_downloaded_by_gpt56sol_script'
        destination = f'data/raw/dhimmel_gene_ontology_962a5e1/{b}'
        commit = 'gitignore_or_data_lfs'
        description = 'Genuine commit-pinned Entrez-native GO annotation TSV downloaded by the validated assistant script.'
    elif b.startswith(('2016-06-01-', '2016-12-23-', '2026-08-14-')) or b.startswith(('gene_association.', 'goa_human.', 'gp_association.')):
        origin = 'external_raw_data'
        if 'gene2go' in b:
            destination = f'data/raw/ncbi_gene/{b}'
        elif 'gp2protein' in b or 'go.' in b or 'ontology' in b or 'annotations-readme' in b:
            destination = f'data/raw/gene_ontology_release_2016_06_01/{b}'
        else:
            destination = f'data/raw/ebi_goa_human/{b}'
        commit = 'gitignore_or_data_lfs'
        description = 'Historical Gene Ontology, GOA, or NCBI Gene source file.'
    elif b.startswith('bioconductor-annotation-org.Hs.eg.db_'):
        origin = 'external_raw_data'
        destination = f'data/raw/bioconductor/{b}'
        commit = 'gitignore_or_data_lfs'
        description = 'Historical Bioconductor annotation package.'
    elif b.startswith('msigdb_'):
        origin = 'external_raw_data'
        destination = f'data/raw/msigdb/{b}'
        commit = 'gitignore_or_data_lfs'
        description = 'Historical MSigDB release archive.'
    elif b in {'graphsage_ppi.zip', 'dgl_ppi.zip'}:
        origin = 'external_raw_data'
        destination = f'data/raw/ppi_benchmarks/{b}'
        commit = 'gitignore_or_data_lfs'
        description = 'Released GraphSAGE or DGL PPI benchmark archive.'
    elif b.startswith('bio-tissue-') or b == 'ohmnet-master.zip':
        origin = 'external_raw_data_or_code'
        destination = f'data/raw/ohmnet/{b}'
        commit = 'gitignore_or_data_lfs'
        description = 'Released OhmNet tissue data, README, hierarchy, labels, networks, or source-code archive.'
    elif b in {'Greene2015.pdf', 'Greene2015_sup.pdf', 'Greene2015_Table6.xlsx', 'Greene2015_Table9.xlsx', 'OhmNet.pdf'}:
        origin = 'external_reference_document'
        destination = f'data/reference_papers/{b}'
        commit = 'gitignore_or_data_lfs'
        description = 'Manuscript or supplementary source used for documentary verification.'
    elif b.startswith(('HumanBase-', 'HuamnBase-')) or b == 'blood_sample.tsv.gz':
        origin = 'external_raw_or_derived_data'
        destination = f'data/raw/humanbase/{b}' if b != 'blood_sample.tsv.gz' else f'data/derived/humanbase/{b}'
        commit = 'gitignore_or_data_lfs'
        description = 'HumanBase/GIANT source or locally sampled derived data.'
        if b.startswith('HuamnBase-'):
            notes = 'Filename contains a likely typo; preserve until content/hash verification supports a rename.'
    elif b == 'idmapping_2026_08_27.tsv.gz':
        origin = 'external_current_mapping_output'
        destination = f'data/derived/uniprot_current/{b}'
        commit = 'gitignore_or_data_lfs'
        description = 'Current UniProt ID mapping result used only as temporal sensitivity evidence.'
    elif b == 'blood_sample.tsv.gz':
        origin = 'external_derived_data'
        destination = f'data/derived/humanbase/{b}'
        commit = 'gitignore_or_data_lfs'
    else:
        origin = 'unclassified_review_needed'
        destination = f'review/{p}'
        action = 'review_before_move'
        commit = 'review'
        description = 'Path was listed by the user but does not match a high-confidence classification rule.'

    src = known_source.get(b, {})
    return {
        'current_relative_path': p,
        'entry_kind': kind,
        'origin_class': origin,
        'recommended_destination': destination,
        'recommended_action': action,
        'git_tracking_recommendation': commit,
        'description': description,
        'notes': notes,
        'known_size_bytes': src.get('size_bytes', ''),
        'known_sha256': src.get('sha256', ''),
        'known_source_url': src.get('direct_or_canonical_source_url', ''),
        'classification_basis': 'filename/path rules; hashes only where previously recorded',
    }


def read_known_sources() -> dict[str, dict[str, str]]:
    mapping: dict[str, dict[str, str]] = {}
    candidates = [
        DIRECT_CURRENT / 'results' / 'source_ledger_through_B104I_FINAL_20260830T114918Z.csv',
        DIRECT_CURRENT / 'results' / 'actual_input_file_manifest_through_B104I_20260830T114918Z.csv',
    ]
    for path in candidates:
        if not path.is_file():
            continue
        with path.open(newline='', encoding='utf-8-sig') as f:
            for row in csv.DictReader(f):
                name = row.get('artifact_name', '')
                rel = row.get('user_local_relative_path', '')
                for key in {name, os.path.basename(rel), os.path.basename(row.get('local_path', ''))}:
                    if key:
                        mapping.setdefault(key, {}).update({k: row.get(k, '') for k in ['size_bytes', 'sha256', 'direct_or_canonical_source_url', 'source_page_url']})
    return mapping


def collect_occurrences() -> list[Occurrence]:
    out: list[Occurrence] = []
    if DIRECT_CURRENT.is_dir():
        for path in sorted(DIRECT_CURRENT.rglob('*')):
            if path.is_file():
                out.append(Occurrence('direct_current', str(path), path.relative_to(DIRECT_CURRENT).as_posix(), 'ppi_repro_corrected'))
    if DIRECT_INITIAL.is_dir():
        for path in sorted(DIRECT_INITIAL.rglob('*')):
            if path.is_file():
                out.append(Occurrence('direct_initial', str(path), path.relative_to(DIRECT_INITIAL).as_posix(), 'ppi_repro_initial'))
    for path in STANDALONE_FILES:
        if path.is_file():
            out.append(Occurrence('standalone', str(path), path.name, 'standalone'))
    for zpath in BUNDLE_PATHS:
        with zipfile.ZipFile(zpath) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                member = safe_member(info.filename)
                out.append(Occurrence('zip_member', str(zpath), member, zpath.name, size=info.file_size))
    return out


def hash_occurrences(occurrences: list[Occurrence]) -> None:
    direct_cache: dict[str, tuple[str, int]] = {}
    zip_cache: dict[str, zipfile.ZipFile] = {}
    try:
        for idx, occ in enumerate(occurrences, 1):
            if occ.kind in {'direct_current', 'direct_initial', 'standalone'}:
                if occ.source not in direct_cache:
                    p = Path(occ.source)
                    direct_cache[occ.source] = (file_sha256(p), p.stat().st_size)
                occ.sha256, occ.size = direct_cache[occ.source]
            else:
                zf = zip_cache.get(occ.source)
                if zf is None:
                    zf = zipfile.ZipFile(occ.source)
                    zip_cache[occ.source] = zf
                with zf.open(occ.member) as f:
                    occ.sha256, actual_size = sha256_stream(f)
                if actual_size != occ.size:
                    raise RuntimeError(f'ZIP member size mismatch: {occ.source}:{occ.member}')
    finally:
        for zf in zip_cache.values():
            zf.close()


def choose_occurrence(occurrences: list[Occurrence]) -> Occurrence:
    direct_current = [o for o in occurrences if o.kind == 'direct_current']
    if direct_current:
        return sorted(direct_current, key=lambda o: (len(o.member), o.member))[0]
    standalone = [o for o in occurrences if o.kind == 'standalone']
    if standalone:
        return standalone[0]
    direct_initial = [o for o in occurrences if o.kind == 'direct_initial']
    if direct_initial:
        return sorted(direct_initial, key=lambda o: (len(o.member), o.member))[0]
    return sorted(occurrences, key=lambda o: (-bundle_rank(o.logical_group), len(o.member), o.member))[0]


def destination_for(occ: Occurrence) -> Path:
    if occ.kind == 'direct_current':
        return GPT_ROOT / occ.member
    if occ.kind == 'direct_initial':
        return GPT_ROOT / 'archive' / 'initial_incomplete' / occ.member
    if occ.kind == 'standalone':
        return GPT_ROOT / 'docs' / occ.member
    bundle_stem = Path(occ.source).stem
    return GPT_ROOT / 'bundle_only' / bundle_stem / safe_member(occ.member)


def materialize(occ: Occurrence, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    if occ.kind in {'direct_current', 'direct_initial', 'standalone'}:
        shutil.copy2(occ.source, dest)
    else:
        with zipfile.ZipFile(occ.source) as zf, zf.open(occ.member) as src, dest.open('wb') as dst:
            shutil.copyfileobj(src, dst, length=1024 * 1024)
    if file_sha256(dest) != occ.sha256:
        raise RuntimeError(f'Materialized hash mismatch for {dest}')


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = list(rows[0].keys()) if rows else []
    with path.open('w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        w.writeheader()
        w.writerows(rows)


def autosize_sheet(ws, max_width: int = 60) -> None:
    for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), 1):
        width = 0
        for cell in col[: min(len(col), 500)]:
            if cell.value is not None:
                width = max(width, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), max_width)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet(title=title[:31])
    if not rows:
        ws['A1'] = 'No rows'
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    header_fill = PatternFill('solid', fgColor='1F4E78')
    for cell in ws[1]:
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical='top', wrap_text=True)
    for row in rows:
        ws.append([row.get(h, '') for h in headers])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    autosize_sheet(ws)


def main() -> None:
    if SNAPSHOT_ROOT.exists():
        shutil.rmtree(SNAPSHOT_ROOT)
    GPT_ROOT.mkdir(parents=True)

    occurrences = collect_occurrences()
    hash_occurrences(occurrences)
    by_sha: dict[str, list[Occurrence]] = defaultdict(list)
    for occ in occurrences:
        by_sha[occ.sha256].append(occ)

    artifact_rows: list[dict[str, object]] = []
    occurrence_rows: list[dict[str, object]] = []
    canonical_by_sha: dict[str, str] = {}

    for sha in sorted(by_sha):
        occs = by_sha[sha]
        chosen = choose_occurrence(occs)
        dest = destination_for(chosen)
        materialize(chosen, dest)
        rel = dest.relative_to(GPT_ROOT).as_posix()
        canonical_by_sha[sha] = rel
        artifact_rows.append({
            'canonical_relative_path': rel,
            'size_bytes': chosen.size,
            'sha256': sha,
            'category': category_for(rel),
            'batch': batch_from_text(rel),
            'status': status_for(rel),
            'description': description_for(rel),
            'recommended_git_tracking': git_tracking_for(chosen.size, rel),
            'canonical_source_kind': chosen.kind,
            'canonical_source_container': os.path.basename(chosen.source),
            'canonical_source_member_or_path': chosen.member,
            'occurrence_count': len(occs),
            'duplicate_alias_count': len(occs) - 1,
        })

    for occ in sorted(occurrences, key=lambda o: (o.logical_group, o.member, o.sha256)):
        occurrence_rows.append({
            'occurrence_kind': occ.kind,
            'source_container_or_path': occ.source,
            'logical_group': occ.logical_group,
            'member_or_relative_path': occ.member,
            'size_bytes': occ.size,
            'sha256': occ.sha256,
            'canonical_relative_path': canonical_by_sha[occ.sha256],
            'is_canonical_materialized_copy': 'yes' if (choose_occurrence(by_sha[occ.sha256]) is occ) else 'no',
        })

    bundle_rows: list[dict[str, object]] = []
    for zpath in BUNDLE_PATHS:
        with zipfile.ZipFile(zpath) as zf:
            members = [i for i in zf.infolist() if not i.is_dir()]
            member_shas = []
            for info in members:
                member = safe_member(info.filename)
                found = [o for o in occurrences if o.kind == 'zip_member' and o.source == str(zpath) and o.member == member]
                if len(found) != 1:
                    raise RuntimeError(f'Occurrence lookup failure for {zpath}:{member}')
                member_shas.append(found[0].sha256)
            represented = all(s in canonical_by_sha for s in member_shas)
        bundle_rows.append({
            'omitted_redundant_bundle': zpath.name,
            'bundle_size_bytes': zpath.stat().st_size,
            'bundle_sha256': file_sha256(zpath),
            'file_members': len(members),
            'unique_member_hashes': len(set(member_shas)),
            'all_members_represented_in_snapshot_by_sha256': 'yes' if represented else 'no',
            'reason_not_nested': 'All unique contents are expanded in this snapshot; nesting prior bundles would duplicate data.',
        })

    known_sources = read_known_sources()
    local_rows = [classify_local(normalize_user_path(line), known_sources) for line in USER_LOCAL_LISTING if line.strip()]

    metadata = GPT_ROOT / 'metadata'
    write_csv(metadata / 'ASSISTANT_ARTIFACT_INVENTORY.csv', artifact_rows)
    write_csv(metadata / 'ARTIFACT_OCCURRENCES_AND_ALIASES.csv', occurrence_rows)
    write_csv(metadata / 'OMITTED_REDUNDANT_BUNDLE_COVERAGE.csv', bundle_rows)
    write_csv(metadata / 'LOCAL_DIRECTORY_CLASSIFICATION_AND_MOVE_PLAN.csv', local_rows)

    # Current canonical index
    current_candidates = [r for r in artifact_rows if r['status'] in {'current_core', 'current_provenance'}]
    current_candidates = sorted(current_candidates, key=lambda r: (0 if 'B104I' in r['canonical_relative_path'] else 1 if 'B104H' in r['canonical_relative_path'] else 2, r['canonical_relative_path']))
    current_lines = [
        '# Current canonical / start-here index',
        '',
        'This index points to the latest core reconstruction and provenance artifacts. Earlier batches remain in the snapshot for auditability.',
        '',
        '## Recommended reading order',
        '',
        '1. `batches/B104G_20260829T150633Z/B104G_REPORT_20260829T151452Z.md` — complete row identity, features, labels, and ambiguity review.',
        '2. `batches/B104H_20260830T110259Z/GRAPHSAGE_NODE_MAPPING_README_20260830T110259Z.md` — node-to-Entrez mapping schema and interpretation.',
        '3. `batches/B104H_20260830T110259Z/B104H_SOURCE_PROVENANCE_AND_NEXT_PRIORITIES_20260830T110259Z.md` — source-provenance assessment.',
        '4. `batches/B104I_20260830T114918Z/B104I_PPI_PROVENANCE_MENCHE_AND_SPLIT_REPORT_20260830T114918Z.md` — PPI provenance and split semantics.',
        '5. `batches/B104I_20260830T114918Z/B104I_EVIDENCE_VS_LITERATURE_REGISTER_20260830T114918Z.csv` — exact/inferred/documented/open claim register.',
        '6. `batches/B104A_20260828T145842Z/B104A_REPORT_20260828T145842Z.md` — exact GO-label transformation breakthrough.',
        '7. `batches/B104E_20260829T121535Z/B104E_GOA_DATE_SCREEN_UNIPROT_AND_FEATURE_RULE_REPORT_20260829T121535Z.md` — GOA date and MSigDB feature results.',
        '8. `batches/B104C_20260828T194921Z/B104C_MSIGDB50_COLUMN_ORDER_REPORT_20260828T200948Z.md` — label-column order fingerprint.',
        '',
        '## Latest provenance snapshots',
        '',
        '- `results/actual_input_file_manifest_through_B104I_20260830T114918Z.csv`',
        '- `results/source_ledger_through_B104I_FINAL_20260830T114918Z.csv`',
        '- `results/provenance_events_through_B104I_20260830T114918Z.csv`',
        '',
        '## Current core files discovered automatically',
        '',
    ]
    for row in current_candidates:
        current_lines.append(f"- `{row['canonical_relative_path']}` — {row['description']}")
    (GPT_ROOT / 'CURRENT_CANONICAL_INDEX.md').write_text('\n'.join(current_lines) + '\n', encoding='utf-8')

    repo_layout = f'''# Recommended repository layout

```text
repo/
├── README.md
├── .gitignore
├── data/                         # local raw/reference data; normally not committed
│   ├── raw/
│   ├── derived/
│   ├── reference_papers/
│   ├── cache/
│   └── manifests/
├── gpt56sol/                     # this snapshot: scripts, reports, results, provenance
├── other_agent/                  # other investigator's reports/scripts/results
└── notebooks_or_analysis/        # your later clean external reproduction
```

Use `metadata/LOCAL_DIRECTORY_CLASSIFICATION_AND_MOVE_PLAN.csv` as the proposed path-by-path move plan for the directory listing supplied on {STAMP}.

## Git recommendation

- Commit `gpt56sol/`, `other_agent/`, and small provenance manifests.
- Keep `data/raw/`, `data/cache/`, and most binary archives out of ordinary Git.
- Use download scripts plus hashes for reproducibility; use Git LFS only when a binary must be versioned.
- Do not commit credentials, authenticated MSigDB cookies, or transient download caches.
- Quarantine or delete the four `dhimmel-gene-ontology-...tsv` files that are actually saved GitHub HTML pages.
'''
    (GPT_ROOT / 'RECOMMENDED_REPO_LAYOUT.md').write_text(repo_layout, encoding='utf-8')

    gitignore = '''# Local external data and caches
/data/raw/
/data/cache/
/data/reference_papers/
/data/derived/

# Keep manifests and explanatory READMEs
!/data/manifests/
!/data/README.md

# Python
__pycache__/
*.py[cod]
.venv/

# OS/editor
.DS_Store
Thumbs.db
*.swp

# Temporary downloads
*.part
*.tmp
'''
    (GPT_ROOT / 'GITIGNORE_TEMPLATE').write_text(gitignore, encoding='utf-8')

    gitattributes = '''# Optional Git LFS suggestions for assistant-generated compact binaries
*.zip filter=lfs diff=lfs merge=lfs -text
*.npy filter=lfs diff=lfs merge=lfs -text
*.npz filter=lfs diff=lfs merge=lfs -text
*.gz filter=lfs diff=lfs merge=lfs -text
'''
    (GPT_ROOT / 'GITATTRIBUTES_LFS_TEMPLATE').write_text(gitattributes, encoding='utf-8')

    # Local move plan narrative
    counts_origin = Counter(r['origin_class'] for r in local_rows)
    move_md = [
        '# Local directory classification and move plan',
        '',
        'This is a filename/path-based classification of the local `find .` listing supplied by the user. It does not silently assert hashes for files that were not uploaded or previously inventoried.',
        '',
        '## Summary by origin class',
        '',
        '| Origin class | Entries |',
        '|---|---:|',
    ]
    for key, count in sorted(counts_origin.items()):
        move_md.append(f'| `{key}` | {count} |')
    move_md += [
        '',
        '## Important actions',
        '',
        '- Move large external source files to `data/` and normally keep them out of ordinary Git.',
        '- Move assistant scripts, reports, script outputs, and provenance records to `gpt56sol/`.',
        '- Keep the other agent’s artifacts separate under `other_agent/` so claims remain attributable.',
        '- Delete or quarantine the four `dhimmel-gene-ontology-...tsv` files; prior inspection established they are HTML pages, not annotation TSVs.',
        '- Treat `goa_date_screen_cache/` as disposable working cache, not a canonical result.',
        '- Preserve the misspelled `HuamnBase-kidney.dat` filename until content/hash evidence supports a safe rename.',
        '',
        'The complete path-level plan is in `metadata/LOCAL_DIRECTORY_CLASSIFICATION_AND_MOVE_PLAN.csv` and the workbook.',
    ]
    (GPT_ROOT / 'LOCAL_DIRECTORY_MOVE_PLAN.md').write_text('\n'.join(move_md) + '\n', encoding='utf-8')

    # Verification script
    verify_script = r'''#!/usr/bin/env python3
import csv, hashlib, pathlib, sys
root = pathlib.Path(__file__).resolve().parents[1]
manifest = root / 'metadata' / 'SNAPSHOT_FILE_MANIFEST.csv'
failures = []
with manifest.open(newline='', encoding='utf-8') as f:
    for row in csv.DictReader(f):
        path = root / row['relative_path']
        if not path.is_file():
            failures.append(f"missing: {row['relative_path']}")
            continue
        h = hashlib.sha256()
        with path.open('rb') as fh:
            for block in iter(lambda: fh.read(1024 * 1024), b''):
                h.update(block)
        if h.hexdigest() != row['sha256']:
            failures.append(f"hash mismatch: {row['relative_path']}")
if failures:
    print('\n'.join(failures), file=sys.stderr)
    raise SystemExit(1)
print('All snapshot files verified.')
'''
    scripts_dir = GPT_ROOT / 'scripts'
    scripts_dir.mkdir(exist_ok=True)
    (scripts_dir / 'verify_snapshot.py').write_text(verify_script, encoding='utf-8')
    shutil.copy2(__file__, scripts_dir / 'build_gpt56sol_snapshot.py')

    # Workbook
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    summary_rows = [
        {'metric': 'snapshot_created_at_utc', 'value': STAMP},
        {'metric': 'unique_assistant_artifacts', 'value': len(artifact_rows)},
        {'metric': 'artifact_occurrences_and_aliases', 'value': len(occurrence_rows)},
        {'metric': 'unique_artifact_bytes', 'value': sum(int(r['size_bytes']) for r in artifact_rows)},
        {'metric': 'historical_bundles_expanded_not_nested', 'value': len(bundle_rows)},
        {'metric': 'user_local_entries_classified', 'value': len(local_rows)},
        {'metric': 'bundle_member_coverage', 'value': 'all represented by SHA-256' if all(r['all_members_represented_in_snapshot_by_sha256']=='yes' for r in bundle_rows) else 'INCOMPLETE'},
    ]
    add_sheet(wb, 'Summary', summary_rows)
    add_sheet(wb, 'Artifacts', artifact_rows)
    add_sheet(wb, 'Occurrences', occurrence_rows)
    add_sheet(wb, 'Local move plan', local_rows)
    add_sheet(wb, 'Bundle coverage', bundle_rows)
    workbook_path = metadata / 'GPT56SOL_COMPLETE_INVENTORY.xlsx'
    wb.save(workbook_path)
    # Reopen to validate workbook structure.
    wb2 = load_workbook(workbook_path, read_only=True)
    expected_sheets = {'Summary', 'Artifacts', 'Occurrences', 'Local move plan', 'Bundle coverage'}
    if set(wb2.sheetnames) != expected_sheets:
        raise RuntimeError(f'Workbook sheet mismatch: {wb2.sheetnames}')
    wb2.close()

    # README after counts known
    status_counts = Counter(str(r['status']) for r in artifact_rows)
    category_counts = Counter(str(r['category']) for r in artifact_rows)
    readme = [
        '# gpt56sol complete investigation snapshot',
        '',
        f'Created: `{STAMP}`',
        '',
        'This repo-ready snapshot contains every unique assistant-generated artifact currently recoverable from the working directories and historical delivery bundles. Historical bundles are expanded and deduplicated by SHA-256; the bundle ZIPs themselves are not nested.',
        '',
        '## Contents',
        '',
        f'- **{len(artifact_rows):,} unique assistant artifacts** ({sum(int(r["size_bytes"]) for r in artifact_rows):,} bytes before snapshot metadata).',
        f'- **{len(occurrence_rows):,} original occurrences/aliases** across direct files and prior bundles.',
        f'- **{len(bundle_rows)} prior delivery bundles** with complete member-by-member SHA-256 coverage.',
        f'- **{len(local_rows)} local paths** classified into `data/`, `gpt56sol/`, `other_agent/`, cache, or quarantine destinations.',
        '',
        '## Start here',
        '',
        '- `CURRENT_CANONICAL_INDEX.md`',
        '- `RECOMMENDED_REPO_LAYOUT.md`',
        '- `LOCAL_DIRECTORY_MOVE_PLAN.md`',
        '- `metadata/GPT56SOL_COMPLETE_INVENTORY.xlsx`',
        '- `metadata/ASSISTANT_ARTIFACT_INVENTORY.csv`',
        '- `metadata/LOCAL_DIRECTORY_CLASSIFICATION_AND_MOVE_PLAN.csv`',
        '',
        '## Snapshot policy',
        '',
        '- All unique artifact bytes are retained once.',
        '- Duplicate filenames and historical aliases are preserved in `ARTIFACT_OCCURRENCES_AND_ALIASES.csv`.',
        '- Prior delivery bundles are omitted as redundant containers; `OMITTED_REDUNDANT_BUNDLE_COVERAGE.csv` proves that every bundle member is represented by SHA-256.',
        '- Early incomplete or superseded deliveries are retained under `archive/` or marked in the inventory.',
        '- Raw external datasets are not added to this snapshot; the local move plan places those under `data/`.',
        '',
        '## Artifact status counts',
        '',
        '| Status | Files |',
        '|---|---:|',
    ]
    for key, count in sorted(status_counts.items()):
        readme.append(f'| `{key}` | {count} |')
    readme += ['', '## Artifact category counts', '', '| Category | Files |', '|---|---:|']
    for key, count in sorted(category_counts.items()):
        readme.append(f'| `{key}` | {count} |')
    readme += [
        '',
        '## Verification',
        '',
        'Run:',
        '',
        '```bash',
        'python gpt56sol/scripts/verify_snapshot.py',
        '```',
        '',
        '`metadata/SNAPSHOT_FILE_MANIFEST.csv` intentionally excludes itself and the ZIP container to avoid a self-referential checksum.',
    ]
    (GPT_ROOT / 'README.md').write_text('\n'.join(readme) + '\n', encoding='utf-8')

    # Snapshot-level manifest after all metadata except the manifest itself.
    manifest_rows = []
    for path in sorted(GPT_ROOT.rglob('*')):
        if not path.is_file():
            continue
        rel = path.relative_to(GPT_ROOT).as_posix()
        if rel == 'metadata/SNAPSHOT_FILE_MANIFEST.csv':
            continue
        manifest_rows.append({
            'relative_path': rel,
            'size_bytes': path.stat().st_size,
            'sha256': file_sha256(path),
        })
    write_csv(metadata / 'SNAPSHOT_FILE_MANIFEST.csv', manifest_rows)

    # Human-readable tree.
    tree_lines = []
    for path in sorted(GPT_ROOT.rglob('*')):
        rel = path.relative_to(GPT_ROOT)
        depth = len(rel.parts) - 1
        marker = '/' if path.is_dir() else ''
        tree_lines.append('  ' * depth + rel.name + marker)
    (metadata / 'FILE_TREE.txt').write_text('\n'.join(tree_lines) + '\n', encoding='utf-8')

    # Snapshot summary JSON.
    summary = {
        'created_at_utc': STAMP,
        'snapshot_root_name': GPT_ROOT.name,
        'unique_assistant_artifacts': len(artifact_rows),
        'artifact_occurrences': len(occurrence_rows),
        'unique_artifact_bytes_before_snapshot_metadata': sum(int(r['size_bytes']) for r in artifact_rows),
        'bundle_coverage_rows': len(bundle_rows),
        'all_prior_bundle_members_represented': all(r['all_members_represented_in_snapshot_by_sha256'] == 'yes' for r in bundle_rows),
        'local_paths_classified': len(local_rows),
        'snapshot_files_excluding_manifest_itself': len(manifest_rows),
        'snapshot_bytes_excluding_manifest_itself': sum(int(r['size_bytes']) for r in manifest_rows),
    }
    (metadata / 'SNAPSHOT_SUMMARY.json').write_text(json.dumps(summary, indent=2, sort_keys=True) + '\n', encoding='utf-8')

    # Recreate final manifest to include FILE_TREE and summary JSON.
    manifest_rows = []
    for path in sorted(GPT_ROOT.rglob('*')):
        if not path.is_file():
            continue
        rel = path.relative_to(GPT_ROOT).as_posix()
        if rel == 'metadata/SNAPSHOT_FILE_MANIFEST.csv':
            continue
        manifest_rows.append({'relative_path': rel, 'size_bytes': path.stat().st_size, 'sha256': file_sha256(path)})
    write_csv(metadata / 'SNAPSHOT_FILE_MANIFEST.csv', manifest_rows)

    # Verify all manifest rows.
    for row in manifest_rows:
        path = GPT_ROOT / str(row['relative_path'])
        if file_sha256(path) != row['sha256']:
            raise RuntimeError(f'Final manifest mismatch: {path}')

    # Zip the top-level gpt56sol directory.
    if ZIP_PATH.exists():
        ZIP_PATH.unlink()
    with zipfile.ZipFile(ZIP_PATH, 'w', compression=zipfile.ZIP_DEFLATED, compresslevel=6, allowZip64=True) as zf:
        for path in sorted(GPT_ROOT.rglob('*')):
            if path.is_file():
                arc = Path('gpt56sol') / path.relative_to(GPT_ROOT)
                zf.write(path, arc.as_posix())
    with zipfile.ZipFile(ZIP_PATH) as zf:
        bad = zf.testzip()
        if bad is not None:
            raise RuntimeError(f'ZIP integrity failure at {bad}')
        zip_members = [i for i in zf.infolist() if not i.is_dir()]
    zip_sha = file_sha256(ZIP_PATH)

    final_info = {
        'zip_path': str(ZIP_PATH),
        'zip_size_bytes': ZIP_PATH.stat().st_size,
        'zip_sha256': zip_sha,
        'zip_file_members': len(zip_members),
        'snapshot_directory': str(GPT_ROOT),
        **summary,
    }
    (ROOT / f'gpt56sol_snapshot_build_summary_{STAMP}.json').write_text(json.dumps(final_info, indent=2, sort_keys=True) + '\n', encoding='utf-8')
    print(json.dumps(final_info, indent=2, sort_keys=True))

if __name__ == '__main__':
    main()
