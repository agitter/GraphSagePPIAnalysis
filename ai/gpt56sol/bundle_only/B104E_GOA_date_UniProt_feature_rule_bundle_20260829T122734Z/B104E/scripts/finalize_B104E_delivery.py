#!/usr/bin/env python3
from __future__ import annotations

import csv
import hashlib
import json
import os
import shutil
import zipfile
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

STAMP = '20260829T122734Z'
EVENT_TIME = '2026-08-29T12:27:34Z'
BATCH = 'B104E'
BASE = Path('/mnt/data/ppi_repro_corrected/batches/B104E_20260829T121535Z')
RESULTS = Path('/mnt/data/ppi_repro_corrected/results')
RESULTS.mkdir(parents=True, exist_ok=True)

OLD_ACTUAL = RESULTS / 'actual_input_file_manifest_through_B104D_20260829T010311Z.csv'
OLD_LEDGER = RESULTS / 'source_ledger_through_B104D_FINAL_20260829T010311Z.csv'
OLD_EVENTS = RESULTS / 'provenance_events_through_B104D_FINAL_20260829T010311Z.csv'

UP_ZIP = Path('/mnt/data/uniprot_2016_mapping_audit_ledger.zip')
GOA_ZIP = Path('/mnt/data/goa_date_screen_results.zip')

ACTUAL_OUT = RESULTS / f'actual_input_file_manifest_through_B104E_{STAMP}.csv'
ACTUAL_MD = RESULTS / f'actual_input_file_manifest_through_B104E_{STAMP}.md'
LEDGER_PRE = RESULTS / f'source_ledger_through_B104E_PREBUNDLE_{STAMP}.csv'
LEDGER_FINAL = RESULTS / f'source_ledger_through_B104E_FINAL_{STAMP}.csv'
LEDGER_MD = RESULTS / f'source_ledger_through_B104E_FINAL_{STAMP}.md'
EVENTS_PRE = RESULTS / f'provenance_events_through_B104E_PREBUNDLE_{STAMP}.csv'
EVENTS_FINAL = RESULTS / f'provenance_events_through_B104E_FINAL_{STAMP}.csv'
CLEARANCE = BASE / f'B104E_DELETION_CLEARANCE_{STAMP}.md'
CHECKSUMS_PRE = BASE / f'B104E_OUTPUT_CHECKSUMS_PREBUNDLE_{STAMP}.csv'
VALIDATION = RESULTS / f'B104E_FINAL_DELIVERY_VALIDATION_{STAMP}.json'
WORKBOOK = RESULTS / f'B104E_provenance_workbook_FINAL_{STAMP}.xlsx'
BUNDLE = Path(f'/mnt/data/B104E_GOA_date_UniProt_feature_rule_bundle_{STAMP}.zip')
BUNDLE_HASH_FILE = RESULTS / f'B104E_bundle_sha256_{STAMP}.txt'
FINAL_CHECKSUMS = RESULTS / f'B104E_final_delivery_checksums_{STAMP}.csv'


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open('rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def read_csv(path: Path):
    with path.open(newline='', encoding='utf-8-sig') as f:
        r = csv.DictReader(f)
        return r.fieldnames or [], list(r)


def write_csv(path: Path, fields: list[str], rows: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, '') for k in fields})


def append_semicolon(existing: str, text: str) -> str:
    if not existing:
        return text
    parts = [p.strip() for p in existing.split(';') if p.strip()]
    if text not in parts:
        parts.append(text)
    return '; '.join(parts)


def md_escape(x) -> str:
    return str(x).replace('|', '\\|').replace('\n', '<br>')


def write_manifest_md(path: Path, rows: list[dict]):
    cols = ['record_type','artifact_name','local_status','origin_in_this_run','analysis_role','size_bytes','sha256','batch_id','direct_or_canonical_source_url','raw_retention_status','raw_available_in_conversation']
    with path.open('w', encoding='utf-8') as f:
        f.write('# Actual input manifest through B104E\n\n')
        f.write(f'Generated: `{EVENT_TIME}`\n\n')
        f.write('This is the current-state manifest. Append-only history is preserved separately in the source ledger and provenance-event table.\n\n')
        f.write('| ' + ' | '.join(cols) + ' |\n')
        f.write('| ' + ' | '.join(['---'] * len(cols)) + ' |\n')
        for row in rows:
            f.write('| ' + ' | '.join(md_escape(row.get(c,'')) for c in cols) + ' |\n')


def write_ledger_md(path: Path, rows: list[dict]):
    # Focused readable view of B104E additions, while CSV retains full history.
    cols = ['record_type','artifact_name','local_status','analysis_role','direct_or_canonical_source_url','size_bytes','sha256','deletion_state','hash_authority','runtime_verification_status','retained_derivative_paths']
    b_rows = [r for r in rows if r.get('batch_id') == BATCH]
    with path.open('w', encoding='utf-8') as f:
        f.write('# Source ledger through B104E\n\n')
        f.write(f'Generated: `{EVENT_TIME}`\n\n')
        f.write(f'The CSV contains the complete append-only history. This view shows the {len(b_rows)} B104E additions.\n\n')
        f.write('| ' + ' | '.join(cols) + ' |\n')
        f.write('| ' + ' | '.join(['---'] * len(cols)) + ' |\n')
        for row in b_rows:
            f.write('| ' + ' | '.join(md_escape(row.get(c,'')) for c in cols) + ' |\n')


def make_actual_manifest():
    fields, rows = read_csv(OLD_ACTUAL)
    by_name = {r['artifact_name']: r for r in rows}
    # Update major reused inputs.
    if 'graphsage_ppi.zip' in by_name:
        by_name['graphsage_ppi.zip']['used_by'] = append_semicolon(by_name['graphsage_ppi.zip'].get('used_by',''), 'B104E exact feature reconstruction and source-date validation')
    for n in ['msigdb_v5.1_files_to_download_locally.zip','msigdb_v5.2_files_to_download_locally.zip','msigdb_v6.0_files_to_download_locally.zip']:
        if n in by_name:
            by_name[n]['used_by'] = append_semicolon(by_name[n].get('used_by',''), 'B104E provenance-safe MSigDB feature-version screen')
            by_name[n]['notes'] = append_semicolon(by_name[n].get('notes',''), 'User previously reported deleting older MSigDB files locally to make space; B104E reused already-retained/runtime-accessible bytes and did not assume local availability.')
    n='B104C_msigdb_v5.0_normalized_entrez_gene_sets_20260828T194921Z.tsv.gz'
    if n in by_name:
        by_name[n]['used_by'] = append_semicolon(by_name[n].get('used_by',''), 'B104E provenance-safe MSigDB v5.0 feature-version screen')
    # Supersede incomplete B104D UniProt package status in current-state manifest.
    if 'uniprot_2016_mapping.zip' in by_name:
        by_name['uniprot_2016_mapping.zip']['notes'] = append_semicolon(by_name['uniprot_2016_mapping.zip'].get('notes',''), 'Superseded by complete B104E package uniprot_2016_mapping_audit_ledger.zip containing extracted records and per-release provenance.')
        by_name['uniprot_2016_mapping.zip']['local_status'] = 'superseded_ledger_only_package'
    # New actual uploaded packages.
    retained_up = BASE / 'retained_inputs/uniprot_2016_mapping_audit'
    retained_goa = BASE / 'retained_inputs/goa_date_screen_results'
    rows.extend([
        {
            'record_type':'actual_input','artifact_name':UP_ZIP.name,'local_path':str(UP_ZIP),'local_status':'present_verified_complete_package','origin_in_this_run':'user_upload',
            'analysis_role':'complete user-side UniProt 2016 mapping audit outputs for O95073 and Q9Y620','used_by':'B104E UniProt record audit and mapping interpretation',
            'direct_or_canonical_source_url':'multiple official UniProt previous-release URLs recorded per release in retained provenance JSON and ledger',
            'source_page_url':'https://ftp.uniprot.org/pub/databases/uniprot/previous_releases/','url_status':'official source URLs retained per release',
            'retrieval_status':'uploaded_by_user; three parent archives downloaded, verified, streamed, and deleted on user system','retrieved_at_utc':'',
            'size_bytes':str(UP_ZIP.stat().st_size),'sha256':sha256(UP_ZIP),'parent_or_derivation':'output of download_extract_uniprot_2016_mapping_audit.py',
            'notes':'ZIP integrity passed; 16 members; contains three extracted DAT files, three TSV summaries, three provenance JSON files, three official RELEASE metalinks, and the append-only audit ledger.',
            'batch_id':BATCH,'inventory_sha256':'','received_at_utc':'2026-08-29T12:08:14Z','integrity_verified_at_utc':EVENT_TIME,'analysis_completed_at_utc':EVENT_TIME,
            'deletion_clearance_issued_at_utc':EVENT_TIME,'user_deletion_confirmed_at_utc':'','raw_retention_status':'all compact extracted records and provenance retained; uploaded wrapper ZIP dispensable after clearance',
            'raw_available_in_conversation':'yes_pending_user_deletion','retained_derivative_paths':str(retained_up),
            'derivative_sha256s':'per-file hashes recorded in B104E output checksum table and per-release provenance JSON',
            'parser_version_or_script_sha256':'validate_uploaded_audit_archives.py:' + sha256(BASE/'scripts/validate_uploaded_audit_archives.py') + '; parse_uniprot_target_records.py:' + sha256(BASE/'scripts/parse_uniprot_target_records.py'),
            'raw_to_derived_reconciliation':'ZIP test passed; 16/16 members extracted; all small-output hashes agree with audit ledger/provenance; both O95073 and Q9Y620 found in all three releases.',
            'reacquisition_url':'regenerate with retained download/extract script from official UniProt previous releases','deletion_notes':'Safe to delete conversation ZIP after final B104E bundle validation; keep user-side extracted result directory or frozen bundle.'
        },
        {
            'record_type':'actual_input','artifact_name':GOA_ZIP.name,'local_path':str(GOA_ZIP),'local_status':'present_verified_complete_package','origin_in_this_run':'user_upload',
            'analysis_role':'user-side GOA release 158–169 source-date screening outputs','used_by':'B104E GOA source-date, term-selection, and column-order trajectory analyses',
            'direct_or_canonical_source_url':'per-release EBI GOA GAF/GPI URLs recorded in each retained JSON','source_page_url':'https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/',
            'url_status':'official URLs retained per release','retrieval_status':'uploaded_by_user; raw GAF/GPI pairs downloaded, verified, analyzed, and deleted sequentially on user system','retrieved_at_utc':'',
            'size_bytes':str(GOA_ZIP.stat().st_size),'sha256':sha256(GOA_ZIP),'parent_or_derivation':'output of screen_goa_release_date_range.py using B104D reference pack',
            'notes':'ZIP integrity passed; 15 members; includes one detailed JSON per release 158–169, summary CSV, metadata JSON, and append-only events CSV.',
            'batch_id':BATCH,'inventory_sha256':'','received_at_utc':'2026-08-29T12:08:14Z','integrity_verified_at_utc':EVENT_TIME,'analysis_completed_at_utc':EVENT_TIME,
            'deletion_clearance_issued_at_utc':EVENT_TIME,'user_deletion_confirmed_at_utc':'','raw_retention_status':'complete compact per-release results and source hashes retained; uploaded wrapper ZIP dispensable after clearance',
            'raw_available_in_conversation':'yes_pending_user_deletion','retained_derivative_paths':str(retained_goa),
            'derivative_sha256s':'per-file hashes recorded in B104E output checksum table',
            'parser_version_or_script_sha256':'validate_uploaded_audit_archives.py:' + sha256(BASE/'scripts/validate_uploaded_audit_archives.py') + '; summarize_goa_date_screen.py:' + sha256(BASE/'scripts/summarize_goa_date_screen.py'),
            'raw_to_derived_reconciliation':'ZIP test passed; 12 release JSONs present and internally agree with summary; 3 source-integrity/analysis/deletion events per release; 81 total package checks passed.',
            'reacquisition_url':'https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/','deletion_notes':'Safe to delete conversation ZIP after final B104E bundle validation; keep frozen bundle.'
        },
        {
            'record_type':'retained_input_collection','artifact_name':'B104E retained UniProt audit outputs','local_path':str(retained_up),'local_status':'present_verified','origin_in_this_run':'extracted_from_verified_user_upload',
            'analysis_role':'complete compact records and provenance from three 1.5-GB UniProt parent archives','used_by':'B104E and future mapping-source audits',
            'direct_or_canonical_source_url':'per-release URLs in provenance JSON','source_page_url':'https://ftp.uniprot.org/pub/databases/uniprot/previous_releases/','url_status':'official',
            'retrieval_status':'retained_from_verified_package','retrieved_at_utc':EVENT_TIME,'size_bytes':str(sum(p.stat().st_size for p in retained_up.rglob('*') if p.is_file())),
            'sha256':'collection_manifested_per_file','parent_or_derivation':UP_ZIP.name+':'+sha256(UP_ZIP),'notes':'Four files per release plus audit ledger; all declared hashes validated.',
            'batch_id':BATCH,'received_at_utc':EVENT_TIME,'integrity_verified_at_utc':EVENT_TIME,'analysis_completed_at_utc':EVENT_TIME,
            'raw_retention_status':'required compact retained input','raw_available_in_conversation':'not_applicable_retained_collection','retained_derivative_paths':str(retained_up),
            'derivative_sha256s':'see B104E output checksums','parser_version_or_script_sha256':'','raw_to_derived_reconciliation':'16 ZIP members -> 16 retained files; exact member bytes retained',
            'reacquisition_url':'official UniProt URLs in provenance JSON','deletion_notes':''
        },
        {
            'record_type':'retained_input_collection','artifact_name':'B104E retained GOA 158–169 date-screen results','local_path':str(retained_goa),'local_status':'present_verified','origin_in_this_run':'extracted_from_verified_user_upload',
            'analysis_role':'compact per-release source hashes, headers, mapping statistics, label comparison, term selection, and order scores','used_by':'B104E and future source-date audits',
            'direct_or_canonical_source_url':'per-release GAF/GPI URLs in JSON','source_page_url':'https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/','url_status':'official',
            'retrieval_status':'retained_from_verified_package','retrieved_at_utc':EVENT_TIME,'size_bytes':str(sum(p.stat().st_size for p in retained_goa.rglob('*') if p.is_file())),
            'sha256':'collection_manifested_per_file','parent_or_derivation':GOA_ZIP.name+':'+sha256(GOA_ZIP),'notes':'12 release JSON files plus summary, metadata, and append-only events.',
            'batch_id':BATCH,'received_at_utc':EVENT_TIME,'integrity_verified_at_utc':EVENT_TIME,'analysis_completed_at_utc':EVENT_TIME,
            'raw_retention_status':'required compact retained input','raw_available_in_conversation':'not_applicable_retained_collection','retained_derivative_paths':str(retained_goa),
            'derivative_sha256s':'see B104E output checksums','parser_version_or_script_sha256':'','raw_to_derived_reconciliation':'15 ZIP members -> 15 retained files; exact member bytes retained',
            'reacquisition_url':'official EBI GOA URLs in per-release JSON','deletion_notes':''
        },
    ])
    # Ensure current rows replace updated references in list where needed.
    # rows contains original dict objects; updates above mutate them.
    write_csv(ACTUAL_OUT, fields, rows)
    write_manifest_md(ACTUAL_MD, rows)
    return fields, rows


def base_ledger_row(fields):
    return {k:'' for k in fields}


def make_source_ledger(actual_rows):
    fields, rows = read_csv(OLD_LEDGER)
    def add(**kwargs):
        row=base_ledger_row(fields); row.update(kwargs); rows.append(row)
    up_sha=sha256(UP_ZIP); goa_sha=sha256(GOA_ZIP)
    add(record_type='actual_input',artifact_name=UP_ZIP.name,local_path=str(UP_ZIP),local_status='present_verified_complete_package',origin_in_this_run='user_upload',analysis_role='complete UniProt 2016 mapping-audit package',used_by='B104E UniProt audit',direct_or_canonical_source_url='multiple official URLs in retained per-release provenance',source_page_url='https://ftp.uniprot.org/pub/databases/uniprot/previous_releases/',url_status='official URLs retained',retrieval_status='uploaded_by_user',size_bytes=str(UP_ZIP.stat().st_size),sha256=up_sha,parent_or_derivation='download_extract_uniprot_2016_mapping_audit.py output',notes='16-member ZIP; all compact outputs present and validated.',batch_id=BATCH,deletion_state='deletion_clearance_pending_user_confirmation',event_recorded_at_utc=EVENT_TIME,hash_authority='runtime SHA-256; user-side parent MD5/size against official metalinks',runtime_verification_status='ZIP integrity and 81 package checks passed',retained_derivative_paths=str(BASE/'retained_inputs/uniprot_2016_mapping_audit'),parser_script_sha256=sha256(BASE/'scripts/validate_uploaded_audit_archives.py'),raw_to_derived_reconciliation='16/16 members retained byte-for-byte')
    add(record_type='actual_input',artifact_name=GOA_ZIP.name,local_path=str(GOA_ZIP),local_status='present_verified_complete_package',origin_in_this_run='user_upload',analysis_role='GOA releases 158–169 screen package',used_by='B104E GOA date screen',direct_or_canonical_source_url='per-release official EBI URLs in retained JSON',source_page_url='https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/',url_status='official URLs retained',retrieval_status='uploaded_by_user',size_bytes=str(GOA_ZIP.stat().st_size),sha256=goa_sha,parent_or_derivation='screen_goa_release_date_range.py output',notes='15-member ZIP; 12 detailed release results plus summary, metadata, and events.',batch_id=BATCH,deletion_state='deletion_clearance_pending_user_confirmation',event_recorded_at_utc=EVENT_TIME,hash_authority='runtime SHA-256 for package; user-side SHA-256 for transient source GAF/GPI files',runtime_verification_status='ZIP integrity and 81 package checks passed',retained_derivative_paths=str(BASE/'retained_inputs/goa_date_screen_results'),parser_script_sha256=sha256(BASE/'scripts/validate_uploaded_audit_archives.py'),raw_to_derived_reconciliation='15/15 members retained byte-for-byte')
    # UniProt parent archives and retained outputs.
    up_root=BASE/'retained_inputs/uniprot_2016_mapping_audit'
    ledger_fields, ledger_rows = read_csv(up_root/'uniprot_2016_mapping_audit_ledger.csv')
    by_rel={r['release']:r for r in ledger_rows}
    for rel in ['2016_04','2016_05','2016_06']:
        r=by_rel[rel]
        url=r.get('url','') or f'https://ftp.uniprot.org/pub/databases/uniprot/previous_releases/release-{rel}/knowledgebase/uniprot_sprot-only{rel}.tar.gz'
        retained='; '.join(str(p) for p in sorted((up_root/rel).glob('*')))
        add(record_type='external_parent_input_processed_user_side',artifact_name=f'uniprot_sprot-only{rel}.tar.gz',local_path=r.get('archive_path',''),local_status='deleted_after_success_on_user_system',origin_in_this_run='downloaded_by_user_audit_script',analysis_role='date-matched reviewed UniProtKB records for O95073 and Q9Y620',used_by='B104E mapping-component audit',direct_or_canonical_source_url=url,source_page_url=f'https://ftp.uniprot.org/pub/databases/uniprot/previous_releases/release-{rel}/knowledgebase/',url_status='official',retrieval_status='downloaded_verified_streamed_deleted_on_user_system',retrieved_at_utc=r.get('started_at_utc',''),size_bytes=r.get('archive_size_bytes',''),sha256=r.get('archive_sha256',''),parent_or_derivation='',notes=f"Official MD5 expected={r.get('expected_md5','')}; observed={r.get('archive_md5','')}; records_scanned={r.get('records_scanned','')}; targets_found={r.get('target_records_found','')}",batch_id=BATCH,user_local_status='deleted_after_success',deletion_state='deleted_after_success_on_user_system',event_recorded_at_utc=r.get('finished_at_utc',''),hash_authority='user-side SHA-256 plus official metalink byte-size/MD5 verification',runtime_verification_status='retained provenance JSON, metalink, extracted DAT/TSV, and ledger cross-validated',retained_derivative_paths=retained,parser_script_sha256='',raw_to_derived_reconciliation='complete Swiss-Prot stream scanned; exactly two target records retained; parent archive deleted only after validation')
        for p in sorted((up_root/rel).glob('*')):
            add(record_type='retained_input',artifact_name=p.name,local_path=str(p),local_status='present_verified',origin_in_this_run='extracted_from_complete_user_audit_package',analysis_role='retained UniProt audit record/provenance',used_by='B104E UniProt audit and future reproduction',direct_or_canonical_source_url=url if p.suffix in {'.dat','.tsv'} else (f'https://ftp.uniprot.org/pub/databases/uniprot/previous_releases/release-{rel}/knowledgebase/RELEASE.metalink' if 'metalink' in p.name else ''),source_page_url=f'https://ftp.uniprot.org/pub/databases/uniprot/previous_releases/release-{rel}/knowledgebase/',url_status='official parent source recorded',retrieval_status='retained_from_verified_user_package',retrieved_at_utc=EVENT_TIME,size_bytes=str(p.stat().st_size),sha256=sha256(p),parent_or_derivation=UP_ZIP.name+':'+up_sha,notes='',batch_id=BATCH,deletion_state='retained',event_recorded_at_utc=EVENT_TIME,hash_authority='runtime SHA-256; matched package member/provenance where declared',runtime_verification_status='present and hash-verified',retained_derivative_paths=str(p),raw_to_derived_reconciliation='exact retained ZIP member')
    # Ledger file itself.
    p=up_root/'uniprot_2016_mapping_audit_ledger.csv'
    add(record_type='retained_input',artifact_name=p.name,local_path=str(p),local_status='present_verified',origin_in_this_run='extracted_from_complete_user_audit_package',analysis_role='append-only user-side UniProt audit ledger',used_by='B104E provenance validation',direct_or_canonical_source_url='per-release URLs in ledger',source_page_url='https://ftp.uniprot.org/pub/databases/uniprot/previous_releases/',url_status='official URLs recorded',retrieval_status='retained_from_verified_user_package',retrieved_at_utc=EVENT_TIME,size_bytes=str(p.stat().st_size),sha256=sha256(p),parent_or_derivation=UP_ZIP.name+':'+up_sha,notes='Three successful release rows.',batch_id=BATCH,deletion_state='retained',event_recorded_at_utc=EVENT_TIME,hash_authority='runtime SHA-256',runtime_verification_status='present and parsed',retained_derivative_paths=str(p),raw_to_derived_reconciliation='exact retained ZIP member')
    # GOA transient source files and compact results.
    goa_root=BASE/'retained_inputs/goa_date_screen_results'
    for rel in range(158,170):
        jp=goa_root/f'goa_release_{rel}_screen.json'
        obj=json.load(jp.open())
        for kind in ['gaf','gpi']:
            sf=obj['source_files'][kind]
            add(record_type='external_parent_input_processed_user_side',artifact_name=Path(sf['url_or_local']).name,local_path=sf.get('path',''),local_status='deleted_after_success_on_user_system',origin_in_this_run='downloaded_by_user_date_screen_script',analysis_role=f'GOA release {rel} {kind.upper()} source-date candidate',used_by=f'B104E release {rel} membership, term-selection, and order screen',direct_or_canonical_source_url=sf['url_or_local'],source_page_url='https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/',url_status='official EBI archive URL',retrieval_status='downloaded_integrity_checked_analyzed_deleted_on_user_system',retrieved_at_utc='',size_bytes=str(sf['size_bytes']),sha256=sf['sha256'],parent_or_derivation='',notes=f"listed_release_date={obj.get('listed_release_date','')}; GO header={next((h for h in obj['gaf']['headers'] if h.startswith('!GO-version:')), '') if kind=='gaf' else ''}",batch_id=BATCH,user_local_status='deleted_after_success',deletion_state='deleted_after_success_on_user_system',event_recorded_at_utc=EVENT_TIME,hash_authority='user-side SHA-256 retained in per-release JSON; releases 158/159 previously independently verified from raw uploads',runtime_verification_status='compact result JSON internally validated; raw 160–169 not independently re-downloaded in runtime due DNS failure',retained_derivative_paths=str(jp),parser_script_sha256='',raw_to_derived_reconciliation=f"source integrity passed; analyzed; raw downloaded pair deleted; result={jp.name}")
        add(record_type='retained_input',artifact_name=jp.name,local_path=str(jp),local_status='present_verified',origin_in_this_run='extracted_from_verified_user_package',analysis_role=f'complete compact GOA release {rel} screening result',used_by='B104E release-date trajectory and future auditing',direct_or_canonical_source_url='per-source URLs inside JSON',source_page_url='https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/',url_status='official source URLs retained',retrieval_status='retained_from_verified_user_package',retrieved_at_utc=EVENT_TIME,size_bytes=str(jp.stat().st_size),sha256=sha256(jp),parent_or_derivation=GOA_ZIP.name+':'+goa_sha,notes=f"mismatches={obj['label_comparison']['total_mismatches']}; exact_columns={obj['label_comparison']['exact_columns']}; order_lcs={obj['column_order']['lcs']}",batch_id=BATCH,deletion_state='retained',event_recorded_at_utc=EVENT_TIME,hash_authority='runtime SHA-256',runtime_verification_status='parsed and reconciled against summary',retained_derivative_paths=str(jp),raw_to_derived_reconciliation='exact retained ZIP member')
    for name in ['goa_release_date_screen_summary_20260829T120614Z.csv','goa_date_screen_run_metadata.json','goa_date_screen_events.csv']:
        p=goa_root/name
        add(record_type='retained_input',artifact_name=p.name,local_path=str(p),local_status='present_verified',origin_in_this_run='extracted_from_verified_user_package',analysis_role='GOA date-screen summary/metadata/event history',used_by='B104E validation and provenance',direct_or_canonical_source_url='per-release URLs in detailed JSON',source_page_url='https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/',url_status='official source URLs retained',retrieval_status='retained_from_verified_user_package',retrieved_at_utc=EVENT_TIME,size_bytes=str(p.stat().st_size),sha256=sha256(p),parent_or_derivation=GOA_ZIP.name+':'+goa_sha,notes='',batch_id=BATCH,deletion_state='retained',event_recorded_at_utc=EVENT_TIME,hash_authority='runtime SHA-256',runtime_verification_status='parsed and validated',retained_derivative_paths=str(p),raw_to_derived_reconciliation='exact retained ZIP member')
    # Generated scientific outputs and scripts.
    for p in sorted(BASE.rglob('*')):
        if not p.is_file():
            continue
        if 'retained_inputs' in p.parts or 'logs' in p.parts or p.name == Path(__file__).name:
            continue
        if p in {CLEARANCE,CHECKSUMS_PRE}:
            continue
        record_type='generated_analysis' if 'analysis' in p.parts else ('analysis_script' if 'scripts' in p.parts else 'generated_report')
        add(record_type=record_type,artifact_name=p.name,local_path=str(p),local_status='present',origin_in_this_run='generated_by_B104E',analysis_role='B104E scientific result, report, diagnostic, or script',used_by='B104E delivery and future reproduction',direct_or_canonical_source_url='',source_page_url='',url_status='',retrieval_status='generated_locally',retrieved_at_utc=EVENT_TIME,size_bytes=str(p.stat().st_size),sha256=sha256(p),parent_or_derivation='B104E analysis inputs and retained prior compact derivatives',notes='',batch_id=BATCH,deletion_state='retained',event_recorded_at_utc=EVENT_TIME,hash_authority='runtime SHA-256',runtime_verification_status='present',retained_derivative_paths=str(p),parser_script_sha256=sha256(p) if record_type=='analysis_script' else '',raw_to_derived_reconciliation='see report and diagnostics')
    write_csv(LEDGER_PRE, fields, rows)
    return fields, rows


def make_events():
    fields, rows = read_csv(OLD_EVENTS)
    def ev(kind, artifact, status, details, when=EVENT_TIME):
        rows.append({'event_time_utc':when,'batch_id':BATCH,'event_type':kind,'artifact_name':artifact,'status':status,'details':details})
    ev('input_received',UP_ZIP.name,'received',f'bytes={UP_ZIP.stat().st_size}; sha256={sha256(UP_ZIP)}; 16-member complete audit package')
    ev('input_received',GOA_ZIP.name,'received',f'bytes={GOA_ZIP.stat().st_size}; sha256={sha256(GOA_ZIP)}; 15-member complete date-screen package')
    ev('package_integrity_validation','B104E uploaded audit packages','passed','ZIP integrity and 81 structural/hash/reconciliation checks all passed.')
    ev('goa_date_screen','releases 158-169','accepted','Release 159 uniquely exact: 121/121 columns, 0/516428 differences. Closest alternative release 160 has 652 differences.')
    ev('goa_term_selection_date_screen','releases 158-169','accepted','Only release 159 simultaneously gives exact memberships, exact top-121 candidate set, and exactly 121 terms at the >=1000 full-human threshold.')
    ev('column_order_date_screen','releases 158-169','no_perfect_model','Best LCS remains 94/121 for releases 158-166 and 93/121 for 167-169; duplicate orientation 001 stable across all releases.')
    ev('uniprot_parent_archive_validation','2016_04, 2016_05, 2016_06 reviewed-only archives','passed','Official byte sizes and MD5 values matched; parent SHA-256 values retained; all three full archives streamed and deleted after successful target extraction.')
    ev('historical_mapping_correction','O95073 -> 25788','interpretation_revised','Official reviewed UniProt records in 2016_04/05/06 explicitly carry O95073/FSBP -> GeneIDs 100861412 and 25788; edge is genuine historical UniProt data, not a parser artifact.')
    ev('mapping_semantics','O95073/Q9Y620 component','accepted_with_caution','Exact labels require FSBP annotations not to be projected to GraphSAGE GeneID 25788/RAD54B. This behavior can result from Entrez-native annotations or symbol-aware component resolution; original mechanism remains unproven.')
    ev('feature_reconstruction','GraphSAGE ppi-feats.npy','exact','Collection order C1,C3,C7; preserve GMT row order; retain sets with >=200 unique Entrez IDs; stop after 50. All 2,820,550 resolved feature cells match exactly in two implementations.')
    ev('feature_zero_column','feature column 10','identified','Column 10 is chryq11: 204 genes in full MSigDB, none among 4,268 resolved GraphSAGE genes.')
    ev('feature_version_attribution','MSigDB v5.0-v6.0','prior_claim_superseded','All supplied versions 5.0, 5.1, 5.2, and 6.0 yield the same 50 membership vectors and exact feature matrix. v5.2 is not uniquely identifiable.')
    ev('provenance_correction','MSigDB v5.0 screen','authoritative_rerun','Initial exploratory run touched a residual raw mount after user deletion confirmation. It was moved to logs/superseded; authoritative rerun uses the hash-verified retained normalized v5.0 derivative.')
    ev('independent_download_attempt','GOA releases 160 and 168','failed_environment_dns','Independent runtime re-download attempted but DNS resolution failed; user-side source hashes/results retained. This limitation is recorded, not treated as source failure.')
    ev('deletion_clearance_issued',UP_ZIP.name,'safe_after_bundle_validation',f'Complete compact outputs retained; raw upload sha256={sha256(UP_ZIP)}')
    ev('deletion_clearance_issued',GOA_ZIP.name,'safe_after_bundle_validation',f'Complete compact per-release results retained; raw upload sha256={sha256(GOA_ZIP)}')
    write_csv(EVENTS_PRE, fields, rows)
    return fields, rows


def write_clearance():
    text=f'''# SAFE TO DELETE — BATCH B104E\n\nIssued: `{EVENT_TIME}`\n\nThe following conversation attachments may be deleted after the final B104E bundle and validation files are available:\n\n```text\n{UP_ZIP.name}\nBytes: {UP_ZIP.stat().st_size}\nSHA-256: {sha256(UP_ZIP)}\n\n{GOA_ZIP.name}\nBytes: {GOA_ZIP.stat().st_size}\nSHA-256: {sha256(GOA_ZIP)}\n```\n\nRetained for the UniProt package:\n\n- three extracted complete DAT record files;\n- three compact TSV summaries;\n- three per-release provenance JSON files;\n- three official RELEASE metalinks;\n- the append-only audit ledger;\n- runtime validation and scientific interpretation outputs.\n\nRetained for the GOA package:\n\n- one detailed JSON for every release 158 through 169;\n- the release summary;\n- run metadata;\n- append-only source-integrity, analysis, and cleanup events;\n- runtime validation and scientific summary outputs.\n\nThe user-side large UniProt and GOA parent downloads were already deleted only after successful validation by their respective scripts. Keep the local compact audit-result directories or the frozen B104E bundle.\n\nAfter deleting the two conversation attachments, report:\n\n```text\nDeleted B104E\n```\n'''
    CLEARANCE.write_text(text, encoding='utf-8')


def create_pre_checksums():
    rows=[]
    for p in sorted(BASE.rglob('*')):
        if p.is_file() and 'logs/superseded' not in str(p):
            rows.append({'relative_path':str(p.relative_to(BASE)),'size_bytes':p.stat().st_size,'sha256':sha256(p)})
    write_csv(CHECKSUMS_PRE,['relative_path','size_bytes','sha256'],rows)
    return rows


def create_workbook(actual_rows, ledger_rows, event_rows, checksum_rows, path: Path):
    wb=Workbook(); wb.remove(wb.active)
    def add_sheet(name, rows, fields):
        ws=wb.create_sheet(name)
        ws.sheet_view.showGridLines=False
        ws.append(fields)
        for cell in ws[1]:
            cell.font=Font(bold=True,color='FFFFFF')
            cell.fill=PatternFill('solid', fgColor='1F4E78')
            cell.alignment=Alignment(wrap_text=True,vertical='top')
        for row in rows:
            ws.append([row.get(f,'') for f in fields])
        ws.freeze_panes='A2'
        ws.auto_filter.ref=ws.dimensions
        for col_idx, field in enumerate(fields,1):
            vals=[str(field)] + [str(r.get(field,'')) for r in rows[:500]]
            width=min(max(10,max(len(v) for v in vals)+2),60)
            ws.column_dimensions[get_column_letter(col_idx)].width=width
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment=Alignment(vertical='top',wrap_text=True)
                # imported / linked values in green as per spreadsheet convention
                c.font=Font(color='008000')
        return ws
    actual_fields=['record_type','artifact_name','local_path','local_status','origin_in_this_run','analysis_role','used_by','direct_or_canonical_source_url','source_page_url','retrieval_status','size_bytes','sha256','batch_id','raw_retention_status','raw_available_in_conversation','retained_derivative_paths','raw_to_derived_reconciliation','deletion_notes']
    ledger_fields=['record_type','artifact_name','local_path','local_status','analysis_role','direct_or_canonical_source_url','size_bytes','sha256','batch_id','deletion_state','hash_authority','runtime_verification_status','retained_derivative_paths','raw_to_derived_reconciliation']
    event_fields=['event_time_utc','batch_id','event_type','artifact_name','status','details']
    check_fields=['relative_path','size_bytes','sha256']
    add_sheet('Actual Inputs',actual_rows,actual_fields)
    add_sheet('B104E Ledger Additions',[r for r in ledger_rows if r.get('batch_id')==BATCH],ledger_fields)
    add_sheet('Provenance Events',event_rows,event_fields)
    add_sheet('Checksums',checksum_rows,check_fields)
    wb.save(path)


def create_bundle():
    include=[]
    for p in BASE.rglob('*'):
        if p.is_file() and 'logs/superseded' not in str(p):
            include.append((p, Path('B104E')/p.relative_to(BASE)))
    # Include prebundle manifests/events and readable manifest.
    for p in [ACTUAL_OUT,ACTUAL_MD,LEDGER_PRE,EVENTS_PRE,WORKBOOK]:
        include.append((p, Path('results')/p.name))
    with zipfile.ZipFile(BUNDLE,'w',compression=zipfile.ZIP_DEFLATED,compresslevel=9) as z:
        for src, arc in sorted(include,key=lambda x:str(x[1])):
            z.write(src, arcname=str(arc))
    # Test archive.
    with zipfile.ZipFile(BUNDLE) as z:
        bad=z.testzip()
        if bad is not None:
            raise RuntimeError(f'Bad ZIP member: {bad}')
    bh=sha256(BUNDLE)
    BUNDLE_HASH_FILE.write_text(f'{bh}  {BUNDLE.name}\n',encoding='utf-8')
    return bh


def finalize_ledgers(ledger_fields, ledger_rows, event_fields, event_rows, bundle_hash):
    row={k:'' for k in ledger_fields}
    row.update(record_type='generated_bundle',artifact_name=BUNDLE.name,local_path=str(BUNDLE),local_status='present_verified',origin_in_this_run='generated_by_B104E',analysis_role='frozen B104E reports, retained compact inputs, scripts, analyses, and prebundle provenance tables',used_by='delivery and future reproduction',direct_or_canonical_source_url='',source_page_url='',url_status='',retrieval_status='generated_locally',retrieved_at_utc=EVENT_TIME,size_bytes=str(BUNDLE.stat().st_size),sha256=bundle_hash,parent_or_derivation='B104E batch directory plus prebundle manifests/events',notes='ZIP integrity test passed. Internal source ledger is intentionally prebundle to avoid self-referential hashing; external final ledger contains this row.',batch_id=BATCH,deletion_state='retained',event_recorded_at_utc=EVENT_TIME,hash_authority='runtime SHA-256',runtime_verification_status='ZIP test passed',retained_derivative_paths=str(BUNDLE),raw_to_derived_reconciliation='all intended B104E files and prebundle provenance included')
    ledger_rows.append(row)
    write_csv(LEDGER_FINAL,ledger_fields,ledger_rows)
    write_ledger_md(LEDGER_MD,ledger_rows)
    event_rows.append({'event_time_utc':EVENT_TIME,'batch_id':BATCH,'event_type':'bundle_finalized','artifact_name':BUNDLE.name,'status':'passed','details':f'ZIP integrity passed; bytes={BUNDLE.stat().st_size}; sha256={bundle_hash}'})
    event_rows.append({'event_time_utc':EVENT_TIME,'batch_id':BATCH,'event_type':'deletion_gate_finalized','artifact_name':f'{UP_ZIP.name}; {GOA_ZIP.name}','status':'safe_to_delete_conversation_attachments','details':f'Frozen bundle={bundle_hash}; compact extracted inputs and provenance retained.'})
    write_csv(EVENTS_FINAL,event_fields,event_rows)


def final_validation(bundle_hash):
    checks=[]
    def ck(name,passed,details=''):
        checks.append({'check':name,'passed':bool(passed),'details':details})
    # Core findings.
    summary_fields, summary_rows=read_csv(BASE/'analysis/B104E_GOA_release158_169_validated_summary_20260829T121535Z.csv')
    by_rel={int(r['release']):r for r in summary_rows}
    ck('goa_12_releases',len(by_rel)==12,str(sorted(by_rel)))
    ck('goa_159_exact',int(by_rel[159]['total_mismatches'])==0 and int(by_rel[159]['exact_label_columns'])==121,str(by_rel[159]))
    ck('goa_159_unique_exact',sum(int(r['total_mismatches'])==0 for r in summary_rows)==1,'')
    ck('goa_duplicate_orientation_stable',set(r['duplicate_orientation'] for r in summary_rows)=={'001'},str(set(r['duplicate_orientation'] for r in summary_rows)))
    feature=json.load(open(BASE/'analysis/B104E_exact_MSigDB52_feature_generation_validation_20260829T121535Z.json'))
    independent=json.load(open(BASE/'analysis/B104E_independent_MSigDB52_feature_validation_20260829T121535Z.json'))
    ck('feature_main_exact',feature.get('total_resolved_feature_mismatches')==0 and feature.get('exact_feature_columns')==50,json.dumps(feature,sort_keys=True))
    ck('feature_independent_exact',independent.get('mismatches')==0,json.dumps(independent,sort_keys=True))
    vers_fields,vers=read_csv(BASE/'analysis/B104E_MSigDB_feature_version_screen_PROVENANCE_SAFE_20260829T121535Z.csv')
    ck('four_msigdb_versions',len(vers)==4,str([r.get('version') for r in vers]))
    ck('all_versions_exact',all(int(r['total_mismatches'])==0 and int(r['exact_columns'])==50 for r in vers),str(vers))
    up_fields,up=read_csv(BASE/'analysis/B104E_UniProt_O95073_Q9Y620_flatfile_record_comparison_20260829T121535Z.csv')
    ck('six_uniprot_records',len(up)==6,str(len(up)))
    ck('O95073_dual_geneids',all(r['geneids']=='100861412|25788' for r in up if r['primary_accession']=='O95073'),str([r for r in up if r['primary_accession']=='O95073']))
    ck('Q9Y620_25788',all(r['geneids']=='25788' for r in up if r['primary_accession']=='Q9Y620'),str([r for r in up if r['primary_accession']=='Q9Y620']))
    ck('uploaded_package_validation',json.load(open(BASE/'analysis/B104E_uploaded_audit_archives_validation_20260829T121535Z.json')).get('all_checks_passed') is True,'')
    # Files/manifests/bundle.
    for p in [ACTUAL_OUT,ACTUAL_MD,LEDGER_FINAL,LEDGER_MD,EVENTS_FINAL,CLEARANCE,CHECKSUMS_PRE,WORKBOOK,BUNDLE,BUNDLE_HASH_FILE]:
        ck(f'file_exists:{p.name}',p.exists(),str(p))
    ck('bundle_sha256',sha256(BUNDLE)==bundle_hash,sha256(BUNDLE))
    try:
        with zipfile.ZipFile(BUNDLE) as z: bad=z.testzip()
    except Exception as e:
        bad=str(e)
    ck('bundle_integrity',bad is None,str(bad))
    ck('input_zip_sha_uniprot',sha256(UP_ZIP)=='4ac8cb1a900215ded9dc35a4fc44a4abaaff3e5e774cf62769592f1ab153a7b0',sha256(UP_ZIP))
    ck('input_zip_sha_goa',sha256(GOA_ZIP)=='502c8ffdb7b809c1665e82d31db52b72e7e855e7e9fcc06a8a2f46a64bc30de9',sha256(GOA_ZIP))
    obj={'created_at_utc':EVENT_TIME,'batch_id':BATCH,'all_checks_passed':all(x['passed'] for x in checks),'checks':checks,'bundle':{'path':str(BUNDLE),'size_bytes':BUNDLE.stat().st_size,'sha256':bundle_hash}}
    VALIDATION.write_text(json.dumps(obj,indent=2,sort_keys=True),encoding='utf-8')
    if not obj['all_checks_passed']:
        raise RuntimeError('Final validation failed')
    return obj


def create_final_checksums():
    paths=[]
    for p in [ACTUAL_OUT,ACTUAL_MD,LEDGER_FINAL,LEDGER_MD,EVENTS_FINAL,CLEARANCE,CHECKSUMS_PRE,WORKBOOK,BUNDLE,BUNDLE_HASH_FILE,VALIDATION]:
        paths.append(p)
    for p in sorted(BASE.rglob('*')):
        if p.is_file() and 'logs/superseded' not in str(p):
            paths.append(p)
    seen=set(); rows=[]
    for p in paths:
        rp=str(p)
        if rp in seen: continue
        seen.add(rp)
        rows.append({'path':rp,'size_bytes':p.stat().st_size,'sha256':sha256(p)})
    write_csv(FINAL_CHECKSUMS,['path','size_bytes','sha256'],rows)


def main():
    actual_fields, actual_rows=make_actual_manifest()
    ledger_fields, ledger_rows=make_source_ledger(actual_rows)
    event_fields, event_rows=make_events()
    write_clearance()
    checksum_rows=create_pre_checksums()
    create_workbook(actual_rows,ledger_rows,event_rows,checksum_rows,WORKBOOK)
    bundle_hash=create_bundle()
    finalize_ledgers(ledger_fields,ledger_rows,event_fields,event_rows,bundle_hash)
    final_validation(bundle_hash)
    create_final_checksums()
    print(json.dumps({
        'actual_manifest':str(ACTUAL_OUT),
        'source_ledger':str(LEDGER_FINAL),
        'events':str(EVENTS_FINAL),
        'clearance':str(CLEARANCE),
        'validation':str(VALIDATION),
        'workbook':str(WORKBOOK),
        'bundle':str(BUNDLE),
        'bundle_sha256':bundle_hash,
        'bundle_size':BUNDLE.stat().st_size,
        'final_checksums':str(FINAL_CHECKSUMS)
    },indent=2))

if __name__=='__main__':
    main()
