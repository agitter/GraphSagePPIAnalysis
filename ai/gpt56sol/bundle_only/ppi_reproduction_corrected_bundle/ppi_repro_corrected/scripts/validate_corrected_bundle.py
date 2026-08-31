#!/usr/bin/env python3
from __future__ import annotations
import csv, hashlib, json, re, sys
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path('/mnt/data/ppi_repro_corrected')
R=ROOT/'results'

def h(p):
    x=hashlib.sha256()
    with p.open('rb') as f:
        for c in iter(lambda:f.read(1<<20),b''): x.update(c)
    return x.hexdigest()

checks=[]
def check(name, ok, detail=''):
    checks.append({'check':name,'pass':bool(ok),'detail':detail})
    if not ok: print('FAIL',name,detail,file=sys.stderr)

actual=list(csv.DictReader((R/'actual_input_file_manifest.csv').open(encoding='utf-8')))
ledger=list(csv.DictReader((R/'source_ledger.csv').open(encoding='utf-8')))
check('actual_manifest_row_count',len(actual)==18,str(len(actual)))
check('actual_manifest_only_actual_inputs',all(r['record_type']=='actual_input' for r in actual))
public=[r for r in actual if r['artifact_name'] not in {'investigation_summary_2026_08_23.md','historical_go_mapping_inventory.md'}]
check('public_actual_inputs_have_urls',all(r['direct_or_canonical_source_url'] for r in public),','.join(r['artifact_name'] for r in public if not r['direct_or_canonical_source_url']))
missing=[]; bad=[]
for r in actual:
    p=Path(r['local_path'])
    if not p.exists(): missing.append(r['artifact_name']); continue
    if r['sha256'] and h(p)!=r['sha256']: bad.append(r['artifact_name'])
check('actual_input_paths_exist',not missing,','.join(missing))
check('actual_input_hashes_match',not bad,','.join(bad))
check('source_ledger_has_historical_candidates',sum(r['record_type']=='historical_candidate_not_materialized' for r in ledger)>=25)
check('all_historical_candidates_have_urls',all(r['direct_or_canonical_source_url'] for r in ledger if r['record_type']=='historical_candidate_not_materialized'))
for x in ['actual_input_file_manifest.xlsx','source_ledger.xlsx']:
    try:
        wb=load_workbook(R/x, read_only=True, data_only=True)
        n=wb.active.max_row
        check(f'{x}_opens',True,f'{n} rows')
    except Exception as e:
        check(f'{x}_opens',False,repr(e))
report=(R/'MASTER_REPRODUCTION_REPORT.md').read_text(encoding='utf-8')
diag=(R/'EXECUTION_DIAGNOSTICS.md').read_text(encoding='utf-8')
check('master_contains_execution_diagnostics_section','## 10. Execution diagnostics' in report)
check('master_contains_exact_superseded_errors','IndexError: list index out of range' in report and 'ValueError: No candidate records for MSigDB 5.1 C5 BP' in report)
check('diagnostics_contains_final_and_superseded','## Final successful runs' in diag and '## Superseded attempts and errors' in diag)
check('no_dangling_old_error_phrase','see error section' not in report.lower())
status=(R/'RUN_STATUS.md').read_text(encoding='utf-8')
check('run_status_no_missing_output','| MISSING |' not in status)
required=['MASTER_REPRODUCTION_REPORT.md','EXECUTION_DIAGNOSTICS.md','SOURCE_ACQUISITION.md','actual_input_file_manifest.csv','source_ledger.csv','core_verification_summary.json','local_label_source_screen_summary.json']
check('required_outputs_exist',all((R/x).exists() for x in required),','.join(x for x in required if not (R/x).exists()))
# Verify direct report artifact references that are expected under results.
refs=set(re.findall(r'`([^`]+\.(?:csv|json|md|xlsx|stdout|stderr))`',report+diag))
ignore={'scripts/download_or_verify_sources.py'}
missing_refs=[]
input_paths={Path(r['local_path']).name: Path(r['local_path']) for r in actual if r.get('local_path')}
for ref in refs:
    if '/' in ref:
        continue
    if not (R/ref).exists() and not (ref in input_paths and input_paths[ref].exists()):
        missing_refs.append(ref)
check('report_artifact_references_exist',not missing_refs,','.join(sorted(missing_refs)))
res={'status':'PASS' if all(c['pass'] for c in checks) else 'FAIL','checks':checks}
(R/'bundle_validation.json').write_text(json.dumps(res,indent=2),encoding='utf-8')
print(json.dumps(res,indent=2))
return_code=0 if res['status']=='PASS' else 1
raise SystemExit(return_code)
