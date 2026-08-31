#!/usr/bin/env python3
from __future__ import annotations

import csv
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path('/mnt/data')
BASE = ROOT / 'ppi_repro_corrected'
OUT = BASE / 'results'


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open('rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def actual_row(filename, role, source_url, source_page_url, used_by, notes='', url_status='official_or_canonical_url_recorded'):
    path = ROOT / filename
    present = path.exists()
    return {
        'record_type': 'actual_input',
        'artifact_name': filename,
        'local_path': str(path) if present else '',
        'local_status': 'present' if present else 'missing',
        'origin_in_this_run': 'user_upload' if present else 'not_present',
        'analysis_role': role,
        'used_by': used_by,
        'direct_or_canonical_source_url': source_url,
        'source_page_url': source_page_url,
        'url_status': url_status,
        'retrieval_status': 'supplied_by_user; not downloaded by corrected run' if present else 'not available',
        'retrieved_at_utc': '',
        'size_bytes': path.stat().st_size if present else '',
        'sha256': sha256(path) if present else '',
        'parent_or_derivation': '',
        'notes': notes,
    }


def candidate_row(name, role, url, page='', notes='', url_status='exact_candidate_url_recorded'):
    return {
        'record_type': 'historical_candidate_not_materialized',
        'artifact_name': name,
        'local_path': '',
        'local_status': 'not_materialized',
        'origin_in_this_run': 'web_research_or_prior_inventory',
        'analysis_role': role,
        'used_by': 'planned GO source grid; not used in current numeric results',
        'direct_or_canonical_source_url': url,
        'source_page_url': page,
        'url_status': url_status,
        'retrieval_status': 'not materialized in runtime',
        'retrieved_at_utc': '',
        'size_bytes': '',
        'sha256': '',
        'parent_or_derivation': '',
        'notes': notes,
    }


def web_ref(name, role, url, notes=''):
    return {
        'record_type': 'web_reference',
        'artifact_name': name,
        'local_path': '',
        'local_status': 'web_reference_only',
        'origin_in_this_run': 'web_research',
        'analysis_role': role,
        'used_by': 'documentary verification',
        'direct_or_canonical_source_url': url,
        'source_page_url': url,
        'url_status': 'official_page_or_repository',
        'retrieval_status': 'inspected online; no local byte copy',
        'retrieved_at_utc': datetime.now(timezone.utc).isoformat(),
        'size_bytes': '',
        'sha256': '',
        'parent_or_derivation': '',
        'notes': notes,
    }


def generated_row(path: Path, role, parents, notes=''):
    return {
        'record_type': 'generated_output',
        'artifact_name': path.name,
        'local_path': str(path),
        'local_status': 'present' if path.exists() else 'missing',
        'origin_in_this_run': 'generated',
        'analysis_role': role,
        'used_by': 'report and reproducibility bundle',
        'direct_or_canonical_source_url': '',
        'source_page_url': '',
        'url_status': 'not_applicable_generated_file',
        'retrieval_status': 'generated locally',
        'retrieved_at_utc': datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).isoformat() if path.exists() else '',
        'size_bytes': path.stat().st_size if path.exists() else '',
        'sha256': sha256(path) if path.exists() else '',
        'parent_or_derivation': parents,
        'notes': notes,
    }


def build_rows():
    rows = []
    rows += [
        actual_row('graphsage_ppi.zip', 'primary dataset under investigation',
                   'https://snap.stanford.edu/graphsage/ppi.zip',
                   'https://snap.stanford.edu/graphsage/',
                   'run_core_verification.py',
                   'Official GraphSAGE project page links this preprocessed PPI archive.'),
        actual_row('dgl_ppi.zip', 'downstream DGL dataset to reproduce',
                   'https://data.dgl.ai/dataset/ppi.zip',
                   'https://github.com/dmlc/dgl/blob/master/python/dgl/data/ppi.py',
                   'run_core_verification.py',
                   'Direct URL used by the DGL dataset loader; supplied archive was hash-checked locally.'),
        actual_row('bio-tissue-networks.tar.gz', 'OhmNet tissue PPI source candidate',
                   'https://snap.stanford.edu/ohmnet/bio-tissue-networks.tar.gz',
                   'https://snap.stanford.edu/ohmnet/',
                   'run_core_verification.py',
                   'Official OhmNet project download.'),
        actual_row('bio-tissue-labels.tar.gz', 'OhmNet tissue-specific GO labels',
                   'https://snap.stanford.edu/ohmnet/bio-tissue-labels.tar.gz',
                   'https://snap.stanford.edu/ohmnet/',
                   'run_local_label_source_screen.py',
                   'Official OhmNet project download.'),
        actual_row('bio-tissue-hierarchy.tar.gz', 'OhmNet tissue hierarchy; retained for later work',
                   'https://snap.stanford.edu/ohmnet/bio-tissue-hierarchy.tar.gz',
                   'https://snap.stanford.edu/ohmnet/',
                   'not used in current numeric checks',
                   'Official OhmNet project download.'),
        actual_row('bio-tissue-readme.txt', 'OhmNet data documentation',
                   'https://snap.stanford.edu/ohmnet/bio-tissue-readme.txt',
                   'https://snap.stanford.edu/ohmnet/',
                   'report documentation',
                   'States Entrez Gene IDs and August-September 2016 collection date.'),
        actual_row('msigdb_v5.1_files_to_download_locally.zip', 'historical MSigDB comparison',
                   'https://www.gsea-msigdb.org/gsea/downloads.jsp',
                   'https://www.gsea-msigdb.org/gsea/downloads.jsp',
                   'run_local_label_source_screen.py',
                   'MSigDB direct archive links require an authenticated account; the official download page is recorded instead.',
                   'official_source_page; authenticated direct URL unavailable'),
        actual_row('msigdb_v5.2_files_to_download_locally.zip', 'feature recovery and historical label comparison',
                   'https://www.gsea-msigdb.org/gsea/downloads.jsp',
                   'https://www.gsea-msigdb.org/gsea/downloads.jsp',
                   'run_core_verification.py; run_local_label_source_screen.py',
                   'MSigDB direct archive links require an authenticated account; archive version and hash identify the supplied bytes.',
                   'official_source_page; authenticated direct URL unavailable'),
        actual_row('msigdb_v5.2_chip_files_to_download_locally.zip', 'historical chip mappings; retained for later GO work',
                   'https://www.gsea-msigdb.org/gsea/downloads.jsp',
                   'https://www.gsea-msigdb.org/gsea/downloads.jsp',
                   'not used in current numeric checks',
                   'MSigDB authenticated download.',
                   'official_source_page; authenticated direct URL unavailable'),
        actual_row('msigdb_v6.0_files_to_download_locally.zip', 'later MSigDB comparison',
                   'https://www.gsea-msigdb.org/gsea/downloads.jsp',
                   'https://www.gsea-msigdb.org/gsea/downloads.jsp',
                   'run_local_label_source_screen.py',
                   'MSigDB authenticated download.',
                   'official_source_page; authenticated direct URL unavailable'),
        actual_row('Greene2015.pdf', 'Greene et al. manuscript',
                   'https://doi.org/10.1038/ng.3259',
                   'https://www.nature.com/articles/ng.3259',
                   'documentary verification',
                   'User-supplied publisher PDF.'),
        actual_row('Greene2015_sup.pdf', 'Greene et al. supplementary information',
                   'https://www.nature.com/articles/ng.3259#Sec23',
                   'https://www.nature.com/articles/ng.3259',
                   'documentary verification',
                   'User-supplied supplementary PDF; source page recorded because publisher asset URLs may change.'),
        actual_row('Greene2015_Table6.xlsx', 'expert-curated GO term list',
                   'https://www.nature.com/articles/ng.3259#Sec23',
                   'https://www.nature.com/articles/ng.3259',
                   'run_local_label_source_screen.py',
                   'Supplementary Table 6; 973 GO IDs after parsing.'),
        actual_row('Greene2015_Table9.xlsx', 'GO-to-tissue mapping',
                   'https://www.nature.com/articles/ng.3259#Sec23',
                   'https://www.nature.com/articles/ng.3259',
                   'run_local_label_source_screen.py',
                   'Supplementary Table 9; 6,172 unique GO IDs after parsing.'),
        actual_row('OhmNet.pdf', 'OhmNet manuscript',
                   'https://doi.org/10.1093/bioinformatics/btx252',
                   'https://academic.oup.com/bioinformatics/article/33/14/i190/3953967',
                   'documentary verification',
                   'User-supplied manuscript PDF.'),
        actual_row('investigation_summary_2026_08_23.md', 'prior hypotheses to independently test',
                   '', '', 'hypothesis list only; never used as mapping input',
                   'No external URL was supplied.', 'no_external_url_supplied'),
        actual_row('Pasted markdown(1).md', 'copied EBI GOA archive listing',
                   'https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/',
                   'https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/',
                   'historical source discovery only',
                   'Local file is a pasted listing, not an archive data file.'),
        actual_row('historical_go_mapping_inventory.md', 'prior source-discovery inventory',
                   '', '', 'historical source discovery only',
                   'Generated in an earlier turn; individual underlying URLs are represented below.', 'derived_local_artifact'),
    ]

    rows += [
        web_ref('GraphSAGE project page', 'official dataset provenance', 'https://snap.stanford.edu/graphsage/'),
        web_ref('GraphSAGE repository', 'input format and code provenance', 'https://github.com/williamleif/GraphSAGE'),
        web_ref('GraphSAGE paper', 'manuscript', 'https://arxiv.org/abs/1706.02216'),
        web_ref('OhmNet project page', 'official data provenance', 'https://snap.stanford.edu/ohmnet/'),
        web_ref('OhmNet repository', 'algorithm code provenance', 'https://github.com/mims-harvard/ohmnet'),
        web_ref('DGL current PPI loader source', 'downstream loader behavior', 'https://github.com/dmlc/dgl/blob/master/python/dgl/data/ppi.py'),
        web_ref('DGL PR 395', 'historical introduction of PPI dataset loader', 'https://github.com/dmlc/dgl/pull/395'),
        web_ref('EBI GOA old/HUMAN index', 'historical GOA archive index', 'https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/'),
    ]

    ebi = 'https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/'
    for release, date in [(155, '2016-03-14'), (156, '2016-04-11'), (157, '2016-05-09')]:
        rows.append(candidate_row(f'gene_association.goa_human.{release}.gz', 'historical human GAF candidate', ebi + f'gene_association.goa_human.{release}.gz', ebi, date))
        rows.append(candidate_row(f'gp_association.goa_human.{release}.gz', 'historical human GPAD candidate', ebi + f'gp_association.goa_human.{release}.gz', ebi, date))
        rows.append(candidate_row(f'gp_information.goa_human.{release}.gz', 'historical human GPI predecessor and ID metadata candidate', ebi + f'gp_information.goa_human.{release}.gz', ebi, date))
    for release, date in [(158, '2016-06-07'), (159, '2016-07-04'), (160, '2016-09-14')]:
        rows.append(candidate_row(f'goa_human.gaf.{release}.gz', 'historical human GAF candidate', ebi + f'goa_human.gaf.{release}.gz', ebi, date))
        rows.append(candidate_row(f'goa_human.gpa.{release}.gz', 'historical human GPAD candidate', ebi + f'goa_human.gpa.{release}.gz', ebi, date))
        rows.append(candidate_row(f'goa_human.gpi.{release}.gz', 'historical human GPI and ID metadata candidate', ebi + f'goa_human.gpi.{release}.gz', ebi, date))

    go_release = 'https://release.geneontology.org/2016-06-01/'
    rows += [
        candidate_row('2016-06-01-go.obo', 'historical GO ontology', go_release + 'ontology/go.obo', go_release),
        candidate_row('2016-06-01-go-basic.obo', 'historical GO ontology restricted to safe relations', go_release + 'ontology/go-basic.obo', go_release),
        candidate_row('2016-06-01-gp2protein.geneid.gz', 'Entrez GeneID to UniProt mapping candidate', go_release + 'annotations/gp2protein/gp2protein.geneid.gz', go_release),
    ]

    bioc = 'https://bioconductor.statistik.tu-dortmund.de/packages/{ver}/data/annotation/src/contrib/{file}'
    for ver, package, dates in [
        ('3.1', 'org.Hs.eg.db_3.1.2.tar.gz', 'Entrez 2015-03-17; GO 2015-03-14'),
        ('3.2', 'org.Hs.eg.db_3.2.3.tar.gz', 'Entrez 2015-09-27; GO 2015-09-19'),
        ('3.3', 'org.Hs.eg.db_3.3.0.tar.gz', 'Entrez 2016-03-14; GO 2016-03-05'),
        ('3.4', 'org.Hs.eg.db_3.4.0.tar.gz', 'Entrez 2016-09-26; GO 2016-09-21'),
    ]:
        rows.append(candidate_row(package, 'historical Entrez-to-UniProt and Entrez-to-GO package', bioc.format(ver=ver, file=package), f'https://bioconductor.org/packages/{ver}/data/annotation/html/org.Hs.eg.db.html', dates))

    rows += [
        candidate_row('idmapping_selected.tab.2015_03.gz', 'historical UniProt-wide mapping with GeneID and GO fields',
                      'https://ftp.uniprot.org/pub/databases/uniprot/current_release/knowledgebase/idmapping/idmapping_selected.tab.2015_03.gz',
                      'https://ftp.uniprot.org/pub/databases/uniprot/current_release/knowledgebase/idmapping/'),
        candidate_row('idmapping.dat.2015_03.gz', 'historical UniProt long-form identifier mapping',
                      'https://ftp.uniprot.org/pub/databases/uniprot/current_release/knowledgebase/idmapping/idmapping.dat.2015_03.gz',
                      'https://ftp.uniprot.org/pub/databases/uniprot/current_release/knowledgebase/idmapping/'),
        candidate_row('NCBI current gene2go.gz', 'current rolling reference; not a historical substitute',
                      'https://ftp.ncbi.nlm.nih.gov/gene/DATA/gene2go.gz', 'https://ftp.ncbi.nlm.nih.gov/gene/DATA/'),
        candidate_row('NCBI current gene2accession.gz', 'current rolling GeneID-to-accession mapping',
                      'https://ftp.ncbi.nlm.nih.gov/gene/DATA/gene2accession.gz', 'https://ftp.ncbi.nlm.nih.gov/gene/DATA/'),
        candidate_row('NCBI current gene2refseq.gz', 'current rolling GeneID-to-RefSeq mapping',
                      'https://ftp.ncbi.nlm.nih.gov/gene/DATA/gene2refseq.gz', 'https://ftp.ncbi.nlm.nih.gov/gene/DATA/'),
        candidate_row('NCBI current gene_refseq_uniprotkb_collab.gz', 'current RefSeq-to-UniProt collaboration map',
                      'https://ftp.ncbi.nlm.nih.gov/gene/DATA/gene_refseq_uniprotkb_collab.gz', 'https://ftp.ncbi.nlm.nih.gov/gene/DATA/'),
    ]

    for path, role, parents in [
        (OUT / 'core_verification_summary.json', 'computed core verification summary', 'GraphSAGE, DGL, OhmNet networks, MSigDB 5.2'),
        (OUT / 'local_label_source_screen_summary.json', 'computed local label-source screen summary', 'collapsed labels, MSigDB 5.1/5.2/6.0, OhmNet labels, Greene Tables 6/9'),
        (OUT / 'tissue_partition.csv', 'computed 24-tissue partition', 'GraphSAGE and OhmNet network archives'),
        (OUT / 'feature_column_mapping.csv', 'computed feature identities', 'GraphSAGE and MSigDB 5.2'),
        (OUT / 'dgl_transformation_verification.json', 'computed DGL transformation checks', 'GraphSAGE and DGL archives'),
        (OUT / 'graphsage_row_to_entrez_topology_features.csv', 'computed row-to-Entrez mapping', 'GraphSAGE, OhmNet networks, MSigDB 5.2'),
    ]:
        rows.append(generated_row(path, role, parents))
    return rows


def save_xlsx(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sources'
    fields = list(rows[0].keys())
    ws.append(fields)
    header_fill = PatternFill('solid', fgColor='1F4E78')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    imported_font = Font(color='008000')
    static_font = Font(color='666666')
    caution_fill = PatternFill('solid', fgColor='FCE4D6')
    for row_idx, record in enumerate(rows, start=2):
        ws.append([record[f] for f in fields])
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.font = imported_font if field in {'direct_or_canonical_source_url', 'source_page_url'} else static_font
            if field in {'direct_or_canonical_source_url', 'source_page_url'} and isinstance(cell.value, str) and cell.value.startswith('http'):
                cell.hyperlink = cell.value
                cell.style = 'Hyperlink'
        if record['local_status'] in {'missing', 'not_materialized'}:
            for cell in ws[row_idx]:
                cell.fill = caution_fill
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    widths = {
        'A': 31, 'B': 44, 'C': 52, 'D': 22, 'E': 24, 'F': 42, 'G': 42,
        'H': 58, 'I': 50, 'J': 34, 'K': 42, 'L': 25, 'M': 16, 'N': 68, 'O': 70,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 34

    summary = wb.create_sheet('Summary')
    summary.append(['Metric', 'Value'])
    for c in summary[1]:
        c.fill = header_fill
        c.font = Font(color='FFFFFF', bold=True)
    metrics = [
        ('Generated at UTC', datetime.now(timezone.utc).isoformat()),
        ('Total records', len(rows)),
        ('Actual input records', sum(r['record_type'] == 'actual_input' for r in rows)),
        ('Present actual inputs', sum(r['record_type'] == 'actual_input' and r['local_status'] == 'present' for r in rows)),
        ('Historical candidates not materialized', sum(r['record_type'] == 'historical_candidate_not_materialized' for r in rows)),
        ('Web references', sum(r['record_type'] == 'web_reference' for r in rows)),
        ('Generated outputs', sum(r['record_type'] == 'generated_output' for r in rows)),
        ('URL policy', 'Direct file URL where public; official source page where authenticated or publisher-controlled.'),
    ]
    for metric in metrics:
        summary.append(metric)
    summary.column_dimensions['A'].width = 38
    summary.column_dimensions['B'].width = 100
    summary.freeze_panes = 'A2'
    wb.save(path)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    rows = build_rows()
    fields = list(rows[0].keys())
    csv_path = OUT / 'actual_input_file_manifest_v2.csv'
    with csv_path.open('w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
    (OUT / 'actual_input_file_manifest_v2.json').write_text(json.dumps(rows, indent=2))
    md = [
        '# Corrected source and input manifest', '',
        'This ledger distinguishes actual local inputs, web references, historical candidates that were not materialized, and generated outputs. A URL is recorded for every public source whose location is known. MSigDB archives use the official authenticated download page rather than inventing a session-specific direct link.', '',
        '| Type | Artifact | Local status | SHA-256 | Direct/canonical URL | Source page | Used by |',
        '|---|---|---|---|---|---|---|',
    ]
    for r in rows:
        md.append(f"| {r['record_type']} | `{r['artifact_name']}` | {r['local_status']} | `{r['sha256']}` | {r['direct_or_canonical_source_url']} | {r['source_page_url']} | {r['used_by']} |")
    (OUT / 'actual_input_file_manifest_v2.md').write_text('\n'.join(md))
    save_xlsx(rows, OUT / 'actual_input_file_manifest_v2.xlsx')

    # Validate saved workbook and CSV row counts.
    wb = load_workbook(OUT / 'actual_input_file_manifest_v2.xlsx', read_only=True, data_only=True)
    assert wb['Sources'].max_row == len(rows) + 1
    with csv_path.open() as f:
        assert sum(1 for _ in csv.DictReader(f)) == len(rows)
    print(json.dumps({
        'status': 'PASS',
        'records': len(rows),
        'actual_inputs': sum(r['record_type'] == 'actual_input' for r in rows),
        'public_source_urls_recorded': sum(bool(r['direct_or_canonical_source_url']) for r in rows),
        'csv': str(csv_path),
        'xlsx': str(OUT / 'actual_input_file_manifest_v2.xlsx'),
    }, indent=2))


if __name__ == '__main__':
    main()
