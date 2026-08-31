#!/usr/bin/env python3
from __future__ import annotations

import argparse
import collections
import csv
import json
import re
import tarfile
import time
import zipfile
from pathlib import Path
from typing import Dict, List, Mapping, Sequence, Set, Tuple

import numpy as np
from openpyxl import load_workbook


def write_csv(path: Path, rows: Sequence[Mapping[str, object]], fieldnames=None):
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def load_collapsed(path: Path):
    with path.open() as fh:
        r = csv.DictReader(fh)
        genes = []
        vals = []
        for row in r:
            genes.append(int(row["entrez_gene_id"]))
            vals.append([int(row[f"label_{i}"]) for i in range(121)])
    return genes, np.asarray(vals, dtype=np.uint8)


def bitsets_from_labels(y: np.ndarray) -> List[int]:
    out = []
    for col in range(y.shape[1]):
        bit = 0
        for i, value in enumerate(y[:, col]):
            if value:
                bit |= 1 << i
        out.append(bit)
    return out


def parse_gmt_archive(path: Path, version: str, gene_index: Mapping[int, int]):
    dedup = {}
    with zipfile.ZipFile(path) as zf:
        for member in zf.namelist():
            if not member.lower().endswith(".entrez.gmt"):
                continue
            base = Path(member).name
            low = base.lower()
            collection = next((p[:-1].upper() for p in ["c1.", "c2.", "c3.", "c4.", "c5.", "c6.", "c7."] if low.startswith(p)), "ALL")
            subcollection = "BP" if low.startswith("c5.bp.") else ""
            text = zf.read(member).decode("utf-8", "replace")
            for line in text.splitlines():
                f = line.split("\t")
                if len(f) < 3:
                    continue
                full_genes = frozenset(int(x) for x in f[2:] if x.isdigit())
                key = (f[0], full_genes)
                bit = 0
                for gene in full_genes:
                    idx = gene_index.get(gene)
                    if idx is not None:
                        bit |= 1 << idx
                rec = {
                    "version": version,
                    "source_member": member,
                    "source_basename": base,
                    "collection": collection,
                    "subcollection": subcollection,
                    "name": f[0],
                    "description": f[1],
                    "bit": bit,
                    "support": bit.bit_count(),
                }
                previous = dedup.get(key)
                if previous is None or (previous["collection"] == "ALL" and collection != "ALL") or (not previous.get("subcollection") and subcollection):
                    dedup[key] = rec
    return list(dedup.values())


def parse_go_ids_from_msigdb_xml(path: Path):
    # Stream line-by-line because every GENESET record is one XML line in these archives.
    name_to_go = {}
    attr_name = re.compile(r'STANDARD_NAME="([^"]+)"')
    attr_cat = re.compile(r'CATEGORY_CODE="([^"]+)"')
    attr_sub = re.compile(r'SUB_CATEGORY_CODE="([^"]+)"')
    go_re = re.compile(r'GO:\d{7}')
    with zipfile.ZipFile(path) as zf:
        xml_members = [n for n in zf.namelist() if n.lower().endswith(".xml")]
        if not xml_members:
            return name_to_go
        with zf.open(xml_members[0]) as fh:
            for raw in fh:
                line = raw.decode("utf-8", "replace")
                if "<GENESET " not in line or 'CATEGORY_CODE="C5"' not in line:
                    continue
                mname = attr_name.search(line)
                msub = attr_sub.search(line)
                if not mname or not msub or msub.group(1) != "BP":
                    continue
                gos = go_re.findall(line)
                if gos:
                    name_to_go[mname.group(1)] = gos[0]
    return name_to_go


def metrics(obs: int, pred: int, n: int):
    xor = (obs ^ pred).bit_count()
    tp = (obs & pred).bit_count()
    fp = (pred & ~obs).bit_count()
    fn = (obs & ~pred).bit_count()
    denom = 2 * tp + fp + fn
    f1 = 2 * tp / denom if denom else 1.0
    jac_denom = tp + fp + fn
    jaccard = tp / jac_denom if jac_denom else 1.0
    return xor, 1 - xor / n, tp, fp, fn, f1, jaccard


def best_matches(label_bits: Sequence[int], records: Sequence[dict], n: int, source_type: str, go_by_name=None):
    rows = []
    if not records:
        raise ValueError(f"No candidate records for {source_type}")
    for col, obs in enumerate(label_bits):
        best_xor = n + 1
        best = []
        for rec in records:
            x = (obs ^ rec["bit"]).bit_count()
            if x < best_xor:
                best_xor = x
                best = [rec]
            elif x == best_xor:
                best.append(rec)
        rec = best[0]
        xor, agreement, tp, fp, fn, f1, jaccard = metrics(obs, rec["bit"], n)
        rows.append({
            "source_type": source_type,
            "version": rec.get("version", ""),
            "label_column": col,
            "observed_positive_genes": obs.bit_count(),
            "best_name": rec.get("name", ""),
            "best_go_id": (go_by_name or {}).get(rec.get("name", ""), rec.get("go_id", "")),
            "best_collection": rec.get("collection", ""),
            "best_subcollection": rec.get("subcollection", ""),
            "best_source_member": rec.get("source_member", ""),
            "best_candidate_positive_genes": rec["bit"].bit_count(),
            "mismatch_genes": xor,
            "agreement": agreement,
            "true_positives": tp,
            "false_positives": fp,
            "false_negatives": fn,
            "f1": f1,
            "jaccard": jaccard,
            "tie_count": len(best),
            "tied_names": "|".join(r.get("name", "") for r in best[:25]),
        })
    return rows


def load_mapping(path: Path):
    out = {}
    with path.open() as fh:
        for row in csv.DictReader(fh):
            out[int(row["graphsage_row"])] = int(row["entrez_gene_id"])
    return out


def load_graphsage_labels(graphsage_dir: Path):
    cm = json.loads((graphsage_dir / "ppi" / "ppi-class_map.json").read_text())
    return np.asarray([cm[str(i)] for i in range(len(cm))], dtype=np.uint8)


def parse_ohmnet_labels(extracted: Path, gene_index: Mapping[int, int], selected_tissues: Set[str]):
    individual = []
    union_all: Dict[str, Set[int]] = collections.defaultdict(set)
    union_selected: Dict[str, Set[int]] = collections.defaultdict(set)
    file_data = []
    for path in sorted(extracted.rglob("*.lab")):
        m = re.match(r"(.+)_GO:(\d{7})\.lab$", path.name)
        if not m:
            continue
        tissue = m.group(1)
        go_id = "GO:" + m.group(2)
        positives = set()
        represented = set()
        with path.open() as fh:
            for line in fh:
                if not line.strip() or line.startswith("#"):
                    continue
                parts = line.split()
                if len(parts) < 2:
                    continue
                gene = int(parts[0])
                represented.add(gene)
                if int(parts[1]) == 1:
                    positives.add(gene)
        union_all[go_id].update(positives)
        if tissue in selected_tissues:
            union_selected[go_id].update(positives)
        bit = 0
        for gene in positives:
            idx = gene_index.get(gene)
            if idx is not None:
                bit |= 1 << idx
        rec = {
            "version": "OhmNet supplied archive",
            "name": f"{tissue}_{go_id}",
            "go_id": go_id,
            "tissue": tissue,
            "collection": "individual_tissue_file",
            "subcollection": "",
            "source_member": path.name,
            "bit": bit,
            "represented_gene_count": len(represented),
        }
        individual.append(rec)
        file_data.append((tissue, go_id, positives, represented, path.name))

    union_all_records = []
    for go_id, positives in union_all.items():
        bit = 0
        for gene in positives:
            idx = gene_index.get(gene)
            if idx is not None:
                bit |= 1 << idx
        union_all_records.append({
            "version": "OhmNet supplied archive",
            "name": go_id,
            "go_id": go_id,
            "collection": "union_all_ohmnet_tissues",
            "subcollection": "",
            "source_member": "bio-tissue-labels.tar.gz",
            "bit": bit,
        })
    union_selected_records = []
    for go_id, positives in union_selected.items():
        bit = 0
        for gene in positives:
            idx = gene_index.get(gene)
            if idx is not None:
                bit |= 1 << idx
        union_selected_records.append({
            "version": "OhmNet supplied archive",
            "name": go_id,
            "go_id": go_id,
            "collection": "union_selected_24_tissues",
            "subcollection": "",
            "source_member": "bio-tissue-labels.tar.gz",
            "bit": bit,
        })
    return individual, union_all_records, union_selected_records, file_data


def tissue_specific_screen(
    file_data,
    tissue_partition_csv: Path,
    row_to_gene: Mapping[int, int],
    gs_labels: np.ndarray,
):
    partitions = []
    with tissue_partition_csv.open() as fh:
        for row in csv.DictReader(fh):
            partitions.append({
                "tissue": row["ohmnet_tissue"],
                "start": int(row["row_start_inclusive"]),
                "end": int(row["row_end_exclusive"]),
            })
    by_tissue = {x["tissue"]: x for x in partitions}
    best_by_col = {col: None for col in range(gs_labels.shape[1])}
    for tissue, go_id, positives, represented, member in file_data:
        block = by_tissue.get(tissue)
        if block is None:
            continue
        selected_rows = np.asarray([
            r for r in range(block["start"], block["end"])
            if r in row_to_gene and row_to_gene[r] in represented
        ], dtype=np.int64)
        if selected_rows.size == 0:
            continue
        pred = np.asarray([1 if row_to_gene[int(r)] in positives else 0 for r in selected_rows], dtype=np.uint8)
        truth = gs_labels[selected_rows, :]
        pred2 = pred[:, None]
        mismatch = np.count_nonzero(truth != pred2, axis=0)
        tp = np.count_nonzero((truth == 1) & (pred2 == 1), axis=0)
        fp = np.count_nonzero((truth == 0) & (pred2 == 1), axis=0)
        fn = np.count_nonzero((truth == 1) & (pred2 == 0), axis=0)
        denom = 2 * tp + fp + fn
        f1 = np.divide(2 * tp, denom, out=np.ones_like(denom, dtype=float), where=denom != 0)
        for col in range(gs_labels.shape[1]):
            rec = {
                "tissue": tissue,
                "go_id": go_id,
                "source_member": member,
                "label_column": col,
                "compared_rows": int(selected_rows.size),
                "mismatch_rows": int(mismatch[col]),
                "agreement": 1 - int(mismatch[col]) / int(selected_rows.size),
                "f1": float(f1[col]),
                "true_positives": int(tp[col]),
                "false_positives": int(fp[col]),
                "false_negatives": int(fn[col]),
            }
            prev = best_by_col[col]
            key = (rec["mismatch_rows"] / rec["compared_rows"], -rec["compared_rows"], rec["source_member"])
            if prev is None:
                best_by_col[col] = rec
            else:
                prev_key = (prev["mismatch_rows"] / prev["compared_rows"], -prev["compared_rows"], prev["source_member"])
                if key < prev_key:
                    best_by_col[col] = rec
    return [best_by_col[col] for col in range(gs_labels.shape[1]) if best_by_col[col] is not None]


def greene_go_sets(table6: Path, table9: Path):
    wb6 = load_workbook(table6, read_only=True, data_only=True)
    ws6 = wb6.active
    table6_ids = {str(row[1]) for row in ws6.iter_rows(min_row=3, values_only=True) if row[1] and str(row[1]).startswith("GO:")}
    wb9 = load_workbook(table9, read_only=True, data_only=True)
    ws9 = wb9.active
    table9_ids = set()
    for row in ws9.iter_rows(min_row=3, values_only=True):
        for value in row[2:]:
            if value and str(value).startswith("GO:"):
                table9_ids.add(str(value))
    return table6_ids, table9_ids


def summarize(rows: Sequence[dict]):
    return {
        "labels": len(rows),
        "exact_matches": sum(r.get("mismatch_genes", 1) == 0 for r in rows),
        "agreement_at_least_99_percent": sum(r.get("agreement", 0.0) >= 0.99 for r in rows),
        "agreement_at_least_95_percent": sum(r.get("agreement", 0.0) >= 0.95 for r in rows),
        "median_best_agreement": float(np.median([r.get("agreement", 0.0) for r in rows])) if rows else None,
        "minimum_best_mismatch": min((r.get("mismatch_genes", 10**9) for r in rows), default=None),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-dir", type=Path, default=Path("/mnt/data"))
    ap.add_argument("--work-dir", type=Path, default=Path("/mnt/data/ppi_repro_corrected/work"))
    ap.add_argument("--output-dir", type=Path, default=Path("/mnt/data/ppi_repro_corrected/results"))
    args = ap.parse_args()
    inp, work, out = args.input_dir, args.work_dir, args.output_dir
    started = time.time()
    genes, y = load_collapsed(out / "collapsed_gene_labels_topology_features.csv")
    gene_index = {g: i for i, g in enumerate(genes)}
    label_bits = bitsets_from_labels(y)

    msigdb_archives = {
        "5.1": inp / "msigdb_v5.1_files_to_download_locally.zip",
        "5.2": inp / "msigdb_v5.2_files_to_download_locally.zip",
        "6.0": inp / "msigdb_v6.0_files_to_download_locally.zip",
    }
    all_msig_rows = []
    msig_summaries = {}
    msig_csv = out / "msigdb_label_source_screen.csv"
    if msig_csv.exists() and msig_csv.stat().st_size > 0:
        print("reuse completed MSigDB screen", flush=True)
        with msig_csv.open() as fh:
            raw_rows = list(csv.DictReader(fh))
        for row in raw_rows:
            for key in ["label_column", "observed_positive_genes", "best_candidate_positive_genes", "mismatch_genes", "true_positives", "false_positives", "false_negatives", "tie_count"]:
                row[key] = int(row[key])
            for key in ["agreement", "f1", "jaccard"]:
                row[key] = float(row[key])
        all_msig_rows = raw_rows
        for version in msigdb_archives:
            all_rows = [r for r in all_msig_rows if r["version"] == version and r["source_type"] == f"MSigDB {version} all collections"]
            bp_rows = [r for r in all_msig_rows if r["version"] == version and r["source_type"] == f"MSigDB {version} C5 BP"]
            msig_summaries[version] = {
                "unique_gene_sets_all_collections": None,
                "unique_gene_sets_c5_bp": None,
                "all_collections": summarize(all_rows),
                "c5_bp": summarize(bp_rows),
            }
    else:
        for version, archive in msigdb_archives.items():
            print(f"parse MSigDB {version}", flush=True)
            records = parse_gmt_archive(archive, version, gene_index)
            go_map = parse_go_ids_from_msigdb_xml(archive)
            all_rows = best_matches(label_bits, records, len(genes), f"MSigDB {version} all collections", go_map)
            bp_records = [r for r in records if r["collection"] == "C5" and r["subcollection"] == "BP"]
            bp_rows = best_matches(label_bits, bp_records, len(genes), f"MSigDB {version} C5 BP", go_map)
            all_msig_rows.extend(all_rows)
            all_msig_rows.extend(bp_rows)
            msig_summaries[version] = {
                "unique_gene_sets_all_collections": len(records),
                "unique_gene_sets_c5_bp": len(bp_records),
                "all_collections": summarize(all_rows),
                "c5_bp": summarize(bp_rows),
            }
        write_csv(msig_csv, all_msig_rows)

    labels_extract = work / "ohmnet_labels"
    if not list(labels_extract.rglob("*.lab")):
        labels_extract.mkdir(parents=True, exist_ok=True)
        with tarfile.open(inp / "bio-tissue-labels.tar.gz", "r:gz") as tf:
            tf.extractall(labels_extract)
    selected_tissues = set()
    with (out / "tissue_partition.csv").open() as fh:
        for row in csv.DictReader(fh):
            selected_tissues.add(row["ohmnet_tissue"])
    individual, union_all, union_selected, file_data = parse_ohmnet_labels(labels_extract, gene_index, selected_tissues)
    ohm_rows = []
    for source_type, records in [
        ("OhmNet individual tissue label files, absent genes treated as 0", individual),
        ("OhmNet union by GO term across all supplied tissues", union_all),
        ("OhmNet union by GO term across selected 24 tissues", union_selected),
    ]:
        rows = best_matches(label_bits, records, len(genes), source_type)
        ohm_rows.extend(rows)
    write_csv(out / "ohmnet_global_label_source_screen.csv", ohm_rows)

    row_to_gene = load_mapping(out / "graphsage_row_to_entrez_topology_features.csv")
    gs_labels = load_graphsage_labels(work / "graphsage")
    tissue_best = tissue_specific_screen(file_data, out / "tissue_partition.csv", row_to_gene, gs_labels)
    write_csv(out / "ohmnet_same_tissue_label_screen_best.csv", tissue_best)

    table6_ids, table9_ids = greene_go_sets(inp / "Greene2015_Table6.xlsx", inp / "Greene2015_Table9.xlsx")
    # Enrich GO-bearing rows with Greene membership flags.
    for path in [out / "msigdb_label_source_screen.csv", out / "ohmnet_global_label_source_screen.csv"]:
        with path.open() as fh:
            rows = list(csv.DictReader(fh))
        for row in rows:
            go = row.get("best_go_id", "")
            row["best_go_in_greene_table6"] = go in table6_ids
            row["best_go_in_greene_table9"] = go in table9_ids
        write_csv(path, rows)

    ohm_groups = collections.defaultdict(list)
    for row in ohm_rows:
        ohm_groups[row["source_type"]].append(row)
    summary = {
        "generated_at_utc": __import__("datetime").datetime.now(__import__("datetime").timezone.utc).isoformat(),
        "runtime_seconds": time.time() - started,
        "comparison_gene_universe": len(genes),
        "label_columns": y.shape[1],
        "gene_identity_basis": "topology plus MSigDB feature membership; GraphSAGE labels were not used to resolve identities",
        "msigdb": msig_summaries,
        "ohmnet": {key: summarize(rows) for key, rows in ohm_groups.items()},
        "ohmnet_same_tissue": {
            "selected_tissue_label_files_compared": sum(1 for x in file_data if x[0] in selected_tissues),
            "label_columns_with_any_exact_best_same_tissue_match": sum(r["mismatch_rows"] == 0 for r in tissue_best),
            "label_columns_with_best_same_tissue_agreement_at_least_99_percent": sum(r["agreement"] >= 0.99 for r in tissue_best),
            "label_columns_with_best_same_tissue_agreement_at_least_95_percent": sum(r["agreement"] >= 0.95 for r in tissue_best),
        },
        "greene": {
            "table6_unique_go_ids": len(table6_ids),
            "table9_unique_go_ids": len(table9_ids),
            "intersection": len(table6_ids & table9_ids),
        },
        "scope_limit": "No historical GOA, NCBI gene2go, GO ontology snapshot, or Bioconductor annotation package binary was supplied in this turn or successfully materialized in the runtime. This screen therefore verifies only the locally available MSigDB, OhmNet, and Greene candidates; it does not claim to resolve the exact GO source.",
    }
    (out / "local_label_source_screen_summary.json").write_text(json.dumps(summary, indent=2))
    print(json.dumps({"status": "PASS", **summary}, indent=2))


if __name__ == "__main__":
    main()
