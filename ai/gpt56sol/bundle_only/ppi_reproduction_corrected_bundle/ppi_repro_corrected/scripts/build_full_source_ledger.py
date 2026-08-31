#!/usr/bin/env python3
from __future__ import annotations
import csv, hashlib, json
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT=Path('/mnt/data/ppi_repro_corrected')
R=ROOT/'results'; S=ROOT/'scripts'
BASE=R/'actual_input_file_manifest_v2.csv'
FIELDS=['record_type','artifact_name','local_path','local_status','origin_in_this_run','analysis_role','used_by','direct_or_canonical_source_url','source_page_url','url_status','retrieval_status','retrieved_at_utc','size_bytes','sha256','parent_or_derivation','notes']

def sha(p):
    h=hashlib.sha256()
    with p.open('rb') as f:
        for c in iter(lambda:f.read(1<<20),b''): h.update(c)
    return h.hexdigest()

def row_for(p:Path, typ:str, role:str, used_by:str='', parent:str=''):
    return {
        'record_type':typ,'artifact_name':p.name,'local_path':str(p),'local_status':'present',
        'origin_in_this_run':'generated locally','analysis_role':role,'used_by':used_by,
        'direct_or_canonical_source_url':'','source_page_url':'','url_status':'not_applicable_generated_file',
        'retrieval_status':'generated locally','retrieved_at_utc':datetime.fromtimestamp(p.stat().st_mtime,timezone.utc).isoformat(),
        'size_bytes':str(p.stat().st_size),'sha256':sha(p),'parent_or_derivation':parent,'notes':''
    }

def write_xlsx(rows,path):
    wb=Workbook(); ws=wb.active; ws.title='Source ledger'; ws.append(FIELDS)
    fill=PatternFill('solid',fgColor='D9EAF7')
    for c in ws[1]: c.font=Font(bold=True); c.fill=fill; c.alignment=Alignment(wrap_text=True,vertical='top')
    for r in rows: ws.append([r.get(f,'') for f in FIELDS])
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    for i,f in enumerate(FIELDS,1):
        vals=[f]+[str(r.get(f,'')) for r in rows]
        ws.column_dimensions[get_column_letter(i)].width=min(max(max(map(len,vals))+2,12),60)
        for c in ws[get_column_letter(i)]: c.alignment=Alignment(wrap_text=True,vertical='top')
    wb.save(path)

base=list(csv.DictReader(BASE.open(encoding='utf-8')))
rows=[r for r in base if r['record_type']!='generated_output']
script_roles={
'run_core_verification.py':'Core graph, mapping, feature, DGL, and leakage verification',
'run_local_label_source_screen.py':'Local MSigDB, OhmNet-label, and Greene label-source screen',
'build_source_manifest.py':'Initial input/web/historical candidate source enumeration',
'download_or_verify_sources.py':'Download public direct files and verify recorded SHA-256 values',
'build_corrected_report.py':'Generate corrected reports, manifests, and status artifacts',
'validate_corrected_bundle.py':'Validate report references, manifests, hashes, and workbook readability',
'build_full_source_ledger.py':'Generate comprehensive source, script, and output ledger',
}
for p in sorted(S.glob('*.py')):
    rows.append(row_for(p,'analysis_script',script_roles.get(p.name,'Analysis script'),used_by='reproduction bundle',parent='authored for corrected run'))
exclude={
'source_ledger.csv','source_ledger.json','source_ledger.md','source_ledger.xlsx',
'actual_input_file_manifest_v2.csv','actual_input_file_manifest_v2.json','actual_input_file_manifest_v2.md','actual_input_file_manifest_v2.xlsx',
'RUN_STATUS.md','RUN_STATUS_v2.md','bundle_validation.json','bundle_validation.stdout','bundle_validation.stderr',
'bundle_checksums.csv','bundle_sha256.txt','refresh_run_status.stdout','refresh_run_status.stderr',
'build_full_source_ledger.stdout','build_full_source_ledger.stderr'
}
for p in sorted(x for x in R.iterdir() if x.is_file() and x.name not in exclude):
    role='Generated analysis output'
    if p.name.startswith('MASTER_'): role='Master scientific report'
    elif p.name.startswith('RUN_STATUS'): role='Run and output integrity status'
    elif p.name=='EXECUTION_DIAGNOSTICS.md': role='Complete accepted and superseded execution diagnostics'
    elif p.name=='SOURCE_ACQUISITION.md': role='Source URLs and download/verify instructions'
    elif 'manifest' in p.name: role='Actual input manifest'
    elif p.suffix=='.stdout': role='Captured standard output'
    elif p.suffix=='.stderr': role='Captured standard error or warning stream'
    rows.append(row_for(p,'generated_output',role,used_by='report/reproduction bundle',parent='corrected analysis scripts'))
# Stable ordering.
order={'actual_input':0,'web_reference':1,'historical_candidate_not_materialized':2,'analysis_script':3,'generated_output':4}
rows.sort(key=lambda r:(order.get(r['record_type'],9),r['artifact_name']))
with (R/'source_ledger.csv').open('w',newline='',encoding='utf-8') as f:
    w=csv.DictWriter(f,fieldnames=FIELDS); w.writeheader(); w.writerows(rows)
(R/'source_ledger.json').write_text(json.dumps(rows,indent=2),encoding='utf-8')
# Readable markdown.
lines=['# Comprehensive source and artifact ledger','',f'Generated: {datetime.now(timezone.utc).isoformat()}','',
'This ledger distinguishes supplied inputs, web references, historical candidates not materialized in the runtime, analysis scripts, and generated outputs. URLs are deliberately blank for generated artifacts and local notes.','',
'| Type | Artifact | Status | SHA-256 | Direct/canonical URL | Source page | Role |','|---|---|---|---|---|---|---|']
for r in rows:
    def e(s): return str(s).replace('|','\\|')
    lines.append('| '+' | '.join([e(r['record_type']),f"`{e(r['artifact_name'])}`",e(r['local_status']),f"`{e(r['sha256'])}`" if r['sha256'] else '',e(r['direct_or_canonical_source_url']),e(r['source_page_url']),e(r['analysis_role'])])+' |')
(R/'source_ledger.md').write_text('\n'.join(lines)+'\n',encoding='utf-8')
write_xlsx(rows,R/'source_ledger.xlsx')
print(json.dumps({'status':'PASS','rows':len(rows),'types':{t:sum(r['record_type']==t for r in rows) for t in sorted(set(r['record_type'] for r in rows))}},indent=2))
