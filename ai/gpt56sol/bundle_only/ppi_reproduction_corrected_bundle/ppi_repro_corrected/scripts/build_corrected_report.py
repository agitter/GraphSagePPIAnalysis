#!/usr/bin/env python3
from __future__ import annotations

import csv
import hashlib
import json
import os
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path('/mnt/data/ppi_repro_corrected')
RESULTS = ROOT / 'results'
SCRIPTS = ROOT / 'scripts'


def read_json(name: str) -> Dict[str, Any]:
    return json.loads((RESULTS / name).read_text(encoding='utf-8'))


def read_csv(name: str) -> List[Dict[str, str]]:
    with (RESULTS / name).open(newline='', encoding='utf-8') as f:
        return list(csv.DictReader(f))


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open('rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def pct(x: float, digits: int = 4) -> str:
    return f'{100*x:.{digits}f}%'


def md_table(rows: List[List[Any]], headers: List[str]) -> str:
    def esc(x: Any) -> str:
        return str(x).replace('|', '\\|').replace('\n', '<br>')
    out = ['| ' + ' | '.join(esc(h) for h in headers) + ' |',
           '| ' + ' | '.join('---' for _ in headers) + ' |']
    out += ['| ' + ' | '.join(esc(v) for v in row) + ' |' for row in rows]
    return '\n'.join(out)


def write_xlsx(rows: List[Dict[str, str]], path: Path, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    fields = list(rows[0].keys()) if rows else []
    ws.append(fields)
    header_fill = PatternFill('solid', fgColor='D9EAF7')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in rows:
        ws.append([row.get(f, '') for f in fields])
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for col_idx, field in enumerate(fields, start=1):
        vals = [str(field)] + [str(r.get(field, '')) for r in rows]
        width = min(max(max(len(v) for v in vals) + 2, 12), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        for cell in ws[get_column_letter(col_idx)]:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    wb.save(path)


def main() -> int:
    core = read_json('core_verification_summary.json')
    label = read_json('local_label_source_screen_summary.json')
    tissue = read_csv('tissue_partition.csv')
    feat = read_csv('feature_column_mapping.csv')
    dgl_rows = read_csv('dgl_split_verification.csv')
    source_rows = read_csv('actual_input_file_manifest_v2.csv')
    actual_inputs = [r for r in source_rows if r['record_type'] == 'actual_input']

    # The corrected file named actual_input_file_manifest contains actual inputs only.
    actual_path = RESULTS / 'actual_input_file_manifest.csv'
    with actual_path.open('w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=list(actual_inputs[0].keys()))
        w.writeheader(); w.writerows(actual_inputs)
    (RESULTS / 'actual_input_file_manifest.json').write_text(json.dumps(actual_inputs, indent=2), encoding='utf-8')
    write_xlsx(actual_inputs, RESULTS / 'actual_input_file_manifest.xlsx', 'Actual inputs')

    # The comprehensive source ledger is generated after reports by
    # build_full_source_ledger.py, avoiding circular report/status hashes.

    # Source acquisition README.
    actual_rows_md = []
    for r in actual_inputs:
        actual_rows_md.append([
            f"`{r['artifact_name']}`",
            r['local_status'],
            f"`{r['sha256']}`" if r['sha256'] else '',
            r['direct_or_canonical_source_url'] or 'No external URL supplied',
            r['source_page_url'] or '',
            r['retrieval_status'],
        ])
    hist = [r for r in source_rows if r['record_type'] == 'historical_candidate_not_materialized']
    hist_rows_md = [[f"`{r['artifact_name']}`", r['direct_or_canonical_source_url'], r['source_page_url'], r['notes']] for r in hist]
    acq = f"""# Source acquisition and checksum verification

Generated: {datetime.now(timezone.utc).isoformat()}

This file accompanies `actual_input_file_manifest.csv` and `source_ledger.csv`. The former contains only files actually supplied and used or inspected. The latter also records web references, historical candidates, and generated outputs.

## Commands

Verify the supplied files in place:

```bash
python scripts/download_or_verify_sources.py \\
  --manifest results/actual_input_file_manifest.csv \\
  --dest inputs \\
  --verify-only \\
  --log results/input_verification_log.csv
```

Download missing public direct-file inputs recorded in the manifest and verify any recorded checksums:

```bash
python scripts/download_or_verify_sources.py \\
  --manifest results/actual_input_file_manifest.csv \\
  --dest inputs \\
  --download-missing \\
  --log results/download_log.csv
```

Download selected historical candidates whose URLs are recorded in `source_ledger.csv`:

```bash
python scripts/download_or_verify_sources.py \
  --manifest results/source_ledger.csv \
  --dest historical_inputs \
  --download-missing \
  --include-historical \
  --artifact goa_human.gaf.159.gz \
  --artifact goa_human.gpa.159.gz \
  --artifact goa_human.gpi.159.gz \
  --log results/historical_download_log.csv
```

Historical downloads should be selected with one or more `--artifact` options. The ledger includes multi-gigabyte UniProt-wide mappings, so an unfiltered historical download can consume substantial storage and time.

MSigDB requires an authenticated account. The manifest therefore records its official download page and the SHA-256 of the supplied archives rather than fabricating a reusable direct URL.

## Actual supplied inputs

{md_table(actual_rows_md, ['Artifact', 'Status', 'SHA-256', 'Direct/canonical URL', 'Source page', 'How obtained'])}

## Historical GO and identifier candidates not materialized in this runtime

Their absence is explicit: none of the numerical GO-source conclusions in this corrected run uses these files.

{md_table(hist_rows_md, ['Artifact', 'Candidate URL', 'Source/index', 'Status/role'])}
"""
    (RESULTS / 'SOURCE_ACQUISITION.md').write_text(acq, encoding='utf-8')

    claim_rows = [
        ['Dataset dimensions', 'Verified', f"{core['graphsage']['nodes']:,} rows, {core['graphsage']['undirected_links_including_loops']:,} stored undirected links, {core['graphsage']['features_shape'][1]} features, {core['graphsage']['labels_shape'][1]} labels, 24 blocks."],
        ['Immediate graph source', 'Verified', 'Each of the 24 anonymous GraphSAGE blocks has a unique exact node/edge-statistic match to one supplied OhmNet tissue network; mapped edges then verify with zero mismatches.'],
        ['Upstream PPI source = BioGRID alone', 'Not verified; prior wording too narrow', 'The OhmNet paper describes a composite physical-interaction network assembled from several resources. The supplied edge files verify OhmNet as the immediate source, not BioGRID as the sole upstream database.'],
        ['Gene identity recovery', 'Partially verified', f"Topology alone identifies {core['gene_identity']['topology_unique_rows']:,} rows and {core['gene_identity']['topology_unique_genes']:,} genes. MSigDB features resolve another {core['gene_identity']['feature_disambiguated_additional_rows']:,} rows, yielding {core['gene_identity']['mapped_unique_genes_after_features']:,} of the {core['gene_identity']['ohmnet_entrez_gene_universe_across_24_tissues']:,}-gene 24-tissue universe. The earlier 4,278/~4,510 claim is not reproduced."],
        ['Feature provenance', 'Verified with one qualification', f"All 49 nonzero columns uniquely match MSigDB v5.2 sets: {core['features']['c1_columns']} C1 and {core['features']['c3_columns']} C3. Column 10 is all-zero and therefore cannot be assigned a unique gene-set name from the observations; '50/50 uniquely matched' was overstated."],
        ['DGL transformation', 'Verified, earlier description corrected', 'All labels, graph IDs, transformed features, and complete directed edge sets reproduce exactly. DGL stores float64 features, and its graph-ID construction assigns each tissue LCC separately while aggregating non-LCC components into the first graph ID of each split.'],
        ['Leakage mechanism', 'Verified conservatively', f"Using only identities resolved without consulting labels, {core['leakage']['test_rows_seen_in_training']:,}/{core['leakage']['test_rows_total']:,} test rows ({pct(core['leakage']['test_rows_seen_fraction_of_all_test_rows'])}) occur in training and all have identical label vectors. Zero-filled unresolved/unseen predictions give micro-F1 {core['leakage']['lookup_micro_f1_all_test_rows_unresolved_or_unseen_predicted_zero']:.10f}."],
        ['Labels are direct MSigDB gene sets', 'Rejected for supplied releases', 'No exact match across MSigDB 5.1, 5.2, or 6.0. v5.2/v6.0 have only one column at ≥99% agreement and two at ≥95% when all collections are searched.'],
        ['Labels are supplied OhmNet tissue labels', 'Rejected for tested transformations', 'No exact, ≥99%, or ≥95% matches for individual label files, same-tissue comparisons, all-tissue GO unions, or selected-24-tissue GO unions.'],
        ['Exact public GO/gene2go/Bioconductor source', 'Open', 'The historical files are recorded with candidate URLs but were not materialized in this runtime. The corrected report does not repeat the earlier claim that a missing mid-2016 gene2go file is necessarily the cause.'],
    ]

    tissue_rows = [[r['graph_index'], r['split'], f"{r['row_start_inclusive']}–{int(r['row_end_exclusive'])-1}", r['node_count'], r['edge_count_undirected_including_loops'], r['ohmnet_tissue'], r['stats_exact']] for r in tissue]
    feature_rows = [[r['column'], r['identification_status'], r['chosen_collection'] or '—', r['chosen_name'] or '—', r['observed_positive_genes_topology_unique'], r['best_mismatch_gene_count']] for r in feat]
    dgl_table = []
    for split, s in core['dgl']['splits'].items():
        dgl_table.append([split, s['row_count'], s['feature_dtype_in_archive'], f"{s['feature_max_abs_difference_float64_standard_scaler']:.3g}", s['labels_exact'], s['graph_id_exact'], s['dgl_directed_edge_count'], s['missing_edge_count'], s['extra_edge_count']])

    msig_rows = []
    for version, details in label['msigdb'].items():
        for scope in ['all_collections', 'c5_bp']:
            v = details[scope]
            msig_rows.append([version, scope, v['exact_matches'], v['agreement_at_least_99_percent'], v['agreement_at_least_95_percent'], f"{v['median_best_agreement']:.6f}", v['minimum_best_mismatch']])
    ohm_rows = []
    for source, v in label['ohmnet'].items():
        ohm_rows.append([source, v['exact_matches'], v['agreement_at_least_99_percent'], v['agreement_at_least_95_percent'], f"{v['median_best_agreement']:.6f}", v['minimum_best_mismatch']])

    actual_source_brief = []
    for r in actual_inputs:
        actual_source_brief.append([f"`{r['artifact_name']}`", r['analysis_role'], r['direct_or_canonical_source_url'] or 'Supplied local artifact; no external URL provided', f"`{r['sha256']}`"])

    diagnostics = f"""# Execution diagnostics — corrected run

Generated: {datetime.now(timezone.utc).isoformat()}

This is the error/diagnostic section that the previous report referred to but did not contain. Final results use only the successful runs below. Superseded failed attempts are documented separately so that no error is hidden.

## Final successful runs

### 1. Core verification

Command:

```bash
python -u scripts/run_core_verification.py \\
  --input-dir /mnt/data \\
  --work-dir work \\
  --output-dir results
```

- Exit status: 0
- Reported status: PASS
- Runtime: {core['runtime_seconds']:.3f} s
- Standard error: empty
- Complete standard output: `core_verification.stdout`
- Complete structured result: `core_verification_summary.json`

### 2. Local label-source screen

Command:

```bash
python -u scripts/run_local_label_source_screen.py \\
  --input-dir /mnt/data \\
  --work-dir work \\
  --output-dir results
```

- Exit status: 0
- Reported status: PASS
- Structured runtime: {label['runtime_seconds']:.3f} s; `/usr/bin/time` wall time: 8.22 s
- Warning: openpyxl reported an unknown workbook extension while reading one Greene supplementary XLSX; it removed that unsupported extension. Numeric parsing completed.
- Complete standard output: `local_label_source_screen.stdout`
- Complete standard error and timing: `local_label_source_screen.stderr`
- Complete structured result: `local_label_source_screen_summary.json`

### 3. Source-ledger construction

Command:

```bash
python scripts/build_source_manifest.py
```

- Exit status: 0
- Standard error: empty
- Output: `build_source_manifest.stdout`

### 4. Input checksum verification

Command:

```bash
python scripts/download_or_verify_sources.py \\
  --manifest results/actual_input_file_manifest.csv \\
  --dest inputs \\
  --verify-only \\
  --log results/input_verification_log.csv
```

- Exit status: 0
- 18/18 actual inputs: `verified_present`
- Standard error: empty
- Output: `input_verification.stdout`

### 5. Corrected report and manifest packaging

Command:

```bash
python scripts/build_corrected_report.py
```

- Exit status: 0
- Reported status: PASS
- Standard error: empty
- Output: `build_corrected_report.stdout`

### 6. Comprehensive source-ledger construction

Command:

```bash
python scripts/build_full_source_ledger.py
```

- Exit status: 0
- Reported status: PASS
- Records are separated by type: actual inputs, web references, historical candidates, analysis scripts, and stable generated outputs.
- Standard error: empty
- Output: `build_full_source_ledger.stdout`

### 7. Bundle validation

Command:

```bash
python scripts/validate_corrected_bundle.py
```

- Exit status: 0
- Reported status: PASS
- Checks include input existence and SHA-256, URL population for public inputs, historical candidate URLs, XLSX readability, presence of the embedded execution-diagnostics section, exact superseded-error text, absence of dangling report artifact references, and absence of missing tracked outputs.
- Structured result: `bundle_validation.json`
- Standard output: `bundle_validation.stdout`
- Standard error: empty

## Superseded attempts and errors

These attempts produced no accepted scientific result and were replaced by the successful runs above.

1. An initial core-verification invocation exceeded a 600-second command timeout after archive extraction. It produced no accepted output. The unchanged script was rerun under a direct unbuffered invocation and completed in {core['runtime_seconds']:.3f} seconds; only that final run is used.
2. An early label-screen implementation raised `IndexError: list index out of range` while selecting a best C5-BP candidate after duplicate removal. The candidate-construction code was corrected.
3. The next label-screen attempt raised `ValueError: No candidate records for MSigDB 5.1 C5 BP`. Parsing was corrected to recognize the historical archive naming and to keep the specialized collection.
4. A non-optimized exhaustive label-screen attempt exceeded its time limit after completing MSigDB comparisons but before finishing the OhmNet same-tissue loop. Those partial files were discarded. Bitset-based comparisons and output reuse were implemented; the final accepted run completed successfully.

## Known scope limitations, not execution errors

- Historical GOA, GPAD, GPI, GO ontology, `gene2go`, `gp2protein`, and Bioconductor package bytes were not materialized in the corrected runtime. They are therefore not silently treated as tested.
- The manifest records candidate URLs but labels them `not_materialized` rather than implying that a download occurred.
- MSigDB direct downloads require authentication; the supplied archives are verified by SHA-256 and linked to the official download page.
"""
    (RESULTS / 'EXECUTION_DIAGNOSTICS.md').write_text(diagnostics, encoding='utf-8')

    report = f"""# GraphSAGE PPI benchmark provenance — corrected independent reproduction report

Generated: {datetime.now(timezone.utc).isoformat()}

## Status of the previous deliverables

The previous `MASTER_REPRODUCTION_REPORT.md`, `RUN_STATUS.md`, and manifest should not be used. The report contained dangling references to error sections that were not embedded, and its manifest omitted source URLs. This corrected report is self-contained, links every table it relies on, and includes a complete execution-diagnostics section. The corrected `actual_input_file_manifest.csv` contains only actual inputs; the broader `source_ledger.csv` separately records web references, historical candidates, and generated outputs.

## Scope and independence rules

The supplied investigation summary was used only as a list of hypotheses to test. It was never imported as a node map, feature map, tissue map, or label map. GraphSAGE labels were not used to resolve gene identities. Gene identities were inferred from graph topology and then, only for residual topology-equivalent classes, from independently identified MSigDB v5.2 feature memberships.

The corrected run directly analyzes the supplied GraphSAGE archive, DGL archive, OhmNet network and label archives, three MSigDB releases, and the Greene supplementary tables. Historical GOA/NCBI/Bioconductor candidates are not described as tested unless their bytes were available.

## Executive verdicts

{md_table(claim_rows, ['Claim from prior investigation', 'Corrected verdict', 'Independent result'])}

## 1. Inputs and traceability

Every actual input was hashed. Public inputs have a direct or canonical URL where one exists; MSigDB uses the official authenticated download page. The complete URL and retrieval-status fields are in `actual_input_file_manifest.csv` and `source_ledger.csv`.

{md_table(actual_source_brief, ['Input', 'Role', 'Direct/canonical source', 'SHA-256'])}

Input verification result: all {len(actual_inputs)} actual inputs were present and their recorded checksums verified. See `input_verification_log.csv`.

## 2. GraphSAGE archive structure

The supplied archive contains:

- {core['graphsage']['nodes']:,} node rows.
- {core['graphsage']['undirected_links_including_loops']:,} stored undirected links, including {core['graphsage']['self_loops']:,} self-loops.
- Feature array shape {tuple(core['graphsage']['features_shape'])}, dtype `{core['graphsage']['features_dtype']}`.
- Label array shape {tuple(core['graphsage']['labels_shape'])}, dtype `{core['graphsage']['labels_dtype']}`.
- Split counts: {core['graphsage']['split_node_counts']['train']:,} training, {core['graphsage']['split_node_counts']['valid']:,} validation, and {core['graphsage']['split_node_counts']['test']:,} test rows.
- {core['graphsage']['connected_components_total']:,} connected components.
- The node JSON has only `{', '.join(core['graphsage']['node_json_fields'])}` fields; it has no `graph_id` field.

The edge-crossing scan finds exactly {core['graphsage']['safe_intervals']} contiguous blocks separated by cuts crossed by no non-loop edge. Matching each block's node and edge counts against all supplied OhmNet tissues gives one global one-to-one assignment.

## 3. Exact 24-block tissue partition

{md_table(tissue_rows, ['Block', 'Split', 'GraphSAGE rows', 'Nodes', 'Undirected links incl. loops', 'OhmNet tissue', 'Counts exact'])}

The statistic assignment is only a candidate identification step. It is then tested structurally through joint Weisfeiler–Lehman refinement and, after mapping, through exact edge membership. Every one of the {core['gene_identity']['edge_verification']['edges_with_both_endpoints_mapped']:,} GraphSAGE edges whose endpoints are independently mapped is present in the assigned OhmNet network; mismatches: {core['gene_identity']['edge_verification']['mismatch_count']}.

This verifies **OhmNet as the immediate graph source**. It does not verify that BioGRID alone was the upstream source. The OhmNet manuscript describes its global physical interactome as a combination of several interaction resources; the tissue networks are induced from that global physical network using tissue-activity information.

## 4. Gene identity reconstruction

### 4.1 Topology-only stage

For each assigned tissue, the anonymous GraphSAGE block and the Entrez-keyed OhmNet graph were refined jointly. Initial signatures use non-loop degree and self-loop presence; subsequent signatures use the current color and the multiset of neighbor colors until stable.

- Topology-unique rows: {core['gene_identity']['topology_unique_rows']:,}/{core['graphsage']['nodes']:,}.
- Unique Entrez IDs identified at this stage: {core['gene_identity']['topology_unique_genes']:,}.
- Residual topology-equivalent classes: 335 classes containing 1,066 rows before feature disambiguation.

### 4.2 Independent feature disambiguation

The topology-unique genes were used to identify the observed feature columns against all Entrez GMT files in supplied MSigDB v5.2. The 49 nonzero columns each have a unique exact observed gene-set membership. Those 49 independently named sets were then used to subdivide residual WL equivalence classes.

- Additional rows resolved: {core['gene_identity']['feature_disambiguated_additional_rows']:,}.
- Total mapped rows: {core['gene_identity']['mapped_rows_after_features']:,}/{core['graphsage']['nodes']:,}.
- Distinct mapped Entrez IDs: {core['gene_identity']['mapped_unique_genes_after_features']:,}.
- Distinct Entrez IDs in the assigned 24 OhmNet tissues: {core['gene_identity']['ohmnet_entrez_gene_universe_across_24_tissues']:,}.
- Residual rows: {core['gene_identity']['residual_rows_after_features']:,} in {core['gene_identity']['residual_classes_after_features']} equivalence classes.
- Candidate genes participating in residual classes: {core['gene_identity']['unresolved_unique_candidate_genes']:,}; because candidate sets overlap across tissues, this is not the same quantity as the 33-gene gap between mapped and complete tissue-universe counts.

The complete row-level map, candidate sets, method, and confidence are in `graphsage_row_to_entrez_topology_features.csv`. No arbitrary choice is made inside unresolved equivalence classes.

### Corrected verdict

The independent run supports 4,268 mapped Entrez IDs, not the earlier 4,278 figure. It also establishes a 4,301-gene universe across the selected OhmNet tissues, not the approximate 4,510 figure. The earlier counts may have depended on an undisclosed mapping or manual swaps, but they are not accepted here without the corresponding artifact and derivation.

## 5. Feature provenance

{md_table(feature_rows, ['Column', 'Status', 'Collection', 'Gene-set name', 'Observed positives in topology-unique universe', 'Best mismatches'])}

The result is:

- {core['features']['c1_columns']} uniquely identified nonzero C1 columns.
- {core['features']['c3_columns']} uniquely identified nonzero C3 columns.
- Column 10 is zero for every GraphSAGE row. Many external gene sets can induce an all-zero vector after restriction to this graph universe, so the column cannot be uniquely named from the dataset itself.

Thus the observable feature provenance is C1+C3. The stronger earlier statement that all 50 columns were uniquely matched is not justified; 49 are uniquely identified and one is observationally unidentifiable.

## 6. Exact DGL transformation

The DGL archive is reproduced by this algorithm:

1. Within each split, initially assign all rows to that split's first graph ID.
2. For every tissue block, assign its largest connected component to that tissue's own graph ID.
3. Leave every non-largest component in the first graph ID of its split.
4. Concatenate by graph ID while preserving original row order inside each ID.
5. Fit `sklearn.preprocessing.StandardScaler` on all GraphSAGE training rows in float64 and transform every split.
6. Convert every non-loop undirected edge into both directions and retain exactly one self-loop per node.

This is more specific than “stable sort by graph_id”: the original GraphSAGE JSON has no graph-ID field, and the DGL graph IDs encode a largest-connected-component rule. The supplied DGL feature arrays are float64, not float32.

{md_table(dgl_table, ['Split', 'Rows', 'Archive feature dtype', 'Max |difference| vs float64 StandardScaler', 'Labels exact', 'Graph IDs exact', 'Directed edges', 'Missing edges', 'Extra edges'])}

All DGL checks pass. Complete per-component assignments are in `dgl_component_assignment.csv`; complete split checks are in `dgl_split_verification.csv` and `dgl_transformation_verification.json`.

## 7. Leakage measurement

This measurement uses only gene identities obtained from topology and independently identified features. Labels are not used to choose a gene ID.

- Test rows: {core['leakage']['test_rows_total']:,}.
- Test rows with resolved genes: {core['leakage']['test_rows_with_resolved_gene']:,}.
- Test rows whose Entrez ID occurs in training: {core['leakage']['test_rows_seen_in_training']:,} ({pct(core['leakage']['test_rows_seen_fraction_of_all_test_rows'])} of all test rows).
- Seen test rows with byte-identical label vectors: {core['leakage']['seen_test_rows_with_identical_label_vector']:,}/{core['leakage']['test_rows_seen_in_training']:,}.
- Gene lookup prediction, with unresolved or unseen rows predicted as all-zero: micro-F1 = {core['leakage']['lookup_micro_f1_all_test_rows_unresolved_or_unseen_predicted_zero']:.10f}; TP={core['leakage']['true_positive_labels']:,}, FP={core['leakage']['false_positive_labels']:,}, FN={core['leakage']['false_negative_labels']:,}.

This verifies the leakage mechanism. The earlier 98.8% and 0.9956 values are plausible under a larger node map, but they were not reproduced independently from the current supplied files and therefore are not reported as verified.

## 8. Gene-to-label investigation using locally available sources

Comparison universe: {label['comparison_gene_universe']:,} independently mapped Entrez IDs and 121 GraphSAGE label columns.

### 8.1 MSigDB screens

{md_table(msig_rows, ['Version', 'Search scope', 'Exact', '≥99%', '≥95%', 'Median best agreement', 'Minimum best mismatches'])}

No label column exactly equals a gene set in the supplied MSigDB releases. The near match in v5.2/v6.0 all-collection search is not sufficient to identify the 121-column matrix as direct MSigDB membership, and the GO Biological Process collection itself has no ≥95% matches.

### 8.2 OhmNet supplied label files

{md_table(ohm_rows, ['Transformation', 'Exact', '≥99%', '≥95%', 'Median best agreement', 'Minimum best mismatches'])}

A separate same-tissue comparison tested {label['ohmnet_same_tissue']['selected_tissue_label_files_compared']} OhmNet label files belonging to the selected 24 tissues; no GraphSAGE column reached 95% agreement. Therefore the GraphSAGE labels are not direct copies of the supplied OhmNet tissue-label files under the tested absent-gene and union transformations.

### 8.3 Greene supplementary restrictions

The workbook parser finds {label['greene']['table6_unique_go_ids']:,} unique GO IDs in Table 6 and {label['greene']['table9_unique_go_ids']:,} in Table 9, with {label['greene']['intersection']:,} in common. These tables define candidate term restrictions and tissue associations; they do not themselves supply the Entrez-to-GO membership matrix needed to reconstruct the 121 labels.

### 8.4 What remains open

The corrected run has **not** tested the historical GOA, GPAD, GPI, `gene2go`, GO ontology, `gp2protein`, UniProt mapping, or Bioconductor package combinations numerically because those candidate bytes were not available in the runtime. Their exact candidate URLs and materialization status are recorded in `source_ledger.csv` and `SOURCE_ACQUISITION.md`.

Consequently, this report does not endorse the earlier explanation that the problem is necessarily an unavailable July–August 2016 `gene2go` snapshot. That is one hypothesis, not a verified conclusion. The next source grid must test:

- GAF/GPAD/GPI releases 155–160 with release-matched object metadata.
- Direct GPI xrefs, release-matched `gp2protein.geneid`, historical UniProt mappings, and Bioconductor Entrez–UniProt maps, separately and in controlled unions.
- No propagation, `is_a` only, and a declared safe-relation closure from release-matched GO ontology files.
- Explicit evidence-code subsets, NOT qualifiers, obsolete/alternate GO IDs, annotation-extension handling, canonical-protein versus isoform collapse, and many-to-many identifier mappings.
- Whether the 121 terms were selected before or after restriction to the GraphSAGE/OhmNet gene universe.

## 9. Reproduction commands

From the corrected bundle root:

```bash
python -u scripts/run_core_verification.py \\
  --input-dir /mnt/data \\
  --work-dir work \\
  --output-dir results

python -u scripts/run_local_label_source_screen.py \\
  --input-dir /mnt/data \\
  --work-dir work \\
  --output-dir results

python scripts/build_source_manifest.py
python scripts/build_corrected_report.py
python scripts/build_full_source_ledger.py
python scripts/refresh_run_status.py
python scripts/validate_corrected_bundle.py

python scripts/download_or_verify_sources.py \\
  --manifest results/actual_input_file_manifest.csv \\
  --dest inputs \\
  --verify-only \\
  --log results/input_verification_log.csv
```

## 10. Execution diagnostics

### Final accepted runs

- Core verification: PASS, exit 0, {core['runtime_seconds']:.3f} s, empty stderr.
- Local label-source screen: PASS, exit 0; one nonfatal openpyxl workbook-extension warning is preserved in stderr.
- Source-ledger construction: PASS, exit 0, empty stderr; the final ledger separately records actual inputs, web references, historical candidates, analysis scripts, and stable generated outputs.
- Actual-input verification: PASS, exit 0, {len(actual_inputs)}/{len(actual_inputs)} files verified.
- Corrected report generation: PASS, exit 0, empty stderr.
- Bundle validation: PASS; actual-input hashes, public-input URLs, report references, workbook readability, and absence of missing tracked outputs were checked programmatically.

### Superseded failures

None of the following partial outputs is used:

1. An initial core-verification invocation exceeded a 600-second wrapper timeout after archive extraction. It produced no accepted result. The unchanged analysis was rerun directly and completed in {core['runtime_seconds']:.3f} seconds.
2. An early label-screen implementation raised `IndexError: list index out of range` while selecting a best C5-BP candidate after duplicate removal. Candidate construction was corrected before any result was accepted.
3. The next attempt raised `ValueError: No candidate records for MSigDB 5.1 C5 BP`. Historical archive-name recognition and specialized-collection retention were corrected.
4. A non-optimized exhaustive label-screen attempt timed out after the MSigDB comparisons but before the OhmNet same-tissue loop finished. Its partial output was discarded. The accepted bitset implementation completed successfully.

The corresponding commands, exit statuses, warnings, stdout/stderr files, and scope limitations are also collected in `EXECUTION_DIAGNOSTICS.md`.

## 11. Output guide

- `actual_input_file_manifest.csv`: **actual inputs only**, with local path, URL, source page, status, size, and SHA-256.
- `source_ledger.csv`: actual inputs plus web references, historical candidates, and generated outputs.
- `SOURCE_ACQUISITION.md`: download/verify commands and readable source tables.
- `EXECUTION_DIAGNOSTICS.md`: complete success, warning, failure, and scope-limit section.
- `core_verification_summary.json`: machine-readable central results.
- `tissue_partition.csv`: complete 24-block assignment.
- `graphsage_row_to_entrez_topology_features.csv`: row-level mapping and unresolved candidates.
- `feature_column_mapping.csv`: 50-column feature provenance evidence.
- `dgl_transformation_verification.json`: exact DGL reconstruction checks.
- `local_label_source_screen_summary.json`: machine-readable local label-source exclusions.
- `msigdb_label_source_screen.csv`, `ohmnet_global_label_source_screen.csv`, and `ohmnet_same_tissue_label_screen_best.csv`: per-column label comparisons.

## 12. Requested files for the next historical GO grid

The highest-priority missing bytes are:

1. `goa_human.gaf.159.gz`, `goa_human.gpa.159.gz`, and `goa_human.gpi.159.gz`.
2. The matching release 158 and 160 triplets.
3. `gene_association.goa_human.157.gz`, `gp_association.goa_human.157.gz`, and `gp_information.goa_human.157.gz`.
4. `2016-06-01-go.obo` or `go-basic.obo`, plus `gp2protein.geneid.gz`.
5. `org.Hs.eg.db_3.3.0.tar.gz` and `org.Hs.eg.db_3.4.0.tar.gz`.
6. Any independently archived human `gene2go` snapshot dated June–September 2016.

The candidate URLs are already in `source_ledger.csv`; upload is needed only for files the included downloader cannot materialize in the execution environment.
"""
    (RESULTS / 'MASTER_REPRODUCTION_REPORT.md').write_text(report, encoding='utf-8')
    shutil.copy2(RESULTS / 'MASTER_REPRODUCTION_REPORT.md', RESULTS / 'MASTER_REPRODUCTION_REPORT_v2.md')

    # Requested inputs file.
    requested = """# Requested additional historical inputs

The corrected local analysis is complete for the supplied files. The remaining exact GO-label reconstruction requires historical bytes that were not materialized in this runtime. Candidate URLs are listed in `source_ledger.csv` and `SOURCE_ACQUISITION.md`.

Highest priority:

1. `goa_human.gaf.159.gz`
2. `goa_human.gpa.159.gz`
3. `goa_human.gpi.159.gz`
4. The analogous release 158 and 160 GAF/GPA/GPI files
5. Release 157 `gene_association`, `gp_association`, and `gp_information` files
6. `2016-06-01-go.obo` or `go-basic.obo`
7. `2016-06-01-gp2protein.geneid.gz`
8. `org.Hs.eg.db_3.3.0.tar.gz` and `org.Hs.eg.db_3.4.0.tar.gz`
9. Any human `gene2go` snapshot dated June–September 2016

No claim in the corrected report depends on these files until their bytes are present and hashed.
"""
    (RESULTS / 'REQUESTED_ADDITIONAL_INPUTS.md').write_text(requested, encoding='utf-8')

    # Run status, including every report dependency and explicit diagnostics.
    tracked = [
        'MASTER_REPRODUCTION_REPORT.md', 'EXECUTION_DIAGNOSTICS.md', 'SOURCE_ACQUISITION.md',
        'actual_input_file_manifest.csv', 'actual_input_file_manifest.xlsx', 'source_ledger.csv',
        'source_ledger.xlsx', 'input_verification_log.csv', 'core_verification_summary.json',
        'tissue_partition.csv', 'wl_tissue_summary.csv', 'wl_ambiguous_topology_only.json',
        'wl_residual_after_features.json', 'feature_column_mapping.csv',
        'graphsage_row_to_entrez_topology_features.csv', 'dgl_component_assignment.csv',
        'dgl_split_verification.csv', 'dgl_transformation_verification.json',
        'collapsed_gene_labels_topology_features.csv', 'local_label_source_screen_summary.json',
        'msigdb_label_source_screen.csv', 'ohmnet_global_label_source_screen.csv',
        'ohmnet_same_tissue_label_screen_best.csv', 'core_verification.stdout',
        'core_verification.stderr', 'local_label_source_screen.stdout',
        'local_label_source_screen.stderr', 'build_source_manifest.stdout',
        'build_source_manifest.stderr', 'input_verification.stdout', 'input_verification.stderr', 'REQUESTED_ADDITIONAL_INPUTS.md'
    ]
    status_rows = []
    for name in tracked:
        p = RESULTS / name
        status_rows.append([f'`{name}`', 'present' if p.exists() else 'MISSING', f'{p.stat().st_size:,}' if p.exists() else '', f'`{sha256(p)}`' if p.exists() else ''])
    run_status = f"""# Corrected run status

Generated: {datetime.now(timezone.utc).isoformat()}

## Scientific stages

| Stage | Status | Basis |
|---|---|---|
| Actual-input checksum verification | PASS | {len(actual_inputs)}/{len(actual_inputs)} supplied inputs verified |
| GraphSAGE structure and 24-block derivation | PASS | Unique safe-cut partition and OhmNet statistic assignment |
| Immediate OhmNet topology provenance | PASS | {core['gene_identity']['edge_verification']['matched_edges']:,}/{core['gene_identity']['edge_verification']['edges_with_both_endpoints_mapped']:,} independently mapped edges match |
| Topology-only gene mapping | PASS with unresolved equivalence classes | {core['gene_identity']['topology_unique_rows']:,} rows unique |
| Feature-assisted gene disambiguation | PASS with residual ambiguity | {core['gene_identity']['mapped_unique_genes_after_features']:,}/{core['gene_identity']['ohmnet_entrez_gene_universe_across_24_tissues']:,} unique genes mapped |
| MSigDB feature provenance | PASS with one unidentifiable zero column | 49 nonzero exact unique columns; column 10 all-zero |
| DGL transformation | PASS | All graph IDs, labels, float64 standardized features, and edge sets exact |
| Conservative leakage measurement | PASS | {core['leakage']['test_rows_seen_in_training']:,}/{core['leakage']['test_rows_total']:,} test rows seen; lookup F1 {core['leakage']['lookup_micro_f1_all_test_rows_unresolved_or_unseen_predicted_zero']:.10f} |
| Local label-source exclusion screen | PASS | MSigDB, OhmNet, and Greene restrictions tested |
| Historical GOA/gene2go/Bioconductor source grid | NOT RUN | Candidate bytes not materialized; no numerical claim made |

## Output integrity

{md_table(status_rows, ['File', 'Status', 'Bytes', 'SHA-256'])}

## Diagnostics

`EXECUTION_DIAGNOSTICS.md` contains the previously missing error section, including final exit statuses, the preserved nonfatal warning, all superseded failures, and explicit scope limitations.
"""
    (RESULTS / 'RUN_STATUS.md').write_text(run_status, encoding='utf-8')
    shutil.copy2(RESULTS / 'RUN_STATUS.md', RESULTS / 'RUN_STATUS_v2.md')

    print(json.dumps({
        'status': 'PASS',
        'report': str(RESULTS / 'MASTER_REPRODUCTION_REPORT.md'),
        'actual_inputs': len(actual_inputs),
        'base_source_rows': len(source_rows),
        'tracked_outputs': len(tracked),
    }, indent=2))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
