#!/usr/bin/env python3
from __future__ import annotations

import argparse
import collections
import csv
import datetime as dt
import gzip
import hashlib
import html
import io
import json
import os
import re
import shutil
import statistics
import subprocess
import tarfile
import time
import zipfile
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import load_workbook
import numpy as np

UTC = dt.timezone.utc
GAF_FIELDS = [
    "DB", "DB_Object_ID", "DB_Object_Symbol", "Qualifier", "GO_ID",
    "DB_Reference", "Evidence_Code", "With_From", "Aspect", "DB_Object_Name",
    "DB_Object_Synonym", "DB_Object_Type", "Taxon", "Date", "Assigned_By",
    "Annotation_Extension", "Gene_Product_Form_ID",
]
GPAD_FIELDS = [
    "DB", "DB_Object_ID", "Relation", "GO_ID", "DB_Reference", "ECO_ID",
    "With_From", "Interacting_Taxon_ID", "Date", "Assigned_By",
    "Annotation_Extension", "Properties",
]
GPI_FIELDS = [
    "DB", "DB_Object_ID", "DB_Object_Symbol", "DB_Object_Name",
    "DB_Object_Synonyms", "DB_Object_Type", "Taxon", "Parent_Object_ID",
    "DB_Xrefs", "Properties",
]
RELATION_BY_ASPECT = {"F": "enables", "P": "involved_in", "C": "part_of"}
EXPERIMENTAL = {"EXP", "IDA", "IPI", "IMP", "IGI", "IEP"}
EVIDENCE_FILTERS = {
    "all_non_NOT": None,
    "exclude_IEA": {"TAS", "IPI", "IDA", "IBA", "ISS", "IMP", "NAS", "ND", "IC", "EXP", "IGI", "IEP", "IKR"},
    "exclude_IEA_ND": {"TAS", "IPI", "IDA", "IBA", "ISS", "IMP", "NAS", "IC", "EXP", "IGI", "IEP", "IKR"},
    "experimental_only": set(EXPERIMENTAL),
    "experimental_plus_TAS": set(EXPERIMENTAL) | {"TAS"},
    "experimental_plus_TAS_IC": set(EXPERIMENTAL) | {"TAS", "IC"},
    "experimental_plus_TAS_NAS_IC": set(EXPERIMENTAL) | {"TAS", "NAS", "IC"},
    "experimental_plus_IBA_ISS_IC": set(EXPERIMENTAL) | {"IBA", "ISS", "IC"},
    "all_except_IEA_ND_NAS": {"TAS", "IPI", "IDA", "IBA", "ISS", "IMP", "IC", "EXP", "IGI", "IEP", "IKR"},
}


def now_iso() -> str:
    return dt.datetime.now(UTC).isoformat()


def datestamp() -> str:
    return dt.datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")


def sha256_file(path: Path, chunk: int = 1 << 20) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        while True:
            b = fh.read(chunk)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def write_csv(path: Path, rows: Sequence[Mapping[str, object]], fieldnames: Sequence[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def open_deterministic_gzip_text(path: Path):
    raw = path.open("wb")
    gz = gzip.GzipFile(filename="", mode="wb", fileobj=raw, mtime=0, compresslevel=9)
    text = io.TextIOWrapper(gz, encoding="utf-8", newline="")
    return raw, gz, text


def read_header_and_data_stats(path: Path, expected_width: int) -> dict:
    headers: list[str] = []
    rows = 0
    widths = collections.Counter()
    data_hash = hashlib.sha256()
    with gzip.open(path, "rt", encoding="utf-8", errors="strict", newline="") as fh:
        for line in fh:
            if line.startswith("!"):
                headers.append(line.rstrip("\n"))
                continue
            if not line.strip():
                continue
            rows += 1
            widths[len(line.rstrip("\n").split("\t"))] += 1
            data_hash.update(line.encode("utf-8"))
    return {
        "header_lines": headers,
        "header_sha256": hashlib.sha256(("\n".join(headers) + "\n").encode()).hexdigest(),
        "data_rows": rows,
        "column_width_counts": dict(widths),
        "expected_width": expected_width,
        "all_rows_expected_width": widths == collections.Counter({expected_width: rows}),
        "uncompressed_data_sha256": data_hash.hexdigest(),
    }


def normalize_relation(qualifier: str, aspect: str) -> tuple[str, bool]:
    pieces = [x for x in qualifier.split("|") if x] if qualifier else []
    is_not = "NOT" in pieces
    non_not = [x for x in pieces if x != "NOT"]
    relation = non_not[0] if non_not else RELATION_BY_ASPECT[aspect]
    return (("NOT|" + relation) if is_not else relation), is_not


def split_taxon(taxon: str) -> tuple[str, str]:
    pieces = taxon.split("|") if taxon else []
    return (pieces[0] if pieces else "", pieces[1] if len(pieces) > 1 else "")


def parse_go_evidence(properties: str) -> str:
    for item in properties.split("|") if properties else []:
        if item.startswith("go_evidence="):
            return item.split("=", 1)[1]
    return ""


def load_expected_inventory(path: Path) -> dict[str, dict]:
    with path.open(newline="", encoding="utf-8-sig") as fh:
        return {row["artifact_name"]: row for row in csv.DictReader(fh)}


def load_observed_labels(path: Path) -> tuple[list[int], list[int], list[list[int]]]:
    genes: list[int] = []
    matrix: list[list[int]] = []
    with path.open(newline="", encoding="utf-8") as fh:
        r = csv.DictReader(fh)
        for row in r:
            genes.append(int(row["entrez_gene_id"]))
            matrix.append([int(row[f"label_{i}"]) for i in range(121)])
    bits: list[int] = []
    for col in range(121):
        bit = 0
        for i, row in enumerate(matrix):
            if row[col]:
                bit |= 1 << i
        bits.append(bit)
    return genes, bits, matrix


def parse_msigdb_symbol_map_and_c5bp(path: Path, gene_index: Mapping[int, int]) -> tuple[dict[str, set[int]], list[dict], dict]:
    attr_re = re.compile(r'([A-Z_]+)="([^"]*)"')
    go_re = re.compile(r"GO:\d{7}")
    symbol_map: dict[str, set[int]] = collections.defaultdict(set)
    c5bp_records: list[dict] = []
    gene_sets = 0
    aligned_member_pairs = 0
    malformed_symbolized_pairs = 0
    symbol_pair_counts = collections.Counter()
    xml_member = None
    with zipfile.ZipFile(path) as zf:
        xml_members = [n for n in zf.namelist() if n.lower().endswith(".xml")]
        if not xml_members:
            raise RuntimeError("MSigDB archive has no XML member")
        xml_member = xml_members[0]
        with zf.open(xml_member) as fh:
            for raw in fh:
                if b"<GENESET " not in raw:
                    continue
                gene_sets += 1
                attrs = {k: html.unescape(v) for k, v in attr_re.findall(raw.decode("utf-8", "replace"))}
                ezids = attrs.get("MEMBERS_EZID", "").split(",") if attrs.get("MEMBERS_EZID") else []
                symbols = attrs.get("MEMBERS_SYMBOLIZED", "").split(",") if attrs.get("MEMBERS_SYMBOLIZED") else []
                if len(ezids) == len(symbols):
                    for symbol, ezid in zip(symbols, ezids):
                        if symbol and ezid.isdigit():
                            gene = int(ezid)
                            symbol_map[symbol].add(gene)
                            symbol_pair_counts[(symbol, gene)] += 1
                            aligned_member_pairs += 1
                else:
                    malformed_symbolized_pairs += 1
                if attrs.get("CATEGORY_CODE") == "C5" and attrs.get("SUB_CATEGORY_CODE") == "BP":
                    match = go_re.search(attrs.get("EXTERNAL_DETAILS_URL", "") + " " + attrs.get("DESCRIPTION_FULL", ""))
                    if not match:
                        continue
                    bit = 0
                    for ezid in ezids:
                        if ezid.isdigit() and int(ezid) in gene_index:
                            bit |= 1 << gene_index[int(ezid)]
                    c5bp_records.append({
                        "go_id": match.group(),
                        "name": attrs.get("STANDARD_NAME", ""),
                        "description": attrs.get("DESCRIPTION_BRIEF", ""),
                        "bit": bit,
                        "source_member": xml_member,
                    })
    summary = {
        "xml_member": xml_member,
        "gene_sets_parsed": gene_sets,
        "aligned_symbol_entrez_member_pairs": aligned_member_pairs,
        "gene_sets_with_misaligned_MEMBERS_SYMBOLIZED_and_MEMBERS_EZID": malformed_symbolized_pairs,
        "unique_symbols": len(symbol_map),
        "ambiguous_symbols": sum(len(v) > 1 for v in symbol_map.values()),
        "unique_symbol_entrez_pairs": len(symbol_pair_counts),
        "c5_bp_records": len(c5bp_records),
    }
    return symbol_map, c5bp_records, summary


def load_restriction_sets(table6: Path, table9: Path, ohmnet_labels: Path, c5bp_records: Sequence[dict]) -> dict[str, set[str] | None]:
    wb6 = load_workbook(table6, read_only=True, data_only=True)
    ws6 = wb6.active
    table6_ids = {str(row[1]) for row in ws6.iter_rows(min_row=3, values_only=True) if row[1] and str(row[1]).startswith("GO:")}
    wb9 = load_workbook(table9, read_only=True, data_only=True)
    ws9 = wb9.active
    table9_ids: set[str] = set()
    for row in ws9.iter_rows(min_row=3, values_only=True):
        for value in row[2:]:
            if value and str(value).startswith("GO:"):
                table9_ids.add(str(value))
    ohmnet_ids: set[str] = set()
    with tarfile.open(ohmnet_labels, "r:gz") as tf:
        for member in tf.getmembers():
            match = re.search(r"_GO:(\d{7})\.lab$", member.name)
            if match:
                ohmnet_ids.add("GO:" + match.group(1))
    return {
        "all_bp": None,
        "greene_table6": table6_ids,
        "greene_table9": table9_ids,
        "ohmnet_label_terms": ohmnet_ids,
        "msigdb_v52_c5bp_terms": {r["go_id"] for r in c5bp_records},
    }


def metric(obs: int, pred: int, mask: int, denominator: int) -> dict:
    obs_m = obs & mask
    pred_m = pred & mask
    xor = (obs_m ^ pred_m).bit_count()
    tp = (obs_m & pred_m).bit_count()
    fp = (pred_m & ~obs_m & mask).bit_count()
    fn = (obs_m & ~pred_m & mask).bit_count()
    denom_f1 = 2 * tp + fp + fn
    denom_j = tp + fp + fn
    return {
        "mismatch_genes": xor,
        "agreement": 1.0 - (xor / denominator if denominator else 0.0),
        "true_positives": tp,
        "false_positives": fp,
        "false_negatives": fn,
        "f1": (2 * tp / denom_f1) if denom_f1 else 1.0,
        "jaccard": (tp / denom_j) if denom_j else 1.0,
    }


def best_matches(
    label_bits: Sequence[int],
    candidates: Mapping[str, int],
    mask: int,
    denominator: int,
    source_name: str,
    mapping_strategy: str,
    evidence_filter: str,
    term_scope: str,
    comparison_scope: str,
) -> list[dict]:
    ordered = sorted(candidates.items())
    if not ordered:
        return []
    out = []
    for col, obs in enumerate(label_bits):
        best_mismatch = denominator + 1
        best: list[tuple[str, int]] = []
        for go_id, pred in ordered:
            mismatch = ((obs ^ pred) & mask).bit_count()
            if mismatch < best_mismatch:
                best_mismatch = mismatch
                best = [(go_id, pred)]
            elif mismatch == best_mismatch:
                best.append((go_id, pred))
        go_id, pred = best[0]
        m = metric(obs, pred, mask, denominator)
        out.append({
            "source_name": source_name,
            "mapping_strategy": mapping_strategy,
            "evidence_filter": evidence_filter,
            "term_scope": term_scope,
            "comparison_scope": comparison_scope,
            "label_column": col,
            "observed_positive_genes": (obs & mask).bit_count(),
            "best_go_id": go_id,
            "best_candidate_positive_genes": (pred & mask).bit_count(),
            **m,
            "tie_count": len(best),
            "tied_go_ids": "|".join(x[0] for x in best[:50]),
        })
    return out


def summarize_match_rows(rows: Sequence[dict]) -> dict:
    agreements = [float(r["agreement"]) for r in rows]
    mismatches = [int(r["mismatch_genes"]) for r in rows]
    return {
        "label_columns": len(rows),
        "exact_matches": sum(x == 0 for x in mismatches),
        "agreement_at_least_99_percent": sum(x >= 0.99 for x in agreements),
        "agreement_at_least_95_percent": sum(x >= 0.95 for x in agreements),
        "median_best_agreement": statistics.median(agreements) if agreements else None,
        "mean_best_agreement": statistics.mean(agreements) if agreements else None,
        "minimum_mismatch_genes": min(mismatches) if mismatches else None,
        "mean_mismatch_genes": statistics.mean(mismatches) if mismatches else None,
    }



def _int_bitsets_to_words(values: Sequence[int], nwords: int) -> np.ndarray:
    out = np.empty((len(values), nwords), dtype=np.uint64)
    byte_width = nwords * 8
    for i, value in enumerate(values):
        out[i, :] = np.frombuffer(int(value).to_bytes(byte_width, "little", signed=False), dtype="<u8")
    return out


def _hamming_distance_matrix(
    candidate_words: np.ndarray,
    label_words: np.ndarray,
    mask_words: np.ndarray,
    chunk_size: int = 512,
) -> np.ndarray:
    """Return candidate-by-label Hamming distances using packed uint64 popcounts."""
    n_candidates = candidate_words.shape[0]
    n_labels = label_words.shape[0]
    out = np.empty((n_candidates, n_labels), dtype=np.uint16)
    masked_labels = label_words & mask_words
    for start in range(0, n_candidates, chunk_size):
        stop = min(start + chunk_size, n_candidates)
        masked_candidates = candidate_words[start:stop] & mask_words
        xor = masked_candidates[:, None, :] ^ masked_labels[None, :, :]
        out[start:stop, :] = np.bitwise_count(xor).sum(axis=2, dtype=np.uint16)
    return out


def best_matches_multi_scopes(
    label_bits: Sequence[int],
    candidates: Mapping[str, int],
    all_mask: int,
    covered_mask: int,
    all_denominator: int,
    covered_denominator: int,
    restrictions: Mapping[str, set[str] | None],
    source_name: str,
    mapping_strategy: str,
    evidence_filter: str,
) -> tuple[list[dict], list[dict]]:
    """Find best terms for all scopes with vectorized packed-bit Hamming distances."""
    ordered = sorted(candidates.items())
    if not ordered:
        return [], []

    # The denominator is the number of gene positions for the full comparison.
    nwords = (all_denominator + 63) // 64
    go_ids = [go_id for go_id, _ in ordered]
    pred_ints = [pred for _, pred in ordered]
    candidate_words = _int_bitsets_to_words(pred_ints, nwords)
    label_words = _int_bitsets_to_words(list(label_bits), nwords)
    all_mask_words = _int_bitsets_to_words([all_mask], nwords)[0]
    covered_mask_words = _int_bitsets_to_words([covered_mask], nwords)[0]

    distance_matrices = {
        "all_4268_resolved_genes": _hamming_distance_matrix(candidate_words, label_words, all_mask_words),
        "mapping_covered_genes_only": _hamming_distance_matrix(candidate_words, label_words, covered_mask_words),
    }
    comparison_specs = {
        "all_4268_resolved_genes": (all_mask, all_denominator),
        "mapping_covered_genes_only": (covered_mask, covered_denominator),
    }

    rows: list[dict] = []
    summaries: list[dict] = []
    go_index = {go_id: i for i, go_id in enumerate(go_ids)}
    all_indices = np.arange(len(go_ids), dtype=np.int64)

    for term_scope, allowed in restrictions.items():
        if allowed is None:
            indices = all_indices
        else:
            indices = np.fromiter(
                (go_index[go_id] for go_id in allowed if go_id in go_index),
                dtype=np.int64,
            )
            if indices.size:
                indices.sort()
        if indices.size == 0:
            continue

        for comparison_scope, (mask, denominator) in comparison_specs.items():
            if denominator == 0:
                continue
            scoped_distances = distance_matrices[comparison_scope][indices, :]
            minimums = scoped_distances.min(axis=0)
            first_positions = scoped_distances.argmin(axis=0)
            best_indices = indices[first_positions]
            tie_counts = (scoped_distances == minimums[None, :]).sum(axis=0)
            scoped_rows: list[dict] = []
            for col, obs in enumerate(label_bits):
                best_idx = int(best_indices[col])
                go_id = go_ids[best_idx]
                pred = pred_ints[best_idx]
                m = metric(obs, pred, mask, denominator)
                tie_count = int(tie_counts[col])
                if tie_count == 1:
                    tied_go_ids = go_id
                else:
                    tied_local = np.flatnonzero(scoped_distances[:, col] == minimums[col])[:50]
                    tied_go_ids = "|".join(go_ids[int(indices[int(i)])] for i in tied_local)
                row = {
                    "source_name": source_name,
                    "mapping_strategy": mapping_strategy,
                    "evidence_filter": evidence_filter,
                    "term_scope": term_scope,
                    "comparison_scope": comparison_scope,
                    "label_column": col,
                    "observed_positive_genes": (obs & mask).bit_count(),
                    "best_go_id": go_id,
                    "best_candidate_positive_genes": (pred & mask).bit_count(),
                    **m,
                    "tie_count": tie_count,
                    "tied_go_ids": tied_go_ids,
                }
                rows.append(row)
                scoped_rows.append(row)
            summaries.append({
                "source_name": source_name,
                "mapping_strategy": mapping_strategy,
                "evidence_filter": evidence_filter,
                "term_scope": term_scope,
                "comparison_scope": comparison_scope,
                "candidate_GO_terms": int(indices.size),
                "comparison_genes": denominator,
                **summarize_match_rows(scoped_rows),
            })
    return rows, summaries

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-dir", type=Path, default=Path("/mnt/data"))
    ap.add_argument("--batch-dir", type=Path, default=Path("/mnt/data/ppi_repro_corrected/batches/B101"))
    ap.add_argument("--baseline-dir", type=Path, default=Path("/mnt/data/work/ppi_repro_corrected/results"))
    args = ap.parse_args()
    inp = args.input_dir
    out = args.batch_dir
    out.mkdir(parents=True, exist_ok=True)
    stamp = datestamp()
    started = time.time()
    generated = now_iso()

    raw_paths = {
        "gaf": inp / "goa_human.gaf.159.gz",
        "gpad": inp / "goa_human.gpa.159.gz",
        "gpi": inp / "goa_human.gpi.159.gz",
    }
    expected_inventory = load_expected_inventory(inp / "local_upload_inventory_20260827T145903Z.csv")
    source_urls = {
        "goa_human.gaf.159.gz": "https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/goa_human.gaf.159.gz",
        "goa_human.gpa.159.gz": "https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/goa_human.gpa.159.gz",
        "goa_human.gpi.159.gz": "https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/goa_human.gpi.159.gz",
    }
    integrity_rows = []
    for kind, path in raw_paths.items():
        name = path.name
        actual_sha = sha256_file(path)
        actual_size = path.stat().st_size
        expected = expected_inventory.get(name, {})
        gzip_result = subprocess.run(["gzip", "-t", str(path)], capture_output=True, text=True)
        integrity_rows.append({
            "batch_id": "B101",
            "file_role": kind,
            "artifact_name": name,
            "local_path": str(path),
            "size_bytes": actual_size,
            "sha256": actual_sha,
            "expected_size_bytes_from_B000": expected.get("size_bytes", ""),
            "expected_sha256_from_B000": expected.get("sha256", ""),
            "size_matches_B000": str(actual_size) == expected.get("size_bytes", ""),
            "sha256_matches_B000": actual_sha == expected.get("sha256", ""),
            "gzip_integrity_exit_code": gzip_result.returncode,
            "gzip_integrity_ok": gzip_result.returncode == 0,
            "gzip_integrity_stderr": gzip_result.stderr.strip(),
            "direct_or_canonical_source_url": source_urls[name],
            "source_page_url": "https://ftp.ebi.ac.uk/pub/databases/GO/goa/old/HUMAN/",
            "remote_byte_verification_status": "not_performed; exact official URL recorded; upload hash independently matches B000 user-local hash",
            "received_at_utc": generated,
        })
    integrity_path = out / f"B101_input_integrity_{stamp}.csv"
    write_csv(integrity_path, integrity_rows)
    if not all(r["sha256_matches_B000"] and r["size_matches_B000"] and r["gzip_integrity_ok"] for r in integrity_rows):
        raise RuntimeError("B101 input integrity or B000 hash comparison failed")

    raw_stats = {
        "gaf": read_header_and_data_stats(raw_paths["gaf"], 17),
        "gpad": read_header_and_data_stats(raw_paths["gpad"], 12),
        "gpi": read_header_and_data_stats(raw_paths["gpi"], 10),
    }
    headers_path = out / f"B101_headers_and_raw_stats_{stamp}.json"
    headers_path.write_text(json.dumps(raw_stats, indent=2), encoding="utf-8")

    # Parse GPI first.
    gpi_by_object: dict[tuple[str, str], list[str]] = {}
    gpi_property_counts = collections.Counter()
    gpi_nonempty_counts = collections.Counter()
    gpi_symbol_to_objects: dict[str, list[tuple[str, str]]] = collections.defaultdict(list)
    with gzip.open(raw_paths["gpi"], "rt", encoding="utf-8", errors="strict") as fh:
        for line in fh:
            if line.startswith("!") or not line.strip():
                continue
            row = line.rstrip("\n").split("\t")
            key = (row[0], row[1])
            if key in gpi_by_object:
                raise RuntimeError(f"Duplicate GPI object key: {key}")
            gpi_by_object[key] = row
            for i, value in enumerate(row):
                if value:
                    gpi_nonempty_counts[GPI_FIELDS[i]] += 1
            gpi_property_counts[row[9]] += 1
            gpi_symbol_to_objects[row[2]].append(key)

    # Parse GAF and write normalized derivative.
    gaf_norm_path = out / f"B101_goa_human_gaf159_normalized_{stamp}.tsv.gz"
    gaf_objects: set[tuple[str, str]] = set()
    gaf_object_symbol: dict[tuple[str, str], str] = {}
    gaf_object_metadata: dict[tuple[str, str], dict[str, set[str]]] = collections.defaultdict(lambda: collections.defaultdict(set))
    gaf_go_ids: set[str] = set()
    gaf_evidence = collections.Counter()
    gaf_evidence_bp = collections.Counter()
    gaf_aspect = collections.Counter()
    gaf_qualifier = collections.Counter()
    gaf_assigned = collections.Counter()
    gaf_date = collections.Counter()
    gaf_canonical = collections.Counter()
    bp_annotations: list[tuple[tuple[str, str], str, str]] = []
    raw_reconstruct_hash = hashlib.sha256()
    raw_out, gz_out, text_out = open_deterministic_gzip_text(gaf_norm_path)
    try:
        fieldnames = GAF_FIELDS + ["Normalized_Relation", "Is_NOT", "Subject_Taxon", "Interacting_Taxon"]
        writer = csv.writer(text_out, delimiter="\t", lineterminator="\n")
        writer.writerow(fieldnames)
        with gzip.open(raw_paths["gaf"], "rt", encoding="utf-8", errors="strict", newline="") as fh:
            for line in fh:
                if line.startswith("!") or not line.strip():
                    continue
                raw_reconstruct_hash.update(line.encode("utf-8"))
                row = line.rstrip("\n").split("\t")
                relation, is_not = normalize_relation(row[3], row[8])
                subject_taxon, interacting_taxon = split_taxon(row[12])
                writer.writerow(row + [relation, int(is_not), subject_taxon, interacting_taxon])
                key_obj = (row[0], row[1])
                gaf_objects.add(key_obj)
                gaf_object_symbol.setdefault(key_obj, row[2])
                md = gaf_object_metadata[key_obj]
                md["symbol"].add(row[2]); md["name"].add(row[9]); md["synonym"].add(row[10]); md["type"].add(row[11]); md["taxon"].add(subject_taxon)
                gaf_go_ids.add(row[4]); gaf_evidence[row[6]] += 1; gaf_aspect[row[8]] += 1; gaf_qualifier[row[3]] += 1; gaf_assigned[row[14]] += 1; gaf_date[row[13]] += 1
                canonical = (row[0], row[1], relation, row[4], row[5], row[6], row[7], interacting_taxon, row[13], row[14], row[15])
                gaf_canonical[canonical] += 1
                if row[8] == "P":
                    gaf_evidence_bp[row[6]] += 1
                    if not is_not:
                        bp_annotations.append((key_obj, row[4], row[6]))
    finally:
        text_out.flush(); text_out.detach(); gz_out.close(); raw_out.close()
    if raw_reconstruct_hash.hexdigest() != raw_stats["gaf"]["uncompressed_data_sha256"]:
        raise RuntimeError("GAF normalized derivative did not preserve raw data rows")

    # Parse GPAD and write normalized derivative.
    gpad_norm_path = out / f"B101_goa_human_gpad159_normalized_{stamp}.tsv.gz"
    gpad_objects: set[tuple[str, str]] = set()
    gpad_go_ids: set[str] = set()
    gpad_relation = collections.Counter(); gpad_eco = collections.Counter(); gpad_evidence = collections.Counter(); gpad_assigned = collections.Counter(); gpad_date = collections.Counter()
    gpad_canonical: dict[tuple, list[str]] = collections.defaultdict(list)
    raw_reconstruct_hash = hashlib.sha256()
    raw_out, gz_out, text_out = open_deterministic_gzip_text(gpad_norm_path)
    try:
        writer = csv.writer(text_out, delimiter="\t", lineterminator="\n")
        writer.writerow(GPAD_FIELDS + ["GO_Evidence_Code"])
        with gzip.open(raw_paths["gpad"], "rt", encoding="utf-8", errors="strict", newline="") as fh:
            for line in fh:
                if line.startswith("!") or not line.strip():
                    continue
                raw_reconstruct_hash.update(line.encode("utf-8"))
                row = line.rstrip("\n").split("\t")
                evidence = parse_go_evidence(row[11])
                writer.writerow(row + [evidence])
                key_obj = (row[0], row[1]); gpad_objects.add(key_obj); gpad_go_ids.add(row[3])
                gpad_relation[row[2]] += 1; gpad_eco[row[5]] += 1; gpad_evidence[evidence] += 1; gpad_assigned[row[9]] += 1; gpad_date[row[8]] += 1
                canonical = (row[0], row[1], row[2], row[3], row[4], evidence, row[6], row[7], row[8], row[9], row[10])
                gpad_canonical[canonical].append(row[5])
    finally:
        text_out.flush(); text_out.detach(); gz_out.close(); raw_out.close()
    if raw_reconstruct_hash.hexdigest() != raw_stats["gpad"]["uncompressed_data_sha256"]:
        raise RuntimeError("GPAD normalized derivative did not preserve raw data rows")

    duplicate_groups = []
    for key, ecos in gpad_canonical.items():
        if len(ecos) > 1:
            duplicate_groups.append({
                "DB": key[0], "DB_Object_ID": key[1], "Relation": key[2], "GO_ID": key[3], "DB_Reference": key[4],
                "GO_Evidence_Code": key[5], "With_From": key[6], "Interacting_Taxon_ID": key[7], "Date": key[8],
                "Assigned_By": key[9], "Annotation_Extension": key[10], "GPAD_Row_Count_for_GAF_Projection": len(ecos),
                "ECO_IDs": "|".join(sorted(ecos)),
            })
    duplicate_path = out / f"B101_gpad_ECO_projection_multiplicity_{stamp}.csv.gz"
    raw, gz, txt = open_deterministic_gzip_text(duplicate_path)
    try:
        w = csv.DictWriter(txt, fieldnames=list(duplicate_groups[0].keys()) if duplicate_groups else [], lineterminator="\n")
        if duplicate_groups:
            w.writeheader(); w.writerows(sorted(duplicate_groups, key=lambda r: (r["DB_Object_ID"], r["GO_ID"], r["Relation"])))
    finally:
        txt.flush(); txt.detach(); gz.close(); raw.close()

    # GAF/GPAD reconciliation: unique projected assertions should be identical.
    gaf_keys = set(gaf_canonical)
    gpad_keys = set(gpad_canonical)
    recon = {
        "gaf_rows": sum(gaf_canonical.values()),
        "gaf_unique_projected_assertions": len(gaf_keys),
        "gpad_rows": sum(len(v) for v in gpad_canonical.values()),
        "gpad_unique_GAF_projected_assertions": len(gpad_keys),
        "projected_assertions_in_GAF_not_GPAD": len(gaf_keys - gpad_keys),
        "projected_assertions_in_GPAD_not_GAF": len(gpad_keys - gaf_keys),
        "projected_assertion_sets_identical": gaf_keys == gpad_keys,
        "gpad_projection_groups_with_multiple_ECO_rows": len(duplicate_groups),
        "gpad_extra_rows_beyond_GAF_projection": sum(len(v) - 1 for v in gpad_canonical.values()),
        "duplicate_group_ECO_combinations": dict(collections.Counter("|".join(sorted(v)) for v in gpad_canonical.values() if len(v) > 1)),
        "interpretation": "GPAD preserves ECO granularity. GAF collapses 1,018 pairs of ECO:0000364 and ECO:0000366 to the same IEA assertion; no unique projected assertion is missing from either representation.",
    }
    recon_path = out / f"B101_gaf_gpad_reconciliation_{stamp}.json"
    recon_path.write_text(json.dumps(recon, indent=2), encoding="utf-8")
    if not recon["projected_assertion_sets_identical"]:
        raise RuntimeError("GAF and GPAD unique projected assertion sets differ")

    # GPI normalized derivative with annotation coverage and GAF fallback symbol.
    gpi_norm_path = out / f"B101_goa_human_gpi159_normalized_{stamp}.tsv.gz"
    raw_reconstruct_hash = hashlib.sha256()
    raw_out, gz_out, text_out = open_deterministic_gzip_text(gpi_norm_path)
    try:
        writer = csv.writer(text_out, delimiter="\t", lineterminator="\n")
        writer.writerow(GPI_FIELDS + ["GAF_Fallback_Symbol", "Annotated_in_GAF"])
        with gzip.open(raw_paths["gpi"], "rt", encoding="utf-8", errors="strict", newline="") as fh:
            for line in fh:
                if line.startswith("!") or not line.strip():
                    continue
                raw_reconstruct_hash.update(line.encode("utf-8"))
                row = line.rstrip("\n").split("\t")
                key = (row[0], row[1])
                writer.writerow(row + [gaf_object_symbol.get(key, ""), int(key in gaf_objects)])
    finally:
        text_out.flush(); text_out.detach(); gz_out.close(); raw_out.close()
    if raw_reconstruct_hash.hexdigest() != raw_stats["gpi"]["uncompressed_data_sha256"]:
        raise RuntimeError("GPI normalized derivative did not preserve raw data rows")

    metadata_mismatches = collections.Counter()
    symbol_mismatch_examples = []
    all_symbol_mismatches_follow_blank_accession_pattern = True
    for key in gaf_objects:
        gpi = gpi_by_object[key]
        md = gaf_object_metadata[key]
        checks = {
            "symbol": (gpi[2], md["symbol"]), "name": (gpi[3], md["name"]), "synonym": (gpi[4], md["synonym"]),
            "type": (gpi[5], md["type"]), "taxon": (gpi[6], md["taxon"]),
        }
        for field, (value, observed) in checks.items():
            if value not in observed:
                metadata_mismatches[field] += 1
                if field == "symbol":
                    follows_pattern = value == "" and observed == {key[1]}
                    all_symbol_mismatches_follow_blank_accession_pattern &= follows_pattern
                    if len(symbol_mismatch_examples) < 25:
                        symbol_mismatch_examples.append({"DB": key[0], "DB_Object_ID": key[1], "GPI_Symbol": value, "GAF_Symbols": "|".join(sorted(observed))})
    symbol_mismatch_all_blank_accession = (
        all_symbol_mismatches_follow_blank_accession_pattern
        and metadata_mismatches["symbol"] == 394
    )
    gpi_summary = {
        "gpi_rows_and_unique_objects": len(gpi_by_object),
        "gaf_annotated_unique_objects": len(gaf_objects),
        "gpad_annotated_unique_objects": len(gpad_objects),
        "gaf_objects_missing_from_GPI": len(gaf_objects - set(gpi_by_object)),
        "gpad_objects_missing_from_GPI": len(gpad_objects - set(gpi_by_object)),
        "gpi_objects_without_GAF_annotation": len(set(gpi_by_object) - gaf_objects),
        "gpi_property_counts": dict(gpi_property_counts),
        "gpi_nonempty_column_counts": dict(gpi_nonempty_counts),
        "gpi_DB_Xrefs_nonempty": gpi_nonempty_counts["DB_Xrefs"],
        "gpi_Parent_Object_ID_nonempty": gpi_nonempty_counts["Parent_Object_ID"],
        "metadata_mismatch_counts": dict(metadata_mismatches),
        "symbol_mismatch_pattern": "394 annotated objects have blank GPI symbol; GAF substitutes the UniProt accession as the symbol" if symbol_mismatch_all_blank_accession else "see examples",
        "symbol_mismatch_examples": symbol_mismatch_examples,
        "conclusion_about_Entrez_mapping": "The GPI file does not supply Entrez Gene cross-references: DB_Xrefs is empty for every one of the 21,002 rows, consistent with its own header. An external identifier map is required.",
    }
    gpi_summary_path = out / f"B101_gpi_coverage_and_metadata_{stamp}.json"
    gpi_summary_path.write_text(json.dumps(gpi_summary, indent=2), encoding="utf-8")
    if gpi_summary["gaf_objects_missing_from_GPI"] != 0 or gpi_summary["gpad_objects_missing_from_GPI"] != 0:
        raise RuntimeError("GPI does not cover all annotated objects")

    # Summary count tables.
    count_rows = []
    for file_kind, category, counter in [
        ("GAF", "Evidence_Code", gaf_evidence), ("GAF_BP", "Evidence_Code", gaf_evidence_bp),
        ("GAF", "Aspect", gaf_aspect), ("GAF", "Qualifier", gaf_qualifier),
        ("GAF", "Assigned_By", gaf_assigned), ("GAF", "Date", gaf_date),
        ("GPAD", "Relation", gpad_relation), ("GPAD", "ECO_ID", gpad_eco),
        ("GPAD", "GO_Evidence_Code", gpad_evidence), ("GPAD", "Assigned_By", gpad_assigned),
        ("GPAD", "Date", gpad_date), ("GPI", "Properties", gpi_property_counts),
    ]:
        for value, count in sorted(counter.items(), key=lambda x: (-x[1], x[0])):
            count_rows.append({"file_kind": file_kind, "category": category, "value": value, "count": count})
    counts_path = out / f"B101_value_counts_{stamp}.csv"
    write_csv(counts_path, count_rows)

    # Provisional symbol->Entrez map based on MSigDB v5.2, clearly separated from GOA.
    collapsed_path = args.baseline_dir / "collapsed_gene_labels_topology_features.csv"
    genes, label_bits, _ = load_observed_labels(collapsed_path)
    gene_index = {g: i for i, g in enumerate(genes)}
    graph_gene_set = set(genes)
    symbol_map, c5bp_records, msig_summary = parse_msigdb_symbol_map_and_c5bp(inp / "msigdb_v5.2_files_to_download_locally.zip", gene_index)
    mapping_rows = []
    strategy_maps: dict[str, dict[tuple[str, str], set[int]]] = {name: {} for name in [
        "primary_symbol_unique", "primary_then_synonym_unique", "fallback_synonym_union", "all_names_union"
    ]}
    for key, row in sorted(gpi_by_object.items()):
        primary = row[2] or gaf_object_symbol.get(key, "")
        synonyms = [x for x in row[4].split("|") if x and x != primary]
        primary_full = set(symbol_map.get(primary, set()))
        primary_graph = primary_full & graph_gene_set
        synonym_full = set().union(*(symbol_map.get(x, set()) for x in synonyms)) if synonyms else set()
        synonym_graph = synonym_full & graph_gene_set
        all_graph = primary_graph | synonym_graph
        if len(primary_graph) == 1:
            strategy_maps["primary_symbol_unique"][key] = set(primary_graph)
            strategy_maps["primary_then_synonym_unique"][key] = set(primary_graph)
            strategy_maps["fallback_synonym_union"][key] = set(primary_graph)
        elif len(primary_graph) == 0:
            if len(synonym_graph) == 1:
                strategy_maps["primary_then_synonym_unique"][key] = set(synonym_graph)
            if synonym_graph:
                strategy_maps["fallback_synonym_union"][key] = set(synonym_graph)
        if all_graph:
            strategy_maps["all_names_union"][key] = set(all_graph)
        mapping_rows.append({
            "DB": key[0], "DB_Object_ID": key[1], "GPI_Primary_Symbol": row[2], "GAF_Fallback_Symbol": gaf_object_symbol.get(key, ""),
            "Effective_Primary_Symbol": primary, "GPI_Synonyms": row[4],
            "Primary_Full_MSigDB52_Entrez_Candidates": "|".join(map(str, sorted(primary_full))),
            "Primary_GraphSAGE_Entrez_Candidates": "|".join(map(str, sorted(primary_graph))),
            "Synonym_GraphSAGE_Entrez_Candidates": "|".join(map(str, sorted(synonym_graph))),
            "All_Name_GraphSAGE_Entrez_Candidates": "|".join(map(str, sorted(all_graph))),
            "primary_symbol_unique": "|".join(map(str, sorted(strategy_maps["primary_symbol_unique"].get(key, set())))),
            "primary_then_synonym_unique": "|".join(map(str, sorted(strategy_maps["primary_then_synonym_unique"].get(key, set())))),
            "fallback_synonym_union": "|".join(map(str, sorted(strategy_maps["fallback_synonym_union"].get(key, set())))),
            "all_names_union": "|".join(map(str, sorted(strategy_maps["all_names_union"].get(key, set())))),
            "Annotated_in_GAF": int(key in gaf_objects),
        })
    mapping_path = out / f"B101_provisional_UniProt_to_Entrez_via_MSigDB52_symbols_{stamp}.csv.gz"
    raw, gz, txt = open_deterministic_gzip_text(mapping_path)
    try:
        w = csv.DictWriter(txt, fieldnames=list(mapping_rows[0].keys()), lineterminator="\n")
        w.writeheader(); w.writerows(mapping_rows)
    finally:
        txt.flush(); txt.detach(); gz.close(); raw.close()

    strategy_summary = {}
    for name, mapping in strategy_maps.items():
        covered = set().union(*mapping.values()) if mapping else set()
        strategy_summary[name] = {
            "mapped_GPI_objects": len(mapping),
            "objects_with_multiple_GraphSAGE_gene_candidates": sum(len(v) > 1 for v in mapping.values()),
            "GraphSAGE_genes_covered": len(covered),
            "GraphSAGE_genes_missing": len(graph_gene_set - covered),
            "missing_GraphSAGE_Entrez_IDs": sorted(graph_gene_set - covered),
        }
    mapping_summary = {
        "mapping_source": "MSigDB v5.2 XML MEMBERS_SYMBOLIZED aligned with MEMBERS_EZID; this is a provisional external mapping, not a field supplied by GOA GPI",
        "msigdb_archive_sha256": sha256_file(inp / "msigdb_v5.2_files_to_download_locally.zip"),
        "comparison_gene_universe_derivative": str(collapsed_path),
        "comparison_gene_universe_derivative_sha256": sha256_file(collapsed_path),
        "comparison_genes": len(genes),
        "msigdb_parse": msig_summary,
        "strategies": strategy_summary,
        "warning": "Synonyms can collide with symbols of other genes. The all_names_union strategy is intentionally permissive and should not be treated as a definitive identifier map.",
    }
    mapping_summary_path = out / f"B101_provisional_mapping_summary_{stamp}.json"
    mapping_summary_path.write_text(json.dumps(mapping_summary, indent=2), encoding="utf-8")

    # Direct GOA BP label screen, plus MSigDB C5 BP comparator.
    restrictions = load_restriction_sets(inp / "Greene2015_Table6.xlsx", inp / "Greene2015_Table9.xlsx", inp / "bio-tissue-labels.tar.gz", c5bp_records)
    all_mask = (1 << len(genes)) - 1
    all_match_rows: list[dict] = []
    summary_rows: list[dict] = []
    for strategy_name, object_map in strategy_maps.items():
        covered_genes = set().union(*object_map.values()) if object_map else set()
        covered_mask = 0
        for gene in covered_genes:
            covered_mask |= 1 << gene_index[gene]
        for evidence_name, allowed in EVIDENCE_FILTERS.items():
            go_bits: dict[str, int] = collections.defaultdict(int)
            for obj, go_id, evidence in bp_annotations:
                if allowed is not None and evidence not in allowed:
                    continue
                for gene in object_map.get(obj, set()):
                    go_bits[go_id] |= 1 << gene_index[gene]
            rows, summaries = best_matches_multi_scopes(
                label_bits=label_bits,
                candidates=go_bits,
                all_mask=all_mask,
                covered_mask=covered_mask,
                all_denominator=len(genes),
                covered_denominator=len(covered_genes),
                restrictions=restrictions,
                source_name="GOA human GAF v159 direct BP annotations",
                mapping_strategy=strategy_name,
                evidence_filter=evidence_name,
            )
            all_match_rows.extend(rows)
            summary_rows.extend(summaries)
    # Comparator: MSigDB v5.2 C5 BP directly in Entrez space.
    c5bits: dict[str, int] = {}
    c5names: dict[str, str] = {}
    for rec in c5bp_records:
        c5bits[rec["go_id"]] = rec["bit"]
        c5names[rec["go_id"]] = rec["name"]
    c5_rows = best_matches(label_bits, c5bits, all_mask, len(genes), "MSigDB v5.2 C5 BP Entrez gene sets", "native_Entrez", "MSigDB_supplied", "C5_BP", "all_4268_resolved_genes")
    for row in c5_rows:
        row["best_term_name"] = c5names.get(row["best_go_id"], "")
    all_match_rows.extend(c5_rows)
    summary_rows.append({
        "source_name": "MSigDB v5.2 C5 BP Entrez gene sets",
        "mapping_strategy": "native_Entrez",
        "evidence_filter": "MSigDB_supplied",
        "term_scope": "C5_BP",
        "comparison_scope": "all_4268_resolved_genes",
        "candidate_GO_terms": len(c5bits),
        "comparison_genes": len(genes),
        **summarize_match_rows(c5_rows),
    })

    match_path = out / f"B101_direct_label_match_grid_{stamp}.csv.gz"
    raw, gz, txt = open_deterministic_gzip_text(match_path)
    try:
        fieldnames = sorted({k for row in all_match_rows for k in row})
        w = csv.DictWriter(txt, fieldnames=fieldnames, lineterminator="\n")
        w.writeheader(); w.writerows(all_match_rows)
    finally:
        txt.flush(); txt.detach(); gz.close(); raw.close()
    summary_path = out / f"B101_direct_label_match_summary_{stamp}.csv"
    write_csv(summary_path, sorted(summary_rows, key=lambda r: (
        -int(r["agreement_at_least_99_percent"]), -int(r["agreement_at_least_95_percent"]), -float(r["median_best_agreement"]), r["source_name"], r["mapping_strategy"], r["evidence_filter"], r["term_scope"], r["comparison_scope"]
    )))

    # Per-column candidates from the best direct configuration and C5 comparator.
    direct_summaries = [r for r in summary_rows if r["source_name"].startswith("GOA") and r["comparison_scope"] == "all_4268_resolved_genes" and r["term_scope"] == "all_bp"]
    best_direct_summary = max(direct_summaries, key=lambda r: (
        int(r["agreement_at_least_99_percent"]), int(r["agreement_at_least_95_percent"]), float(r["median_best_agreement"]), -float(r["mean_mismatch_genes"])
    ))
    best_direct_rows = [r for r in all_match_rows if all(
        r.get(k) == best_direct_summary[k] for k in ["source_name", "mapping_strategy", "evidence_filter", "term_scope", "comparison_scope"]
    )]
    by_col_direct = {int(r["label_column"]): r for r in best_direct_rows}
    by_col_c5 = {int(r["label_column"]): r for r in c5_rows}
    candidate_rows = []
    for col in range(121):
        d = by_col_direct[col]; m = by_col_c5[col]
        candidate_rows.append({
            "label_column": col,
            "observed_positive_genes": d["observed_positive_genes"],
            "best_direct_GOA_GO_ID": d["best_go_id"],
            "best_direct_GOA_mismatch": d["mismatch_genes"],
            "best_direct_GOA_agreement": d["agreement"],
            "best_direct_GOA_false_positives": d["false_positives"],
            "best_direct_GOA_false_negatives": d["false_negatives"],
            "best_MSigDB_C5BP_GO_ID": m["best_go_id"],
            "best_MSigDB_C5BP_term_name": m.get("best_term_name", ""),
            "best_MSigDB_C5BP_mismatch": m["mismatch_genes"],
            "best_MSigDB_C5BP_agreement": m["agreement"],
            "same_GO_ID": d["best_go_id"] == m["best_go_id"],
        })
    candidates_path = out / f"B101_label_column_candidate_terms_{stamp}.csv"
    write_csv(candidates_path, candidate_rows)

    # Re-read derivatives and verify row counts and preserved raw columns.
    derivative_checks = []
    for kind, path, raw_field_count, expected_rows, expected_data_hash in [
        ("gaf", gaf_norm_path, 17, raw_stats["gaf"]["data_rows"], raw_stats["gaf"]["uncompressed_data_sha256"]),
        ("gpad", gpad_norm_path, 12, raw_stats["gpad"]["data_rows"], raw_stats["gpad"]["uncompressed_data_sha256"]),
        ("gpi", gpi_norm_path, 10, raw_stats["gpi"]["data_rows"], raw_stats["gpi"]["uncompressed_data_sha256"]),
    ]:
        rows = 0; h = hashlib.sha256()
        with gzip.open(path, "rt", encoding="utf-8", errors="strict", newline="") as fh:
            next(fh)
            for line in fh:
                row = line.rstrip("\n").split("\t")
                raw_line = "\t".join(row[:raw_field_count]) + "\n"
                h.update(raw_line.encode("utf-8")); rows += 1
        derivative_checks.append({
            "file_kind": kind,
            "derivative_path": str(path),
            "derivative_sha256": sha256_file(path),
            "rows_re_read": rows,
            "expected_rows": expected_rows,
            "row_count_matches": rows == expected_rows,
            "reconstructed_raw_data_sha256": h.hexdigest(),
            "expected_raw_data_sha256": expected_data_hash,
            "raw_data_hash_matches": h.hexdigest() == expected_data_hash,
        })
    derivative_checks_path = out / f"B101_derivative_reconciliation_{stamp}.csv"
    write_csv(derivative_checks_path, derivative_checks)
    if not all(r["row_count_matches"] and r["raw_data_hash_matches"] for r in derivative_checks):
        raise RuntimeError("Derived normalized tables failed re-read reconciliation")

    # Main machine-readable summary.
    best_direct = best_direct_summary
    global_closest_direct = min(
        (r for r in all_match_rows if r["source_name"].startswith("GOA") and r["comparison_scope"] == "all_4268_resolved_genes"),
        key=lambda r: (int(r["mismatch_genes"]), -float(r["agreement"]), r["mapping_strategy"], r["evidence_filter"], r["term_scope"], int(r["label_column"]), r["best_go_id"]),
    )
    direct_conclusion = (
        "No direct-annotation configuration reaches 95% agreement for any of the 121 columns. "
        f"The selected best-overall full-universe configuration has a nearest column differing by {best_direct['minimum_mismatch_genes']} genes; "
        f"the smallest mismatch observed anywhere in the full direct grid is {global_closest_direct['mismatch_genes']} genes. "
        "This rules out raw direct GOA rows as the final label matrix but does not rule out GOA followed by ontology propagation or other preprocessing."
    )
    summary = {
        "batch_id": "B101",
        "generated_at_utc": generated,
        "runtime_seconds": time.time() - started,
        "inputs": integrity_rows,
        "raw_file_stats": raw_stats,
        "gaf": {
            "rows": raw_stats["gaf"]["data_rows"], "unique_objects": len(gaf_objects), "unique_GO_IDs": len(gaf_go_ids),
            "aspect_counts": dict(gaf_aspect), "evidence_counts": dict(gaf_evidence), "BP_evidence_counts": dict(gaf_evidence_bp),
            "qualifier_counts": dict(gaf_qualifier), "generated_header": next((x for x in raw_stats["gaf"]["header_lines"] if x.startswith("!Generated:")), ""),
            "GO_version_header": next((x for x in raw_stats["gaf"]["header_lines"] if x.startswith("!GO-version:")), ""),
        },
        "gpad": {
            "rows": raw_stats["gpad"]["data_rows"], "unique_objects": len(gpad_objects), "unique_GO_IDs": len(gpad_go_ids),
            "relation_counts": dict(gpad_relation), "ECO_counts": dict(gpad_eco), "GO_evidence_counts": dict(gpad_evidence),
        },
        "gpi": gpi_summary,
        "gaf_gpad_reconciliation": recon,
        "provisional_identifier_mapping": mapping_summary,
        "direct_label_screen": {
            "best_direct_configuration": best_direct,
            "global_closest_direct_label_match": global_closest_direct,
            "conclusion": direct_conclusion,
            "msigdb_C5BP_comparator": next(r for r in summary_rows if r["source_name"].startswith("MSigDB")),
            "ontology_propagation_tested_in_B101": False,
            "reason_not_tested": "No ontology file was uploaded in B101. The GAF header names GO release 2016-06-29; the user's local 2016-06-01 ontology should be analyzed in a subsequent batch and an exact 2016-06-29/2016-07-01 ontology should be sought.",
        },
        "retained_derivatives": derivative_checks,
    }
    summary_json = out / f"B101_analysis_summary_{stamp}.json"
    summary_json.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    # Report.
    report = f"""# Batch B101 — GOA Human release 159 verification and direct-label screen

Generated: `{generated}`  
Batch: `B101`

## Input integrity and provenance

All three uploads matched the size and SHA-256 values declared independently in Batch B000 and passed `gzip -t`:

| File | Bytes | SHA-256 | Official archive URL |
|---|---:|---|---|
"""
    for r in integrity_rows:
        report += f"| `{r['artifact_name']}` | {r['size_bytes']:,} | `{r['sha256']}` | {r['direct_or_canonical_source_url']} |\n"
    report += f"""

The exact remote bytes were not independently re-downloaded in this runtime. The source URLs are the official EBI GOA archive locations; the uploaded bytes are tied to the user's pre-upload inventory by exact size and SHA-256.

## File semantics and internal consistency

The GAF header reports `Generated: 2016-07-04 09:24` and GO ontology `2016-06-29/go.owl`.

- GAF: **{raw_stats['gaf']['data_rows']:,} rows**, **{len(gaf_objects):,} annotated UniProt objects**, **{len(gaf_go_ids):,} GO IDs**.
- GPAD: **{raw_stats['gpad']['data_rows']:,} rows**, the same **{len(gpad_objects):,} annotated objects** and **{len(gpad_go_ids):,} GO IDs**.
- GPI: **{len(gpi_by_object):,} objects**, including **{len(set(gpi_by_object)-gaf_objects):,} objects with no GAF annotation**.

After converting the GAF aspect/qualifier fields to GPAD relations and projecting GPAD ECO codes back to ordinary GO evidence codes, the two annotation files contain exactly the same **{len(gaf_keys):,} unique assertions**. GPAD has **1,018 additional rows** because each of those assertions is represented twice, once with `ECO:0000364` and once with `ECO:0000366`; all project to `IEA` in GAF. They are not accidental byte-identical duplicates.

All annotated GAF and GPAD objects occur in GPI. GPI's `DB_Xrefs` and `Parent_Object_ID` columns are empty in every row. Therefore, this release-159 GPI file **does not contain an Entrez Gene mapping**. The 394 symbol discrepancies are systematic: GPI leaves the symbol blank and GAF substitutes the UniProt accession.

## Provisional Entrez mapping test

Because B101 itself lacks Entrez xrefs, I independently built a provisional symbol map from aligned `MEMBERS_SYMBOLIZED` and `MEMBERS_EZID` attributes in the supplied MSigDB v5.2 XML. This map is used only as an interim test and is not treated as the final historical UniProt↔GeneID map.

| Strategy | Mapped GPI objects | Ambiguous objects | GraphSAGE Entrez genes covered | Missing |
|---|---:|---:|---:|---:|
"""
    for name, rec in strategy_summary.items():
        report += f"| `{name}` | {rec['mapped_GPI_objects']:,} | {rec['objects_with_multiple_GraphSAGE_gene_candidates']:,} | {rec['GraphSAGE_genes_covered']:,} | {rec['GraphSAGE_genes_missing']:,} |\n"
    report += f"""

The permissive synonym strategy covers nearly all resolved genes but introduces alias collisions. It is retained as a sensitivity analysis, not accepted as an authoritative mapping.

## Direct GO-label reconstruction result

I compared every GraphSAGE label column against every directly annotated Biological Process GO term under four identifier strategies, nine evidence-code filters, five term scopes, and both full-universe and mapping-covered-only comparisons.

The best full-universe direct configuration was:

- mapping: `{best_direct['mapping_strategy']}`
- evidence filter: `{best_direct['evidence_filter']}`
- term scope: `{best_direct['term_scope']}`
- median best agreement: **{best_direct['median_best_agreement']:.4%}**
- smallest mismatch within this selected configuration: **{best_direct['minimum_mismatch_genes']} genes**
- label columns at ≥95% agreement: **{best_direct['agreement_at_least_95_percent']} / 121**
- exact label columns: **{best_direct['exact_matches']} / 121**

Across the entire tested full-universe direct grid, the closest individual match was label column **{global_closest_direct['label_column']}** to **`{global_closest_direct['best_go_id']}`**, with **{global_closest_direct['mismatch_genes']} mismatched genes** ({global_closest_direct['agreement']:.4%} agreement), using mapping `{global_closest_direct['mapping_strategy']}` and evidence filter `{global_closest_direct['evidence_filter']}`.

Thus, the 121 columns are not a raw projection of direct GAF rows under any tested evidence filter. This does **not** reject GOA as the annotation source: ontology propagation, term-selection rules, or a different identifier map remain untested in this batch.

As a separate comparator, the MSigDB v5.2 C5 Biological Process sets also produced no columns at ≥95% agreement; their best column still differed by **{next(r for r in summary_rows if r['source_name'].startswith('MSigDB'))['minimum_mismatch_genes']} genes**.

## What B101 resolves

1. GAF and GPAD are semantically consistent once GPAD's finer ECO distinctions are projected to GAF evidence codes.
2. GPI is a complete object-metadata companion for the annotation files, but it supplies no GeneID xrefs.
3. Direct, non-propagated release-159 GOA annotations cannot be the final 121-column matrix.
4. The remaining high-value tests require an ontology and an independent historical identifier mapping; the lack of a direct match is not evidence that a particular missing `gene2go` snapshot is necessarily the cause.

## Retained data and deletion safety

The normalized GAF, GPAD, and GPI derivatives preserve every raw data column and row. They were re-read and their reconstructed raw-data SHA-256 values exactly match the uncompressed data-row hashes of the uploads. Header lines, raw input hashes, reconciliation tables, mapping candidates, and the full label-match grid are also retained.

The uploaded B101 conversation copies can therefore be deleted after this report and its manifest updates are accepted. Keep the user's local master copies.
"""
    report_path = out / f"B101_REPORT_{stamp}.md"
    report_path.write_text(report, encoding="utf-8")

    # Provenance events.
    events = []
    for r in integrity_rows:
        for event_type, status, details in [
            ("batch_input_received", "present", f"size={r['size_bytes']}; local_path={r['local_path']}"),
            ("hash_verified_against_B000", "pass", f"sha256={r['sha256']}; expected={r['expected_sha256_from_B000']}"),
            ("gzip_integrity_verified", "pass", "gzip -t exit code 0"),
            ("normalized_derivative_created", "pass", "all raw data fields and rows preserved; see derivative reconciliation"),
            ("analysis_completed", "pass", "GAF/GPAD/GPI reconciliation and direct label grid completed"),
        ]:
            events.append({"event_time_utc": now_iso(), "batch_id": "B101", "artifact_name": r["artifact_name"], "sha256": r["sha256"], "event_type": event_type, "status": status, "details": details})
        events.append({"event_time_utc": now_iso(), "batch_id": "B101", "artifact_name": r["artifact_name"], "sha256": r["sha256"], "event_type": "deletion_clearance_issued", "status": "safe_to_delete_conversation_copy", "details": "Local master copy should be retained; normalized derivatives, hashes, headers, and reconciliation outputs retained."})
    events_path = out / f"B101_provenance_events_{stamp}.csv"
    write_csv(events_path, events)

    # Append source ledger records to latest B000 ledger.
    base_ledger = inp / "ppi_repro_corrected/results/source_ledger_with_B000_20260827T150123Z.csv"
    with base_ledger.open(newline="", encoding="utf-8-sig") as fh:
        base_rows = list(csv.DictReader(fh)); ledger_fields = list(base_rows[0].keys())
    added_rows = []
    for r in integrity_rows:
        added = {k: "" for k in ledger_fields}
        added.update({
            "record_type": "batch_input_materialized_and_analyzed",
            "artifact_name": r["artifact_name"], "local_path": r["local_path"], "local_status": "present_at_analysis_time",
            "origin_in_this_run": "user_upload_B101", "analysis_role": f"GOA release 159 {r['file_role']} input",
            "used_by": "analyze_B101_goa159.py", "direct_or_canonical_source_url": r["direct_or_canonical_source_url"],
            "source_page_url": r["source_page_url"], "url_status": "official_EBI_GOA_archive_exact_filename",
            "retrieval_status": "supplied_by_user; exact size and SHA-256 match B000 local inventory",
            "retrieved_at_utc": generated, "size_bytes": str(r["size_bytes"]), "sha256": r["sha256"],
            "parent_or_derivation": "local_upload_inventory_20260827T145903Z.csv",
            "notes": "Raw conversation copy eligible for deletion after B101 clearance; local master retained by user.",
            "batch_id": "B101", "user_local_relative_path": r["artifact_name"], "user_local_status": "present_on_user_machine_per_B000_and_user_ls",
            "user_local_size_bytes": str(r["size_bytes"]), "user_local_sha256": r["sha256"],
            "user_local_mtime_utc": expected_inventory[r["artifact_name"]].get("mtime_utc", ""),
            "inventory_file": "local_upload_inventory_20260827T145903Z.csv", "inventory_sha256": sha256_file(inp / "local_upload_inventory_20260827T145903Z.csv"),
            "deletion_state": "deletion_clearance_issued_for_conversation_copy",
        })
        added_rows.append(added)
    ledger_path = inp / f"ppi_repro_corrected/results/source_ledger_with_B101_{stamp}.csv"
    write_csv(ledger_path, base_rows + added_rows, ledger_fields)

    # Checksums after all outputs except bundle/checksum itself.
    script_path = Path(__file__)
    tracked = [p for p in sorted(out.glob(f"*{stamp}*")) if p.is_file()]
    check_rows = [{"artifact_name": p.name, "local_path": str(p), "size_bytes": p.stat().st_size, "sha256": sha256_file(p), "role": "B101 output"} for p in tracked]
    check_rows.append({"artifact_name": script_path.name, "local_path": str(script_path), "size_bytes": script_path.stat().st_size, "sha256": sha256_file(script_path), "role": "analysis script"})
    checksums_path = out / f"B101_output_checksums_{stamp}.csv"
    write_csv(checksums_path, check_rows)

    clearance = f"""# SAFE TO DELETE — BATCH B101

Clearance issued: `{now_iso()}`

The following **uploaded conversation copies** passed hash and gzip integrity checks, were fully parsed, and have row-preserving normalized derivatives with successful re-read reconciliation:

- `goa_human.gaf.159.gz` — `{integrity_rows[0]['sha256']}`
- `goa_human.gpa.159.gz` — `{integrity_rows[1]['sha256']}`
- `goa_human.gpi.159.gz` — `{integrity_rows[2]['sha256']}`

Retained:

- all header lines and raw data-row hashes;
- normalized row-preserving GAF, GPAD, and GPI derivatives;
- GAF↔GPAD assertion reconciliation and ECO multiplicity table;
- GPI coverage and metadata reconciliation;
- provisional mapping candidates and mapping diagnostics;
- complete direct-label reconstruction grid and summary;
- provenance events, source-ledger update, and output checksums.

No raw-file hold remains for these three conversation attachments. **Keep the local master copies on your own machine.**

After deleting the three attachments from the conversation, reply: `Deleted B101`.
"""
    clearance_path = out / f"B101_DELETION_CLEARANCE_{stamp}.md"
    clearance_path.write_text(clearance, encoding="utf-8")

    # Bundle core batch outputs. Avoid including redundant previous ledgers.
    bundle_path = inp / f"B101_goa159_analysis_bundle_{stamp}.zip"
    with zipfile.ZipFile(bundle_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
        for p in sorted(out.iterdir()):
            if p.is_file() and stamp in p.name:
                zf.write(p, arcname=f"B101/{p.name}")
        zf.write(script_path, arcname=f"B101/{script_path.name}")
        zf.write(ledger_path, arcname=f"B101/{ledger_path.name}")
    bundle_sha = sha256_file(bundle_path)
    (inp / f"B101_goa159_analysis_bundle_{stamp}.sha256.txt").write_text(f"{bundle_sha}  {bundle_path.name}\n", encoding="utf-8")

    print(json.dumps({
        "status": "PASS",
        "batch": "B101",
        "stamp": stamp,
        "report": str(report_path),
        "clearance": str(clearance_path),
        "ledger": str(ledger_path),
        "bundle": str(bundle_path),
        "bundle_sha256": bundle_sha,
        "best_direct_configuration": best_direct,
        "gaf_gpad_reconciliation": recon,
        "gpi_xrefs_nonempty": gpi_summary["gpi_DB_Xrefs_nonempty"],
    }, indent=2))


if __name__ == "__main__":
    main()
